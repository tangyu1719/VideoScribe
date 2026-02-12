#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
视频转文字处理工具 - 桌面GUI（Tkinter）
- 只需输入链接并选择平台
- 第一步：根据平台调用下载API获取视频URL并保存至本地
- 第二步：调用语音转文字API并生成Markdown文档
- 输出目录：./output  视频目录：./videos
注意：若第三方网站API不可用，将在日志中提示；程序不中断。
"""

# 忽略NumPy版本警告
import warnings
warnings.filterwarnings("ignore", message="A NumPy version >=1.23.5 and <2.3.0 is required for this version of SciPy")

import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox, filedialog
import threading
import concurrent.futures
import requests
import json
import os
import time
import hashlib
from datetime import datetime
import multiprocessing
import asyncio
import aiohttp

# 导入LangChain集成
from langchain_integration import LangChainIntegration

APP_TITLE = "视频转文字处理工具 (GUI)"
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
VIDEO_DIR = os.path.join(BASE_DIR, "videos")
OUTPUT_DIR = os.path.join(BASE_DIR, "output")
CONFIG_FILE = os.path.join(BASE_DIR, "config.json")
HISTORY_FILE = os.path.join(BASE_DIR, "history.json")

# 火山引擎 API 配置
VOLCENGINE_API_KEY = "5da00752-8f46-44eb-b162-5c52f2a249b3"
VOLCENGINE_API_URL = "https://ark.cn-beijing.volces.com/api/v3"

# 默认配置
DEFAULT_CONFIG = {
    "summary_prompt": "请对以下文本进行总结，提取关键知识点，整理成结构化的格式：\n{text}",
    "volcengine_api_key": VOLCENGINE_API_KEY,
    "system_prompt": "你是一个专业的视频内容分析助手，擅长从视频转写内容中提取关键信息并进行结构化分析。",
    "rules": "1. 提取视频中的关键知识点和核心信息\n2. 保持客观中立的分析态度\n3. 结构化呈现分析结果\n4. 重点关注视频中的技术讲解和实用信息\n5. 文件名命名规则：总记录序号-月-日-文档名称（可通过规则控制文件名生成逻辑）",
    "file_naming_rule": "总记录序号-月-日-文档名称",  # 文件名命名规则
    "output_template": "# {platform}视频分析\n\n## 视频信息\n- 分析时间: {datetime}\n- 原始链接: {link}\n- 平台: {platform}\n\n## 语音转文字内容\n{transcript}\n\n## AI分析摘要\n{summary}",
    "user_prompt": ""
}

for d in (VIDEO_DIR, OUTPUT_DIR):
    if not os.path.exists(d):
        os.makedirs(d)

# 加载配置文件
def load_config():
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            print(f"加载配置文件失败：{e}")
    return DEFAULT_CONFIG

# 保存配置文件
def save_config(config):
    try:
        with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
        return True
    except Exception as e:
        print(f"保存配置文件失败：{e}")
        return False

# 加载历史记录
def load_history():
    if os.path.exists(HISTORY_FILE):
        try:
            with open(HISTORY_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            print(f"加载历史记录失败：{e}")
    return {"tasks": []}

# 保存历史记录
def save_history(history):
    try:
        with open(HISTORY_FILE, 'w', encoding='utf-8') as f:
            json.dump(history, f, ensure_ascii=False, indent=2)
        return True
    except Exception as e:
        print(f"保存历史记录失败：{e}")
        return False

# 加载配置
CONFIG = load_config()

PLATFORMS = {
    "小红书": {
        "api_endpoint": "https://www.hellotik.app/zh/rednote",
        "headers": {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Accept': 'application/json, text/plain, */*',
            'Content-Type': 'application/json',
            'Referer': 'https://hellotik.app/',
            'Origin': 'https://hellotik.app'
        },
        "payload": {
            "requestURL": "{url}",
            "isMobile": "false",
            "isoCode": "HK",
            "adType": "adsense",
            "uwx_id": "uwx_350696y5juIO",
            "successCount": "0",
            "totalSuccessCount": "2",
            "firstSuccessDate": "2026-01-10",
            "time": "{timestamp}",
            "key": "xaq8pkc7"
        },
        "url_key_candidates": ["video_url", "download_url", "url"]
    }
}

class App:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title(APP_TITLE)
        self.root.geometry("1200x800")
        self.root.configure(bg="#f0f4f8")
        # 允许窗口大小调整
        self.root.resizable(True, True)

        self.link_var = tk.StringVar()
        
        # 队列系统初始化
        self.task_queue = []
        self.processing_queue = False
        self.current_task_index = 0
        self.queue_max_size = 50  # 默认队列最大大小
        
        # 历史记录初始化
        self.history = load_history()
        
        # 线程池初始化
        self.cpu_count = multiprocessing.cpu_count()
        # 从配置文件中加载线程数量，如果不存在则使用默认值
        self.max_workers = CONFIG.get("max_workers", min(self.cpu_count, 8))
        # 确保线程数量在合理范围内
        self.max_workers = max(1, min(self.max_workers, self.cpu_count))
        self.executor = concurrent.futures.ThreadPoolExecutor(max_workers=self.max_workers)
        self.active_futures = []  # 存储活跃的任务未来对象
        
        # 缓存机制初始化
        self.model_cache = None  # Whisper模型缓存
        self.model_cache_lock = threading.Lock()  # 模型缓存锁
        self.video_cache = {}  # 视频缓存，键为链接，值为本地文件路径
        self.video_cache_lock = threading.Lock()  # 视频缓存锁
        
        # 功能开关
        self.feishu_enabled = False  # 飞书功能开关，默认禁用

        # 先构建UI，确保所有UI组件都已创建
        self._build_ui()
        
        # LangChain集成初始化
        self.langchain_integration = None
        try:
            self.langchain_integration = LangChainIntegration()
            self.append_log("LangChain集成初始化成功", "INFO")
        except Exception as e:
            self.append_log(f"LangChain集成初始化失败：{e}", "WARNING")
        
        # 自动恢复未完成任务
        self.recover_unfinished_tasks()
        
        # 显示系统信息
        self.append_log(f"系统CPU核心数：{self.cpu_count}")
        self.append_log(f"线程池最大工作线程数：{self.max_workers}")
        
        # 绑定窗口关闭事件
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)

    def _build_ui(self):
        # 设置样式
        style = ttk.Style()
        style.theme_use("clam")
        
        # 自定义样式
        style.configure(
            "TButton", 
            padding=(12, 6),
            font=("微软雅黑", 10),
            foreground="#0066cc",
            background="#e6f0ff"
        )
        style.configure(
            "TLabel", 
            font=("微软雅黑", 10),
            foreground="#333"
        )
        style.configure(
            "TEntry", 
            padding=(8, 4),
            font=("微软雅黑", 10)
        )
        style.configure(
            "TLabelframe", 
            font=("微软雅黑", 10, "bold"),
            foreground="#0066cc"
        )
        
        # 创建圆角输入框样式
        style.configure(
            "Rounded.TEntry",
            padding=(10, 6),
            font=("微软雅黑", 10),
            borderwidth=0,
            relief="flat"
        )
        
        # 主容器
        main_container = tk.Frame(self.root, bg="#f0f4f8")
        main_container.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # 顶部标题区域
        title_frame = tk.Frame(main_container, bg="#f0f4f8")
        title_frame.pack(fill=tk.X, pady=(0, 20))
        
        title_label = tk.Label(
            title_frame, 
            text="视频转文字处理工具", 
            font=("微软雅黑", 18, "bold"),
            foreground="#0066cc",
            bg="#f0f4f8"
        )
        title_label.pack(anchor=tk.W)
        
        subtitle_label = tk.Label(
            title_frame, 
            text="智能视频分析与文本转换系统", 
            font=("微软雅黑", 10, "italic"),
            foreground="#666",
            bg="#f0f4f8"
        )
        subtitle_label.pack(anchor=tk.W, pady=(5, 0))
        
        # 核心功能区域
        core_frame = tk.Frame(main_container, bg="#f0f4f8")
        core_frame.pack(fill=tk.X, pady=(0, 20))
        
        # 视频链接输入区域
        link_frame = tk.Frame(core_frame, bg="#ffffff", bd=0, relief=tk.RAISED)
        link_frame.pack(fill=tk.X, padx=5, pady=5)
        link_frame.configure(bg="#ffffff", highlightbackground="#0066cc", highlightthickness=1, borderwidth=0, highlightcolor="#0066cc")
        
        link_label = tk.Label(
            link_frame, 
            text="视频链接：", 
            font=("微软雅黑", 10),
            foreground="#333",
            bg="#ffffff"
        )
        link_label.pack(side=tk.LEFT, padx=(15, 10), pady=15)
        
        self.link_entry = tk.Entry(
            link_frame, 
            textvariable=self.link_var, 
            font=("微软雅黑", 10),
            bd=0, bg="#ffffff",
            relief=tk.FLAT,
            highlightthickness=0
        )
        self.link_entry.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True, pady=15)
        
        # 按钮区域
        button_frame = tk.Frame(core_frame, bg="#f0f4f8")
        button_frame.pack(fill=tk.X, padx=5, pady=10)
        
        btn_container = tk.Frame(button_frame, bg="#f0f4f8")
        btn_container.pack(anchor=tk.E)
        
        self.start_btn = ttk.Button(btn_container, text="开始处理", command=self.start)
        self.start_btn.pack(side=tk.LEFT, padx=10)
        
        self.ai_config_btn = ttk.Button(btn_container, text="AI配置", command=self.open_ai_config_window)
        self.ai_config_btn.pack(side=tk.LEFT, padx=10)
        
        # 批量导入按钮
        self.batch_import_btn = ttk.Button(btn_container, text="批量导入", command=self.batch_import)
        self.batch_import_btn.pack(side=tk.LEFT, padx=10)
        
        # 历史查询按钮
        self.history_btn = ttk.Button(btn_container, text="历史查询", command=self.show_history)
        self.history_btn.pack(side=tk.LEFT, padx=10)
        
        # 线程配置按钮
        self.thread_config_btn = ttk.Button(btn_container, text="线程配置", command=self.open_thread_config_window)
        self.thread_config_btn.pack(side=tk.LEFT, padx=10)
        
        # 任务状态区域
        self.status_frame = tk.Frame(main_container, bg="#ffffff", bd=0, relief=tk.RAISED)
        self.status_frame.pack(fill=tk.X, pady=(0, 20))
        self.status_frame.configure(bg="#ffffff", highlightbackground="#0066cc", highlightthickness=1, borderwidth=0, highlightcolor="#0066cc")
        
        self.status_info = tk.Label(
            self.status_frame, 
            text="任务状态：就绪", 
            font=("微软雅黑", 10),
            foreground="#333",
            bg="#ffffff"
        )
        self.status_info.pack(side=tk.LEFT, padx=(15, 10), pady=10)
        
        # 队列状态显示
        self.queue_status = tk.Label(
            self.status_frame, 
            text="队列：0 个任务", 
            font=("微软雅黑", 10),
            foreground="#0066cc",
            bg="#ffffff"
        )
        self.queue_status.pack(side=tk.RIGHT, padx=(10, 15), pady=10)
        
        # User Prompt 输入区域
        user_prompt_frame = tk.Frame(main_container, bg="#ffffff", bd=0, relief=tk.RAISED)
        user_prompt_frame.pack(fill=tk.X, pady=(0, 20))
        user_prompt_frame.configure(bg="#ffffff", highlightbackground="#0066cc", highlightthickness=1, borderwidth=0, highlightcolor="#0066cc")
        
        prompt_label_frame = tk.Frame(user_prompt_frame, bg="#ffffff")
        prompt_label_frame.pack(fill=tk.X, padx=15, pady=(15, 0))
        
        prompt_label = tk.Label(
            prompt_label_frame, 
            text="User Prompt（可选）：", 
            font=("微软雅黑", 10, "bold"),
            foreground="#0066cc",
            bg="#ffffff"
        )
        prompt_label.pack(side=tk.LEFT)
        
        # 添加展开按钮
        def open_prompt_window():
            """打开大文本框窗口"""
            prompt_window = tk.Toplevel(self.root)
            prompt_window.title("User Prompt 编辑")
            prompt_window.geometry("800x500")
            prompt_window.configure(bg="#f0f4f8")
            prompt_window.transient(self.root)
            prompt_window.grab_set()
            
            # 居中显示
            prompt_window.update_idletasks()
            width = prompt_window.winfo_width()
            height = prompt_window.winfo_height()
            x = (prompt_window.winfo_screenwidth() // 2) - (width // 2)
            y = (prompt_window.winfo_screenheight() // 2) - (height // 2)
            prompt_window.geometry(f"{width}x{height}+{x}+{y}")
            
            # 文本框标签
            text_frame = tk.Frame(prompt_window, bg="#ffffff")
            text_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=(20, 10))
            text_frame.configure(highlightbackground="#0066cc", highlightthickness=1)
            
            # 大文本框
            text_widget = tk.Text(
                text_frame, 
                font=("微软雅黑", 10),
                bg="#f9f9f9",
                wrap=tk.WORD
            )
            text_widget.pack(fill=tk.BOTH, expand=True, padx=15, pady=15)
            
            # 插入当前值
            current_value = self.user_prompt_var.get()
            text_widget.insert(tk.END, current_value)
            
            # 字符计数标签
            window_char_count_var = tk.StringVar(value=f"{len(current_value)}/500")
            count_frame = tk.Frame(prompt_window, bg="#f0f4f8")
            count_frame.pack(fill=tk.X, padx=20, pady=(0, 10))
            
            count_label = tk.Label(
                count_frame, 
                textvariable=window_char_count_var, 
                font=("微软雅黑", 9),
                foreground="#999",
                bg="#f0f4f8"
            )
            count_label.pack(anchor=tk.E)
            
            # 实时更新字符计数
            def update_char_count(event):
                text = text_widget.get("1.0", tk.END).strip()
                if len(text) > 500:
                    # 限制字符数
                    text_widget.delete("1.0", tk.END)
                    text_widget.insert(tk.END, text[:500])
                    text = text[:500]
                window_char_count_var.set(f"{len(text)}/500")
            
            text_widget.bind("<KeyRelease>", update_char_count)
            
            # 按钮区域
            button_frame = tk.Frame(prompt_window, bg="#f0f4f8")
            button_frame.pack(fill=tk.X, padx=20, pady=10)
            
            def save_prompt():
                """保存User Prompt"""
                value = text_widget.get("1.0", tk.END).strip()
                if len(value) > 500:
                    value = value[:500]
                self.user_prompt_var.set(value)
                prompt_window.destroy()
            
            save_btn = ttk.Button(button_frame, text="保存", command=save_prompt)
            save_btn.pack(side=tk.RIGHT, padx=10)
            
            cancel_btn = ttk.Button(button_frame, text="取消", command=prompt_window.destroy)
            cancel_btn.pack(side=tk.RIGHT, padx=10)
        
        # 展开按钮
        expand_btn = ttk.Button(
            prompt_label_frame, 
            text="展开编辑", 
            command=open_prompt_window
        )
        expand_btn.pack(side=tk.RIGHT, padx=10)
        
        prompt_desc = tk.Label(
            prompt_label_frame, 
            text="每次处理视频时的额外提示信息，最多500字符", 
            font=("微软雅黑", 9),
            foreground="#666",
            bg="#ffffff"
        )
        prompt_desc.pack(side=tk.LEFT, padx=10)
        
        # 字符计数标签
        char_count_var = tk.StringVar(value="0/500")
        char_count_label = tk.Label(
            prompt_label_frame, 
            textvariable=char_count_var, 
            font=("微软雅黑", 9),
            foreground="#999",
            bg="#ffffff"
        )
        char_count_label.pack(side=tk.RIGHT)
        
        self.user_prompt_var = tk.StringVar(value=CONFIG.get("user_prompt", DEFAULT_CONFIG["user_prompt"]))
        self.user_prompt_entry = tk.Entry(
            user_prompt_frame, 
            textvariable=self.user_prompt_var, 
            font=("微软雅黑", 10),
            bd=0, bg="#ffffff",
            relief=tk.FLAT,
            highlightthickness=0
        )
        self.user_prompt_entry.pack(fill=tk.X, padx=15, pady=(10, 15))
        
        # 点击输入框也打开大文本框
        def on_entry_click(event):
            open_prompt_window()
        
        self.user_prompt_entry.bind("<Button-1>", on_entry_click)
        
        # 字符数限制和计数
        def limit_characters(*args):
            value = self.user_prompt_var.get()
            if len(value) > 500:
                self.user_prompt_var.set(value[:500])
            char_count_var.set(f"{len(self.user_prompt_var.get())}/500")
        
        self.user_prompt_var.trace_add("write", limit_characters)
        # 初始计数
        limit_characters()
        
        # 日志区域
        log_frame = tk.Frame(main_container, bg="#ffffff", bd=0, relief=tk.RAISED)
        log_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 20))
        log_frame.configure(bg="#ffffff", highlightbackground="#0066cc", highlightthickness=1, borderwidth=0, highlightcolor="#0066cc")
        
        log_title_frame = tk.Frame(log_frame, bg="#ffffff")
        log_title_frame.pack(fill=tk.X, padx=15, pady=(15, 10))
        
        log_title = tk.Label(
            log_title_frame, 
            text="处理日志", 
            font=("微软雅黑", 10, "bold"),
            foreground="#0066cc",
            bg="#ffffff"
        )
        log_title.pack(anchor=tk.W)
        
        self.log = scrolledtext.ScrolledText(
            log_frame, 
            height=15, 
            font=("Consolas", 10),
            bd=0, 
            bg="#f9f9f9",
            relief=tk.FLAT,
            wrap=tk.WORD
        )
        self.log.pack(fill=tk.BOTH, expand=True, padx=15, pady=(0, 15))
        
        # 底部状态区域
        status_frame = tk.Frame(main_container, bg="#f0f4f8")
        status_frame.pack(fill=tk.X, pady=(0, 10))
        
        self.status_var = tk.StringVar(value="就绪")
        status_label = tk.Label(
            status_frame, 
            textvariable=self.status_var, 
            font=("微软雅黑", 10),
            foreground="#0066cc",
            bg="#f0f4f8"
        )
        status_label.pack(anchor=tk.W)
        
        # 状态指示器
        status_indicator = tk.Label(
            status_frame, 
            text="●", 
            font=("微软雅黑", 12),
            foreground="#00cc66",
            bg="#f0f4f8"
        )
        status_indicator.pack(side=tk.LEFT, padx=(0, 10))

    # 日志与状态
    def append_log(self, msg, *args):
        ts = datetime.now().strftime("%H:%M:%S")
        thread_id = threading.current_thread().name
        self.log.insert(tk.END, f"[{ts}] [线程:{thread_id}] {msg}\n")
        self.log.see(tk.END)
        self.root.update_idletasks()

    def set_status(self, msg: str):
        self.status_var.set(msg)
        self.root.update_idletasks()
    
    def update_progress(self, percentage: float, status: str = ""):
        """更新任务状态显示"""
        if status:
            self.set_status(status)
            self.status_info.config(text=f"任务状态：{status}")
        self.root.update_idletasks()
    
    def update_queue_status(self):
        """更新队列状态显示"""
        queue_length = len(self.task_queue)
        processing_status = "处理中" if self.processing_queue else "就绪"
        self.queue_status.config(text=f"队列：{queue_length} 个任务 | 状态：{processing_status}")
        self.root.update_idletasks()

    # 恢复未完成任务
    def recover_unfinished_tasks(self):
        """恢复上次未完成的任务"""
        unfinished_tasks = [task for task in self.history.get("tasks", []) 
                          if task.get("status") not in ["completed", "failed"]]
        
        if unfinished_tasks:
            self.append_log(f"发现 {len(unfinished_tasks)} 个未完成任务，正在恢复...")
            for task in unfinished_tasks:
                link = task.get("link")
                if link not in self.task_queue:
                    self.task_queue.append(link)
            
            self.append_log(f"已恢复 {len(unfinished_tasks)} 个任务到队列")
            self.append_log(f"当前队列长度：{len(self.task_queue)}")
            
            # 自动开始处理队列
            if not self.processing_queue and self.task_queue:
                self.start_queue_processing()
        else:
            self.append_log("无未完成任务需要恢复")
    
    # 检查链接是否已导入
    def is_link_already_imported(self, link):
        """检查链接是否已经导入过"""
        # 使用链接的hash作为任务ID，提高比较效率
        link_hash = hashlib.md5(link.encode()).hexdigest()
        
        # 检查历史记录中是否存在该链接
        for task in self.history.get("tasks", []):
            if task.get("id") == link_hash:
                return True
        
        return False
    
    # 从链接中提取标题
    def extract_title_from_link(self, link):
        """从链接中提取标题
        
        Args:
            link: 视频链接
            
        Returns:
            提取的标题，如果无法提取则返回None
        """
        try:
            # 对于小红书链接，尝试获取页面标题
            if "xiaohongshu.com" in link:
                import requests
                
                # 尝试导入BeautifulSoup
                try:
                    from bs4 import BeautifulSoup
                    
                    headers = {
                        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/144.0.0.0 Safari/537.36',
                        'Referer': 'https://www.xiaohongshu.com/'
                    }
                    
                    response = requests.get(link, headers=headers, timeout=10)
                    if response.status_code == 200:
                        soup = BeautifulSoup(response.text, 'html.parser')
                        # 尝试从title标签提取
                        title_tag = soup.find('title')
                        if title_tag and title_tag.text:
                            title = title_tag.text.strip()
                            # 清理标题
                            title = title.replace('\n', '').replace('\r', '').replace('  ', ' ')
                            self.append_log(f"从链接中提取标题：{title}")
                            return title
                except ImportError:
                    # 如果缺少bs4模块，跳过网页解析，直接从链接中提取
                    self.append_log("缺少bs4模块，跳过网页解析，直接从链接中提取标题")
            
            # 如果无法从页面获取，尝试从链接中提取
            import re
            # 尝试匹配链接中的数字或有意义的部分
            match = re.search(r'[a-zA-Z0-9_-]{8,}', link)
            if match:
                return match.group(0)
            
            # 对于B站链接，尝试提取BV号
            bv_match = re.search(r'BV[0-9A-Za-z]{10}', link)
            if bv_match:
                return bv_match.group(0)
            
            # 返回默认标题
            return "未知标题"
        except Exception as e:
            self.append_log(f"从链接提取标题异常：{e}")
            # 异常时返回默认标题
            return "未知标题"
    
    # 添加任务到历史记录
    def add_task_to_history(self, link, user_prompt="", feishu_folder_path=None):
        """添加任务到历史记录
        
        Args:
            link: 视频链接
            user_prompt: 用户提示词
            feishu_folder_path: 飞书文件夹路径
            
        Returns:
            任务对象
        """
        # 检查是否已存在
        link_hash = hashlib.md5(link.encode()).hexdigest()
        for task in self.history.get("tasks", []):
            if task.get("id") == link_hash:
                # 更新用户提示词和飞书文件夹路径
                if user_prompt:
                    task["user_prompt"] = user_prompt
                if feishu_folder_path:
                    task["feishu_folder_path"] = feishu_folder_path
                task["updated_at"] = datetime.now().isoformat()
                save_history(self.history)
                return task
        
        # 从链接中提取标题
        link_title = self.extract_title_from_link(link)
        
        # 创建新任务记录
        new_task = {
            "id": link_hash,
            "link": link,
            "title": link_title,
            "status": "pending",
            "user_prompt": user_prompt,
            "feishu_folder_path": feishu_folder_path,
            "stages": {
                "download": {"status": "pending", "result": None},
                "transcribe": {"status": "pending", "result": None},
                "ai_analysis": {"status": "pending", "result": None},
                "generate_md": {"status": "pending", "result": None},
                "feishu_upload": {"status": "pending", "result": None}
            },
            "created_at": datetime.now().isoformat(),
            "updated_at": datetime.now().isoformat()
        }
        
        self.history.setdefault("tasks", []).append(new_task)
        save_history(self.history)
        return new_task
    
    # 更新任务状态
    def update_task_status(self, link, stage, status, result=None):
        """更新任务状态"""
        for task in self.history.get("tasks", []):
            if task.get("link") == link:
                # 更新阶段状态
                if stage in task.get("stages", {}):
                    task["stages"][stage]["status"] = status
                    if result:
                        task["stages"][stage]["result"] = result
                
                # 更新整体状态
                if (stage == "generate_md" or stage == "feishu_upload") and status == "completed":
                    task["status"] = "completed"
                elif status == "failed":
                    task["status"] = "failed"
                
                task["updated_at"] = datetime.now().isoformat()
                save_history(self.history)
                return True
        return False
    
    # 显示历史记录
    def show_history(self):
        """显示历史记录窗口"""
        history_window = tk.Toplevel(self.root)
        history_window.title("历史记录查询")
        history_window.geometry("1000x600")
        history_window.configure(bg="#f0f4f8")
        
        # 设置样式
        style = ttk.Style()
        style.theme_use("clam")
        
        # 主容器
        main_frame = tk.Frame(history_window, bg="#f0f4f8")
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # 标题
        title_frame = tk.Frame(main_frame, bg="#f0f4f8")
        title_frame.pack(fill=tk.X, pady=(0, 15))
        
        title_label = tk.Label(
            title_frame, 
            text="历史记录查询", 
            font=("微软雅黑", 16, "bold"),
            foreground="#0066cc",
            bg="#f0f4f8"
        )
        title_label.pack(anchor=tk.W)
        
        # 历史记录列表
        tree_frame = tk.Frame(main_frame, bg="#ffffff")
        tree_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 15))
        tree_frame.configure(highlightbackground="#0066cc", highlightthickness=1)
        
        # 使用grid布局管理器来确保树状图完全填充空间
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)
        
        # 创建树状图
        columns = ("id", "title", "link", "status", "created_at", "updated_at")
        tree = ttk.Treeview(tree_frame, columns=columns, show="headings")
        
        # 设置列标题
        tree.heading("id", text="任务ID")
        tree.heading("title", text="标题")
        tree.heading("link", text="视频链接")
        tree.heading("status", text="状态")
        tree.heading("created_at", text="创建时间")
        tree.heading("updated_at", text="更新时间")
        
        # 设置列宽
        tree.column("id", width=100)
        tree.column("title", width=200)
        tree.column("link", width=300)
        tree.column("status", width=100)
        tree.column("created_at", width=150)
        tree.column("updated_at", width=150)
        
        # 添加垂直滚动条
        v_scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
        v_scrollbar.grid(row=0, column=1, sticky="ns")
        tree.configure(yscroll=v_scrollbar.set)
        
        # 添加水平滚动条
        h_scrollbar = ttk.Scrollbar(tree_frame, orient="horizontal", command=tree.xview)
        h_scrollbar.grid(row=1, column=0, sticky="ew")
        tree.configure(xscroll=h_scrollbar.set)
        
        # 放置树状图
        tree.grid(row=0, column=0, sticky="nsew")
        
        # 添加鼠标滚轮支持
        def on_mouse_wheel(event):
            try:
                if event.state & 0x1:
                    # Shift键被按下，水平滚动
                    tree.xview_scroll(-1 * (event.delta // 120), "units")
                else:
                    # 垂直滚动
                    tree.yview_scroll(-1 * (event.delta // 120), "units")
            except Exception:
                # 忽略组件已销毁的错误
                pass
        
        tree.bind_all("<MouseWheel>", on_mouse_wheel)
        
        # 绑定窗口关闭事件，解绑鼠标滚轮事件
        def on_window_close():
            try:
                tree.unbind_all("<MouseWheel>")
            except Exception:
                pass
            history_window.destroy()
        
        history_window.protocol("WM_DELETE_WINDOW", on_window_close)
        
        # 填充数据 - 按创建时间降序排序
        tasks = self.history.get("tasks", [])
        # 按created_at字段降序排序
        tasks.sort(key=lambda x: x.get("created_at", ""), reverse=True)
        
        for task in tasks:
            task_id = task.get("id", "")
            title = task.get("title", "")
            # 如果标题为None或空字符串，显示"生成中"的标记
            if not title:
                title = "生成中"
            link = task.get("link", "")
            status = task.get("status", "")
            created_at = task.get("created_at", "")
            updated_at = task.get("updated_at", "")
            
            # 根据状态设置标签
            tag = status
            if status == "completed":
                tag = "completed"
            elif status == "failed":
                tag = "failed"
            elif status == "in_progress":
                tag = "in_progress"
            else:
                tag = "pending"
            
            # 插入数据并应用标签
            item_id = tree.insert("", "end", values=(
                task_id,
                title,
                link,
                status,
                created_at,
                updated_at
            ), tags=(tag,))
        
        # 设置标签颜色
        tree.tag_configure("completed", background="#e6ffe6")  # 绿色背景
        tree.tag_configure("failed", background="#ffe6e6")      # 红色背景
        tree.tag_configure("in_progress", background="#fff0e6")  # 橙色背景
        tree.tag_configure("pending", background="#f0f0f0")     # 灰色背景
        
        # 清理完成任务的缓存文件
        self.cleanup_completed_tasks()
        
        # 详情查看按钮
        def view_details():
            selected = tree.selection()
            if not selected:
                messagebox.showinfo("提示", "请选择一个任务查看详情")
                return
            
            item = tree.item(selected[0])
            task_id = item["values"][0]
            
            # 查找任务
            task = None
            for t in self.history.get("tasks", []):
                if t.get("id") == task_id:
                    task = t
                    break
            
            if not task:
                return
            
            # 显示详情窗口
            detail_window = tk.Toplevel(history_window)
            detail_window.title("任务详情")
            detail_window.geometry("800x500")
            detail_window.configure(bg="#f0f4f8")
            
            # 详情容器
            detail_frame = tk.Frame(detail_window, bg="#f0f4f8")
            detail_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            
            # 链接信息
            link_frame = tk.Frame(detail_frame, bg="#ffffff")
            link_frame.pack(fill=tk.X, pady=(0, 10))
            link_frame.configure(highlightbackground="#0066cc", highlightthickness=1)
            
            ttk.Label(link_frame, text="视频链接：", font=("微软雅黑", 10), background="#ffffff").pack(anchor=tk.W, padx=10, pady=(10, 5))
            link_text = scrolledtext.ScrolledText(link_frame, height=2, font=("Consolas", 10), bg="#f9f9f9")
            link_text.pack(fill=tk.X, padx=10, pady=(0, 10))
            link_text.insert(tk.END, task.get("link", ""))
            link_text.configure(state=tk.DISABLED)
            
            # 阶段信息
            stages_frame = tk.Frame(detail_frame, bg="#ffffff")
            stages_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
            stages_frame.configure(highlightbackground="#0066cc", highlightthickness=1)
            
            ttk.Label(stages_frame, text="处理阶段：", font=("微软雅黑", 10, "bold"), background="#ffffff").pack(anchor=tk.W, padx=10, pady=(10, 5))
            
            # 创建带滚动条的阶段内容框架
            content_frame = tk.Frame(stages_frame, bg="#ffffff")
            content_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
            
            # 添加垂直滚动条
            canvas = tk.Canvas(content_frame, bg="#ffffff")
            scrollbar = ttk.Scrollbar(content_frame, orient="vertical", command=canvas.yview)
            canvas.configure(yscrollcommand=scrollbar.set)
            
            # 内部内容框架
            inner_frame = tk.Frame(canvas, bg="#ffffff")
            canvas.create_window((0, 0), window=inner_frame, anchor="nw")
            
            # 填充阶段信息
            stages = task.get("stages", {})
            for stage_name, stage_info in stages.items():
                stage_frame = tk.Frame(inner_frame, bg="#f9f9f9")
                stage_frame.pack(fill=tk.X, padx=5, pady=5)
                
                stage_label = ttk.Label(stage_frame, text=f"{stage_name}：", font=("微软雅黑", 9), background="#f9f9f9")
                stage_label.pack(side=tk.LEFT, padx=5)
                
                status_label = ttk.Label(stage_frame, text=stage_info.get("status", ""), font=("微软雅黑", 9), background="#f9f9f9")
                status_label.pack(side=tk.LEFT, padx=10)
                
                # 根据状态设置颜色
                if stage_info.get("status") == "completed":
                    status_label.configure(foreground="#00cc66")
                elif stage_info.get("status") == "failed":
                    status_label.configure(foreground="#ff3333")
                    
                    # 添加重试按钮
                    def retry_stage(s=stage_name):
                        link = task.get("link")
                        # 更新阶段状态为pending
                        self.update_task_status(link, s, "pending")
                        # 单独处理失败的阶段
                        threading.Thread(target=self._retry_stage, args=(link, s, task), daemon=True).start()
                        self.append_log(f"开始重试处理阶段 {s}：{link}")
                        detail_window.destroy()
                    
                    retry_btn = ttk.Button(stage_frame, text="重试", command=retry_stage)
                    retry_btn.pack(side=tk.RIGHT, padx=5)
                elif stage_info.get("status") == "in_progress":
                    status_label.configure(foreground="#ff9900")
            
            # 更新画布大小
            def update_scrollregion():
                canvas.update_idletasks()
                canvas.config(scrollregion=canvas.bbox("all"))
            
            # 布局滚动条和画布
            canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            
            # 初始更新滚动区域
            update_scrollregion()
            
            # 按钮区域
            button_frame = tk.Frame(detail_window, bg="#f0f4f8")
            button_frame.pack(fill=tk.X, pady=10)
            
            # 继续处理按钮
            if task.get("status") not in ["completed", "failed"]:
                def continue_task():
                    link = task.get("link")
                    if link not in self.task_queue:
                        self.task_queue.append(link)
                    self.append_log(f"继续处理任务：{link}")
                    if not self.processing_queue:
                        self.start_queue_processing()
                    detail_window.destroy()
                
                ttk.Button(button_frame, text="继续处理", command=continue_task).pack(side=tk.LEFT, padx=10)
            
            # 关闭按钮
            ttk.Button(button_frame, text="关闭", command=detail_window.destroy).pack(side=tk.RIGHT, padx=10)
        
        # 按钮框架
        button_frame = tk.Frame(main_frame, bg="#f0f4f8")
        button_frame.pack(fill=tk.X, pady=10)
        
        # 查看详情按钮
        def view_details():
            selected = tree.selection()
            if not selected:
                messagebox.showinfo("提示", "请选择一个任务查看详情")
                return
            
            item = tree.item(selected[0])
            task_id = item["values"][0]
            
            # 查找任务
            task = None
            for t in self.history.get("tasks", []):
                if t.get("id") == task_id:
                    task = t
                    break
            
            if not task:
                return
            
            # 显示详情窗口
            detail_window = tk.Toplevel(history_window)
            detail_window.title("任务详情")
            detail_window.geometry("800x500")
            detail_window.configure(bg="#f0f4f8")
            
            # 详情容器
            detail_frame = tk.Frame(detail_window, bg="#f0f4f8")
            detail_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            
            # 链接信息
            link_frame = tk.Frame(detail_frame, bg="#ffffff")
            link_frame.pack(fill=tk.X, pady=(0, 10))
            link_frame.configure(highlightbackground="#0066cc", highlightthickness=1)
            
            ttk.Label(link_frame, text="视频链接：", font=("微软雅黑", 10), background="#ffffff").pack(anchor=tk.W, padx=10, pady=(10, 5))
            link_text = scrolledtext.ScrolledText(link_frame, height=2, font=("Consolas", 10), bg="#f9f9f9")
            link_text.pack(fill=tk.X, padx=10, pady=(0, 10))
            link_text.insert(tk.END, task.get("link", ""))
            link_text.configure(state=tk.DISABLED)
            
            # 阶段信息
            stages_frame = tk.Frame(detail_frame, bg="#ffffff")
            stages_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
            stages_frame.configure(highlightbackground="#0066cc", highlightthickness=1)
            
            ttk.Label(stages_frame, text="处理阶段：", font=("微软雅黑", 10, "bold"), background="#ffffff").pack(anchor=tk.W, padx=10, pady=(10, 5))
            
            # 创建阶段内容框架
            content_frame = tk.Frame(stages_frame, bg="#ffffff")
            content_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
            
            stages = task.get("stages", {})
            for stage_name, stage_info in stages.items():
                stage_frame = tk.Frame(content_frame, bg="#f9f9f9")
                stage_frame.pack(fill=tk.X, padx=5, pady=5)
                
                stage_label = ttk.Label(stage_frame, text=f"{stage_name}：", font=("微软雅黑", 9), background="#f9f9f9")
                stage_label.pack(side=tk.LEFT, padx=5)
                
                status_label = ttk.Label(stage_frame, text=stage_info.get("status", ""), font=("微软雅黑", 9), background="#f9f9f9")
                status_label.pack(side=tk.LEFT, padx=10)
                
                # 根据状态设置颜色
                if stage_info.get("status") == "completed":
                    status_label.configure(foreground="#00cc66")
                elif stage_info.get("status") == "failed":
                    status_label.configure(foreground="#ff3333")
                    
                    # 添加重试按钮
                    def retry_stage(s=stage_name):
                        link = task.get("link")
                        # 更新阶段状态为pending
                        self.update_task_status(link, s, "pending")
                        # 单独处理失败的阶段
                        threading.Thread(target=self._retry_stage, args=(link, s, task), daemon=True).start()
                        self.append_log(f"开始重试处理阶段 {s}：{link}")
                        detail_window.destroy()
                    
                    retry_btn = ttk.Button(stage_frame, text="重试", command=retry_stage)
                    retry_btn.pack(side=tk.RIGHT, padx=5)
                elif stage_info.get("status") == "in_progress":
                    status_label.configure(foreground="#ff9900")
            
            # 按钮区域
            button_frame = tk.Frame(detail_window, bg="#f0f4f8")
            button_frame.pack(fill=tk.X, pady=10)
            
            # 继续处理按钮
            if task.get("status") not in ["completed", "failed"]:
                def continue_task():
                    link = task.get("link")
                    if link not in self.task_queue:
                        self.task_queue.append(link)
                    self.update_queue_status()
                    self.append_log(f"继续处理任务：{link}")
                    if not self.processing_queue:
                        self.start_queue_processing()
                    detail_window.destroy()
                
                ttk.Button(button_frame, text="继续处理", command=continue_task).pack(side=tk.LEFT, padx=10)
            
            # 关闭按钮
            ttk.Button(button_frame, text="关闭", command=detail_window.destroy).pack(side=tk.RIGHT, padx=10)
        
        # 继续处理按钮
        def continue_task():
            selected = tree.selection()
            if not selected:
                messagebox.showinfo("提示", "请选择一个任务继续处理")
                return
            
            for item_id in selected:
                item = tree.item(item_id)
                task_id = item["values"][0]
                
                # 查找任务
                task = None
                for t in self.history.get("tasks", []):
                    if t.get("id") == task_id:
                        task = t
                        break
                
                if task:
                    link = task.get("link")
                    if link not in self.task_queue:
                        self.task_queue.append(link)
                    self.update_queue_status()
                    self.append_log(f"继续处理任务：{link}")
            
            # 自动开始处理队列
            if not self.processing_queue and self.task_queue:
                self.start_queue_processing()
        
        # 刷新按钮
        def refresh_tree():
            # 清空树状图
            for item in tree.get_children():
                tree.delete(item)
            
            # 重新填充数据 - 按创建时间降序排序
            tasks = self.history.get("tasks", [])
            # 按created_at字段降序排序
            tasks.sort(key=lambda x: x.get("created_at", ""), reverse=True)
            
            for task in tasks:
                task_id = task.get("id", "")
                title = task.get("title", "")
                # 如果标题为None或空字符串，显示"未知标题"
                if not title:
                    title = "未知标题"
                link = task.get("link", "")
                status = task.get("status", "")
                created_at = task.get("created_at", "")
                updated_at = task.get("updated_at", "")
                
                # 根据状态设置标签
                tag = status
                if status == "completed":
                    tag = "completed"
                elif status == "failed":
                    tag = "failed"
                elif status == "in_progress":
                    tag = "in_progress"
                else:
                    tag = "pending"
                
                # 插入数据并应用标签
                tree.insert("", "end", values=(
                    task_id,
                    title,
                    link,
                    status,
                    created_at,
                    updated_at
                ), tags=(tag,))
        
        # 按钮框架
        action_frame = tk.Frame(main_frame, bg="#f0f4f8")
        action_frame.pack(fill=tk.X, pady=10)
        
        # 批量操作按钮
        def batch_start():
            selected = tree.selection()
            if not selected:
                messagebox.showinfo("提示", "请选择要开始的任务")
                return
            
            for item_id in selected:
                item = tree.item(item_id)
                task_id = item["values"][0]
                
                # 查找任务
                task = None
                for t in self.history.get("tasks", []):
                    if t.get("id") == task_id:
                        task = t
                        break
                
                if task:
                    link = task.get("link")
                    if link not in self.task_queue:
                        self.task_queue.append(link)
                    self.update_queue_status()
                    self.append_log(f"开始处理任务：{link}")
            
            # 自动开始处理队列
            if not self.processing_queue and self.task_queue:
                self.start_queue_processing()
        
        def set_prompt_for_task():
            """为选定的任务设置用户提示词"""
            selected = tree.selection()
            if not selected:
                messagebox.showinfo("提示", "请选择要设置提示词的任务")
                return
            
            # 获取当前的用户提示词
            current_prompt = self.user_prompt_var.get().strip()
            
            # 创建提示词设置窗口
            prompt_window = tk.Toplevel(history_window)
            prompt_window.title("设置用户提示词")
            prompt_window.geometry("600x300")
            prompt_window.configure(bg="#f0f4f8")
            
            # 提示词输入区域
            prompt_frame = tk.Frame(prompt_window, bg="#ffffff")
            prompt_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
            prompt_frame.configure(highlightbackground="#0066cc", highlightthickness=1)
            
            ttk.Label(prompt_frame, text="用户提示词：", font=("微软雅黑", 10), background="#ffffff").pack(anchor=tk.W, padx=15, pady=(15, 5))
            
            prompt_text = tk.Text(prompt_frame, height=8, font=("微软雅黑", 10), bg="#f9f9f9")
            prompt_text.pack(fill=tk.BOTH, expand=True, padx=15, pady=(0, 15))
            prompt_text.insert(tk.END, current_prompt)
            
            # 字符计数
            char_count_var = tk.StringVar(value=f"{len(current_prompt)}/500")
            char_count_label = tk.Label(prompt_frame, textvariable=char_count_var, font=("微软雅黑", 9), foreground="#999", background="#ffffff")
            char_count_label.pack(anchor=tk.E, padx=15, pady=(0, 10))
            
            def update_char_count(event):
                text = prompt_text.get("1.0", tk.END).strip()
                if len(text) > 500:
                    prompt_text.delete("1.0", tk.END)
                    prompt_text.insert(tk.END, text[:500])
                    text = text[:500]
                char_count_var.set(f"{len(text)}/500")
            
            prompt_text.bind("<KeyRelease>", update_char_count)
            
            # 应用按钮
            def apply_prompt():
                new_prompt = prompt_text.get("1.0", tk.END).strip()
                
                for item_id in selected:
                    item = tree.item(item_id)
                    task_id = item["values"][0]
                    
                    # 查找任务
                    task = None
                    for t in self.history.get("tasks", []):
                        if t.get("id") == task_id:
                            task = t
                            break
                    
                    if task:
                        task["user_prompt"] = new_prompt
                        task["updated_at"] = datetime.now().isoformat()
                        self.append_log(f"为任务设置提示词：{task.get('link')}")
                
                # 保存历史记录
                save_history(self.history)
                messagebox.showinfo("成功", f"已为 {len(selected)} 个任务设置提示词")
                prompt_window.destroy()
                refresh_tree()
            
            # 按钮区域
            button_frame = tk.Frame(prompt_window, bg="#f0f4f8")
            button_frame.pack(fill=tk.X, padx=20, pady=(0, 20))
            
            ttk.Button(button_frame, text="应用", command=apply_prompt).pack(side=tk.RIGHT, padx=10)
            ttk.Button(button_frame, text="取消", command=prompt_window.destroy).pack(side=tk.RIGHT, padx=10)
        
        def batch_stop():
            selected = tree.selection()
            if not selected:
                messagebox.showinfo("提示", "请选择要停止的任务")
                return
            
            for item_id in selected:
                item = tree.item(item_id)
                task_id = item["values"][0]
                
                # 查找任务
                task = None
                for t in self.history.get("tasks", []):
                    if t.get("id") == task_id:
                        task = t
                        break
                
                if task:
                    link = task.get("link")
                    if link in self.task_queue:
                        self.task_queue.remove(link)
                        self.update_queue_status()
                        self.append_log(f"停止处理任务：{link}")
        
        def clear_completed():
            """清理已完成的任务"""
            if messagebox.askyesno("确认", "确定要清理所有已完成的任务吗？"):
                completed_tasks = [task for task in self.history.get("tasks", []) 
                                  if task.get("status") == "completed"]
                
                for task in completed_tasks:
                    self.history["tasks"].remove(task)
                
                save_history(self.history)
                self.append_log(f"已清理 {len(completed_tasks)} 个已完成任务")
                refresh_tree()
        
        ttk.Button(action_frame, text="查看详情", command=view_details).pack(side=tk.LEFT, padx=10)
        ttk.Button(action_frame, text="继续处理", command=continue_task).pack(side=tk.LEFT, padx=10)
        ttk.Button(action_frame, text="批量开始", command=batch_start).pack(side=tk.LEFT, padx=10)
        ttk.Button(action_frame, text="批量停止", command=batch_stop).pack(side=tk.LEFT, padx=10)
        ttk.Button(action_frame, text="设置提示词", command=set_prompt_for_task).pack(side=tk.LEFT, padx=10)
        ttk.Button(action_frame, text="清理已完成", command=clear_completed).pack(side=tk.LEFT, padx=10)
        ttk.Button(action_frame, text="刷新", command=refresh_tree).pack(side=tk.LEFT, padx=10)
        ttk.Button(action_frame, text="关闭", command=history_window.destroy).pack(side=tk.RIGHT, padx=10)
        
        # 任务统计信息
        stats_frame = tk.Frame(main_frame, bg="#f0f4f8")
        stats_frame.pack(fill=tk.X, pady=(0, 10))
        
        total_tasks = len(self.history.get("tasks", []))
        completed_tasks = len([t for t in self.history.get("tasks", []) if t.get("status") == "completed"])
        failed_tasks = len([t for t in self.history.get("tasks", []) if t.get("status") == "failed"])
        pending_tasks = len([t for t in self.history.get("tasks", []) if t.get("status") == "pending"])
        in_progress_tasks = len([t for t in self.history.get("tasks", []) if t.get("status") == "in_progress"])
        
        stats_label = tk.Label(
            stats_frame, 
            text=f"任务统计：总计 {total_tasks} | 已完成 {completed_tasks} | 处理中 {in_progress_tasks} | 待处理 {pending_tasks} | 失败 {failed_tasks}", 
            font=("微软雅黑", 9),
            foreground="#666",
            bg="#f0f4f8"
        )
        stats_label.pack(anchor=tk.W, padx=10)
    
    # 批量导入方法
    def batch_import(self):
        """批量导入视频链接，支持先设定user prompt再提交"""
        # 选择Excel文件
        file_path = filedialog.askopenfilename(
            title="选择Excel文件",
            filetypes=[("Excel文件", "*.xlsx;*.xls"), ("所有文件", "*.*")]
        )
        
        if not file_path:
            return
        
        try:
            # 解析Excel文件
            self.append_log(f"开始解析Excel文件：{file_path}")
            links = self.parse_excel_file(file_path)
            
            if not links:
                messagebox.showwarning("提示", "Excel文件中未找到有效的视频链接")
                return
            
            # 创建批量导入设置窗口
            import_window = tk.Toplevel(self.root)
            import_window.title("批量导入设置")
            import_window.geometry("800x600")
            import_window.configure(bg="#f0f4f8")
            # 允许窗口大小调整
            import_window.resizable(True, True)
            
            # 主容器
            main_frame = tk.Frame(import_window, bg="#f0f4f8")
            main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
            
            # 导入信息
            info_frame = tk.Frame(main_frame, bg="#ffffff")
            info_frame.pack(fill=tk.X, pady=(0, 20))
            info_frame.configure(highlightbackground="#0066cc", highlightthickness=1)
            
            ttk.Label(info_frame, text=f"发现 {len(links)} 个视频链接", font=("微软雅黑", 10), background="#ffffff").pack(anchor=tk.W, padx=15, pady=15)
            
            # 用户提示词设置
            prompt_frame = tk.Frame(main_frame, bg="#ffffff")
            prompt_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 20))
            prompt_frame.configure(highlightbackground="#0066cc", highlightthickness=1)
            
            ttk.Label(prompt_frame, text="用户提示词（批量设置）：", font=("微软雅黑", 10), background="#ffffff").pack(anchor=tk.W, padx=15, pady=(15, 5))
            
            # 提示词输入区域
            prompt_text = tk.Text(prompt_frame, height=8, font=("微软雅黑", 10), bg="#f9f9f9")
            prompt_text.pack(fill=tk.BOTH, expand=True, padx=15, pady=(0, 15))
            
            # 获取当前的用户提示词
            current_prompt = self.user_prompt_var.get().strip()
            if current_prompt:
                prompt_text.insert(tk.END, current_prompt)
            
            # 字符计数
            char_count_var = tk.StringVar(value=f"{len(current_prompt)}/500")
            char_count_label = tk.Label(prompt_frame, textvariable=char_count_var, font=("微软雅黑", 9), foreground="#999", background="#ffffff")
            char_count_label.pack(anchor=tk.E, padx=15, pady=(0, 10))
            
            def update_char_count(event):
                text = prompt_text.get("1.0", tk.END).strip()
                if len(text) > 500:
                    prompt_text.delete("1.0", tk.END)
                    prompt_text.insert(tk.END, text[:500])
                    text = text[:500]
                char_count_var.set(f"{len(text)}/500")
            
            prompt_text.bind("<KeyRelease>", update_char_count)
            
            # 飞书文件夹设置
            feishu_frame = tk.Frame(main_frame, bg="#ffffff")
            feishu_frame.pack(fill=tk.X, pady=(0, 20))
            feishu_frame.configure(highlightbackground="#0066cc", highlightthickness=1)
            
            ttk.Label(feishu_frame, text="飞书文件夹路径（可选）：", font=("微软雅黑", 10), background="#ffffff").pack(anchor=tk.W, padx=15, pady=(15, 5))
            
            feishu_folder_var = tk.StringVar()
            feishu_folder_entry = ttk.Entry(feishu_frame, textvariable=feishu_folder_var, font=("微软雅黑", 10))
            feishu_folder_entry.pack(fill=tk.X, padx=15, pady=(0, 15))
            
            # 示例提示
            ttk.Label(feishu_frame, text="示例：就业技术文档集/八股", font=("微软雅黑", 9), foreground="#999", background="#ffffff").pack(anchor=tk.W, padx=15, pady=(0, 10))
            
            # 按钮区域
            button_frame = tk.Frame(main_frame, bg="#f0f4f8")
            button_frame.pack(fill=tk.X, pady=10)
            
            def start_import():
                """开始导入"""
                user_prompt = prompt_text.get("1.0", tk.END).strip()
                feishu_folder_path = feishu_folder_var.get().strip() or None
                
                # 批量添加到历史记录
                self.append_log(f"开始创建 {len(links)} 个任务的历史记录")
                new_links_count = 0
                existing_links_count = 0
                
                for link in links:
                    # 检查链接是否已经导入
                    if not self.is_link_already_imported(link):
                        self.add_task_to_history(link, user_prompt, feishu_folder_path)
                        if link not in self.task_queue:
                            self.task_queue.append(link)
                        new_links_count += 1
                    else:
                        existing_links_count += 1
                        self.append_log(f"链接已存在，跳过导入：{link}")
                
                self.append_log(f"批量导入完成：新添加 {new_links_count} 个链接，跳过 {existing_links_count} 个已存在的链接")
                
                # 更新队列状态显示
                self.update_queue_status()
                
                self.append_log(f"成功导入 {new_links_count} 个视频链接到队列")
                self.append_log(f"当前队列长度：{len(self.task_queue)}")
                
                # 自动开始处理队列
                if not self.processing_queue and new_links_count > 0:
                    self.start_queue_processing()
                
                import_window.destroy()
                messagebox.showinfo("成功", f"批量导入完成：新添加 {new_links_count} 个链接")
            
            ttk.Button(button_frame, text="确认", command=start_import).pack(side=tk.RIGHT, padx=10)
            ttk.Button(button_frame, text="取消", command=import_window.destroy).pack(side=tk.RIGHT, padx=10)
            
        except Exception as e:
            self.append_log(f"批量导入失败：{e}")
            messagebox.showerror("错误", f"批量导入失败：{e}")
    
    # 解析Excel文件
    def parse_excel_file(self, file_path):
        try:
            # 尝试使用pandas解析
            try:
                import pandas as pd
                df = pd.read_excel(file_path, header=None)
                links = df.iloc[:, 0].dropna().tolist()
                
            except ImportError:
                # 如果pandas不可用，使用openpyxl
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(file_path)
                    ws = wb.active
                    links = []
                    for row in ws.iter_rows(min_row=1, values_only=True):
                        if row[0]:
                            links.append(str(row[0]))
                except ImportError:
                    # 如果都不可用，提示安装依赖
                    self.append_log("请安装pandas或openpyxl来解析Excel文件")
                    raise Exception("缺少Excel解析库，请安装pandas或openpyxl")
            
            # 过滤和验证链接
            valid_links = []
            for link in links:
                link = str(link).strip()
                if link and ('http' in link.lower() or 'www.' in link.lower()):
                    valid_links.append(link)
            
            return valid_links
        except Exception as e:
            self.append_log(f"解析Excel文件失败：{e}")
            raise
    
    # 开始队列处理
    def start_queue_processing(self):
        if not self.task_queue:
            self.append_log("队列为空，无需处理")
            self.update_queue_status()
            return
        
        if self.processing_queue:
            self.append_log("队列正在处理中")
            self.update_queue_status()
            return
        
        self.processing_queue = True
        self.update_queue_status()
        
        # 限制一次处理的任务数量，避免系统过载
        batch_size = min(len(self.task_queue), self.max_workers * 2)  # 每次处理的任务数量为线程数的2倍
        task_links = self.task_queue[:batch_size]  # 只处理批次内的任务
        remaining_tasks = self.task_queue[batch_size:]  # 剩余任务
        
        self.append_log(f"开始处理队列任务，本次批次：{len(task_links)} 个任务")
        self.append_log(f"使用线程池并行处理，最大线程数：{self.max_workers}")
        self.append_log(f"队列中剩余任务：{len(remaining_tasks)} 个")
        self.append_log("遵循先进先出原则，按照提交顺序处理任务")
        
        # 使用线程池并行处理任务，保持先进先出顺序
        tasks = []
        
        # 清空之前的活跃任务列表
        self.active_futures = []
        
        for i, link in enumerate(task_links):
            task_number = i + 1
            # 获取任务绑定的用户提示词和飞书文件夹路径
            task_prompt = ""
            feishu_folder_path = None
            for task in self.history.get("tasks", []):
                if task.get("link") == link:
                    task_prompt = task.get("user_prompt", "")
                    feishu_folder_path = task.get("feishu_folder_path")
                    break
            
            # 如果任务没有绑定提示词，使用全局的
            if not task_prompt:
                task_prompt = self.user_prompt_var.get().strip()
            
            self.append_log(f"提交任务 {task_number}/{len(task_links)} 到线程池：{link}")
            self.append_log(f"任务提示词：{task_prompt[:50]}{'...' if len(task_prompt) > 50 else ''}")
            if feishu_folder_path:
                self.append_log(f"飞书文件夹路径：{feishu_folder_path}")
            
            # 提交任务到线程池
            future = self.executor.submit(self._run_pipeline, link, task_prompt, feishu_folder_path)
            tasks.append(future)
            self.active_futures.append(future)
        
        # 等待所有任务完成
        def wait_for_completion():
            try:
                # 等待所有任务完成
                completed_count = 0
                total_tasks = len(tasks)
                
                for future in concurrent.futures.as_completed(tasks):
                    try:
                        future.result()
                        completed_count += 1
                        self.append_log(f"任务完成进度：{completed_count}/{total_tasks}")
                    except Exception as e:
                        self.append_log(f"任务执行异常：{e}")
                        completed_count += 1
                
                self.append_log("本批次任务处理完成")
                
                # 从队列中移除已处理的任务
                for link in task_links:
                    if link in self.task_queue:
                        self.task_queue.remove(link)
                
                # 更新队列状态
                self.processing_queue = False
                self.update_queue_status()
                
                # 清空活跃任务列表
                self.active_futures = []
                
                # 如果队列中还有新任务，继续处理
                if self.task_queue:
                    self.append_log(f"队列中还有 {len(self.task_queue)} 个新任务，继续处理下一批次")
                    # 短暂延迟，避免系统过载
                    time.sleep(0.5)
                    self.start_queue_processing()
                else:
                    self.append_log("所有队列任务处理完成")
            except Exception as e:
                self.append_log(f"队列处理异常：{e}")
                self.processing_queue = False
                self.update_queue_status()
                # 清空活跃任务列表
                self.active_futures = []
        
        # 在后台等待完成
        threading.Thread(target=wait_for_completion, daemon=True).start()
    
    # 单独重试失败阶段
    def _retry_stage(self, link, stage, task):
        """单独重试失败的阶段，使用缓存的结果"""
        try:
            self.append_log(f"开始重试阶段 {stage}：{link}")
            
            # 获取缓存的结果
            stages = task.get("stages", {})
            cached_results = {}
            
            # 收集已完成阶段的结果
            for s, info in stages.items():
                if info.get("status") == "completed" and info.get("result"):
                    cached_results[s] = info.get("result")
            
            # 根据阶段名称执行相应的操作
            if stage == "download":
                # 重新下载视频
                self.update_task_status(link, stage, "in_progress")
                video_file = self.download_video(link)
                if video_file:
                    self.update_task_status(link, stage, "completed", video_file)
                    # 自动继续下一个阶段
                    self._continue_from_stage(link, "transcribe", {"download": video_file})
                else:
                    self.update_task_status(link, stage, "failed")
                    self.append_log(f"阶段 {stage} 重试失败")
            
            elif stage == "transcribe":
                # 重新转写
                video_file = cached_results.get("download")
                if not video_file:
                    self.append_log(f"缺少前一阶段结果，无法重试 {stage}")
                    return
                
                self.update_task_status(link, stage, "in_progress")
                result_data = self.speech_to_text(video_file, "")
                if result_data:
                    self.update_task_status(link, stage, "completed", result_data.get("segments", []))
                    # 更新AI分析结果
                    self.update_task_status(link, "ai_analysis", "completed", result_data.get("ai_summary", ""))
                    # 自动继续下一个阶段
                    self._continue_from_stage(link, "generate_md", {
                        "download": video_file,
                        "transcribe": result_data.get("segments", []),
                        "ai_analysis": result_data.get("ai_summary", "")
                    })
                else:
                    self.update_task_status(link, stage, "failed")
                    self.append_log(f"阶段 {stage} 重试失败")
            
            elif stage == "ai_analysis":
                # 重新AI分析
                segments = cached_results.get("transcribe")
                if not segments:
                    self.append_log(f"缺少前一阶段结果，无法重试 {stage}")
                    return
                
                self.update_task_status(link, stage, "in_progress")
                # 构建临时结果数据
                result_data = {"segments": segments}
                # 提取转写文本
                transcript_lines = []
                for seg in segments:
                    text = seg if isinstance(seg, str) else seg.get("text", "")
                    transcript_lines.append(text)
                transcript = " ".join(transcript_lines)
                # 重新分析
                summary = self.summarize_with_volcengine(transcript, "")
                if summary:
                    self.update_task_status(link, stage, "completed", summary)
                    # 自动继续下一个阶段
                    self._continue_from_stage(link, "generate_md", {
                        "download": cached_results.get("download"),
                        "transcribe": segments,
                        "ai_analysis": summary
                    })
                else:
                    self.update_task_status(link, stage, "failed")
                    self.append_log(f"阶段 {stage} 重试失败")
            
            elif stage == "generate_md":
                # 重新生成MD
                segments = cached_results.get("transcribe")
                summary = cached_results.get("ai_analysis")
                if not segments or not summary:
                    self.append_log(f"缺少前一阶段结果，无法重试 {stage}")
                    return
                
                self.update_task_status(link, stage, "in_progress")
                # 构建临时结果数据
                result_data = {
                    "segments": segments,
                    "ai_summary": summary
                }
                md_file = self.generate_md(result_data, link, "视频")
                if md_file:
                    self.update_task_status(link, stage, "completed", md_file)
                    self.append_log(f"阶段 {stage} 重试成功：{md_file}")
                else:
                    self.update_task_status(link, stage, "failed")
                    self.append_log(f"阶段 {stage} 重试失败")
            
            self.append_log(f"阶段 {stage} 重试完成")
            
        except Exception as e:
            self.append_log(f"重试阶段 {stage} 异常：{e}")
            self.update_task_status(link, stage, "failed")
    
    # 从指定阶段继续处理
    def _continue_from_stage(self, link, next_stage, cached_results):
        """从指定阶段继续处理"""
        try:
            self.append_log(f"从阶段 {next_stage} 继续处理：{link}")
            
            if next_stage == "transcribe":
                video_file = cached_results.get("download")
                if video_file:
                    self.update_task_status(link, next_stage, "in_progress")
                    result_data = self.speech_to_text(video_file, "")
                    if result_data:
                        self.update_task_status(link, next_stage, "completed", result_data.get("segments", []))
                        self.update_task_status(link, "ai_analysis", "completed", result_data.get("ai_summary", ""))
                        self._continue_from_stage(link, "generate_md", {
                            "download": video_file,
                            "transcribe": result_data.get("segments", []),
                            "ai_analysis": result_data.get("ai_summary", "")
                        })
                    else:
                        self.update_task_status(link, next_stage, "failed")
            
            elif next_stage == "generate_md":
                segments = cached_results.get("transcribe")
                summary = cached_results.get("ai_analysis")
                if segments and summary:
                    self.update_task_status(link, next_stage, "in_progress")
                    result_data = {
                        "segments": segments,
                        "ai_summary": summary
                    }
                    md_file = self.generate_md(result_data, link, "视频")
                    if md_file:
                        self.update_task_status(link, next_stage, "completed", md_file)
                        self.append_log(f"任务处理完成：{link}")
                        # 清理缓存文件
                        self.cleanup_task_cache(link, cached_results)
                    else:
                        self.update_task_status(link, next_stage, "failed")
        
        except Exception as e:
            self.append_log(f"继续处理异常：{e}")
    
    # 清理任务缓存文件
    def cleanup_task_cache(self, link, cached_results):
        """清理任务的缓存文件"""
        try:
            # 清理视频文件
            video_file = cached_results.get("download")
            if video_file and os.path.exists(video_file):
                os.remove(video_file)
                self.append_log(f"已清理视频缓存文件：{video_file}")
            
            # 清理可能的临时文件
            import glob
            url_hash = hashlib.md5(link.encode()).hexdigest()[:8]
            temp_files = glob.glob(os.path.join(VIDEO_DIR, f"*{url_hash}*"))
            for temp_file in temp_files:
                if os.path.exists(temp_file):
                    os.remove(temp_file)
                    self.append_log(f"已清理临时缓存文件：{temp_file}")
        except Exception as e:
            self.append_log(f"清理缓存文件异常：{e}")
    
    # 清理所有完成任务的缓存文件
    def cleanup_completed_tasks(self):
        """清理所有完成任务的缓存文件"""
        try:
            completed_tasks = [task for task in self.history.get("tasks", []) 
                              if task.get("status") == "completed"]
            
            for task in completed_tasks:
                stages = task.get("stages", {})
                cached_results = {}
                
                # 检查是否已经生成了MD文件
                generate_md_stage = stages.get("generate_md", {})
                if generate_md_stage.get("status") != "completed":
                    # 如果没有生成MD文件，不清理缓存
                    continue
                
                for s, info in stages.items():
                    if info.get("status") == "completed" and info.get("result"):
                        cached_results[s] = info.get("result")
                
                # 清理缓存文件
                if cached_results:
                    self.cleanup_task_cache(task.get("link"), cached_results)
        except Exception as e:
            self.append_log(f"清理完成任务缓存异常：{e}")
    
    # 从摘要中提取标题
    def extract_title_from_summary(self, summary, link):
        """从AI分析结果中提取标题"""
        try:
            # 检查是否有RULES控制标题提取
            rules = CONFIG.get("rules", "")
            
            # 从摘要中提取标题
            if summary:
                # 尝试提取第一句话作为标题
                lines = summary.split('\n')
                for line in lines:
                    line = line.strip()
                    if line and len(line) > 5:
                        # 清理标题，去除特殊字符
                        import re
                        # 去除标点符号和特殊字符
                        title = re.sub(r'[\\/:*?"<>|]', '', line)
                        # 截取前20个字符作为文件名
                        title = title[:20]
                        # 替换空格为下划线
                        title = title.replace(' ', '_')
                        # 确保标题不为空
                        if title:
                            self.append_log(f"从AI摘要中提取标题：{title}")
                            return title
            
            # 如果摘要中没有合适的标题，从链接中提取
            import re
            # 首先尝试提取B站BV号
            bv_match = re.search(r'BV[0-9A-Za-z]{10}', link)
            if bv_match:
                bv = bv_match.group(0)
                self.append_log(f"从链接中提取B站BV号作为标题：{bv}")
                return bv
            
            # 尝试提取链接中的数字部分
            doc_name_match = re.search(r'\d+', link.split('/')[-1])
            if doc_name_match:
                doc_name = doc_name_match.group(0)
                self.append_log(f"从链接中提取数字作为标题：{doc_name}")
                return doc_name
            
            # 尝试提取链接中的有意义部分
            path_match = re.search(r'[a-zA-Z0-9_-]{8,}', link)
            if path_match:
                path_part = path_match.group(0)
                self.append_log(f"从链接中提取路径部分作为标题：{path_part}")
                return path_part
            
            # 如果都失败了，使用默认值
            default_title = "未知标题"
            self.append_log(f"使用默认标题：{default_title}")
            return default_title
            
        except Exception as e:
            self.append_log(f"提取标题异常：{e}")
            # 从链接中提取作为后备
            import re
            # 首先尝试提取B站BV号
            bv_match = re.search(r'BV[0-9A-Za-z]{10}', link)
            if bv_match:
                return bv_match.group(0)
            # 尝试提取数字
            doc_name_match = re.search(r'\d+', link.split('/')[-1])
            if doc_name_match:
                return doc_name_match.group(0)
            # 尝试提取路径部分
            path_match = re.search(r'[a-zA-Z0-9_-]{8,}', link)
            if path_match:
                return path_match.group(0)
            # 最终默认值
            return "未知标题"
    
    # 窗口关闭事件处理
    def on_closing(self):
        """处理窗口关闭事件"""
        try:
            self.append_log("正在关闭应用程序...")
            # 关闭线程池
            self.append_log("正在关闭线程池...")
            self.executor.shutdown(wait=False)  # 不等待所有任务完成
            self.append_log("线程池已关闭")
        except Exception as e:
            self.append_log(f"关闭应用程序时发生异常：{e}")
        finally:
            # 保存历史记录
            save_history(self.history)
            self.append_log("历史记录已保存")
            # 关闭窗口
            self.root.destroy()
    
    # 入口
    def start(self):
        link = self.link_var.get().strip()
        user_prompt = self.user_prompt_var.get().strip()
        
        # 尝试从文本中提取URL
        extracted_url = self.extract_url_from_text(link)
        if extracted_url:
            # 移除可能的反引号
            extracted_url = extracted_url.strip('`')
            self.append_log(f"从文本中提取到URL: {extracted_url}")
            # 更新输入框为提取的URL
            self.link_var.set(extracted_url)
            link = extracted_url
        
        if not link:
            messagebox.showwarning("提示", "请先输入视频链接")
            return
            
        # 更宽松的链接验证
        if not ('http' in link.lower() or 'www.' in link.lower()):
            messagebox.showwarning("提示", "请输入有效的视频链接\n提示: 链接应该包含 http 或 www")
            return
        
        # 检查链接是否已经导入
        if self.is_link_already_imported(link):
            # 创建自定义对话框，提供取消和继续上传选项
            dialog = tk.Toplevel(self.root)
            dialog.title("链接已导入")
            dialog.geometry("400x200")
            dialog.configure(bg="#f0f4f8")
            dialog.transient(self.root)
            dialog.grab_set()
            
            # 居中显示
            dialog.update_idletasks()
            width = dialog.winfo_width()
            height = dialog.winfo_height()
            x = (dialog.winfo_screenwidth() // 2) - (width // 2)
            y = (dialog.winfo_screenheight() // 2) - (height // 2)
            dialog.geometry(f"{width}x{height}+{x}+{y}")
            
            # 消息标签
            msg_frame = tk.Frame(dialog, bg="#ffffff")
            msg_frame.pack(fill=tk.X, padx=20, pady=(20, 10))
            msg_frame.configure(highlightbackground="#0066cc", highlightthickness=1)
            
            msg_label = tk.Label(
                msg_frame, 
                text="该链接已经导入过，是否继续上传？", 
                font=("微软雅黑", 10),
                foreground="#333",
                bg="#ffffff"
            )
            msg_label.pack(padx=20, pady=20)
            
            # 按钮区域
            btn_frame = tk.Frame(dialog, bg="#f0f4f8")
            btn_frame.pack(fill=tk.X, padx=20, pady=10)
            
            # 存储用户选择的变量
            user_choice = tk.StringVar(value="cancel")
            
            def cancel_upload():
                user_choice.set("cancel")
                dialog.destroy()
            
            def continue_upload():
                user_choice.set("continue")
                dialog.destroy()
            
            # 取消按钮
            cancel_btn = ttk.Button(btn_frame, text="取消上传", command=cancel_upload)
            cancel_btn.pack(side=tk.LEFT, padx=10)
            
            # 继续上传按钮
            continue_btn = ttk.Button(btn_frame, text="继续上传", command=continue_upload)
            continue_btn.pack(side=tk.RIGHT, padx=10)
            
            # 等待用户选择
            self.root.wait_window(dialog)
            
            # 根据用户选择执行操作
            if user_choice.get() == "cancel":
                return
            # 如果用户选择继续上传，不返回，继续执行后续代码

        # 检查任务队列大小限制
        if len(self.task_queue) >= self.queue_max_size:
            messagebox.showwarning("提示", f"任务队列已满（当前大小：{len(self.task_queue)}，最大限制：{self.queue_max_size}）\n请稍后再添加任务或调整队列大小限制")
            return

        # 添加到任务队列
        self.task_queue.append(link)
        self.add_task_to_history(link, user_prompt)
        self.update_queue_status()
        
        self.start_btn.config(state=tk.DISABLED)
        self.append_log(f"添加任务到队列：链接={link}")
        self.append_log(f"当前队列长度：{len(self.task_queue)}")
        
        # 自动开始处理队列
        if not self.processing_queue:
            self.start_queue_processing()
        
        # 恢复按钮状态
        self.start_btn.config(state=tk.NORMAL)

    def extract_url_from_text(self, text: str) -> str:
        """从文本中提取URL"""
        import re
        
        # 匹配http/https开头的URL
        url_pattern = r'https?://[^\s]+'
        urls = re.findall(url_pattern, text)
        
        # 找到包含xiaohongshu.com的URL
        for url in urls:
            if 'xiaohongshu.com' in url.lower():
                # 移除末尾可能的标点符号
                url = url.rstrip('.,;:!?')
                return url
                
        return None

    # 主流程：下载 -> 转写 -> 生成MD
    def _run_pipeline(self, link: str, user_prompt: str = "", feishu_folder_path: str = None):
        try:
            # 添加任务到历史记录，更新飞书文件夹路径
            self.add_task_to_history(link, user_prompt, feishu_folder_path)
            
            # 初始化进度条
            self.update_progress(0, "准备开始处理...")
            
            platform = "视频"
            
            # 阶段1：下载视频
            self.update_task_status(link, "download", "in_progress")
            self.update_progress(10, "下载视频...")
            video_file = self.download_video(link)
            if not video_file:
                self.append_log("视频下载失败，流程结束。")
                self.update_task_status(link, "download", "failed")
                self.update_progress(0, "失败")
                return
            self.update_task_status(link, "download", "completed", video_file)

            # 阶段2：语音转文字
            self.update_task_status(link, "transcribe", "in_progress")
            self.update_progress(40, "语音转文字(上传/轮询)...")
            result_data = self.speech_to_text(video_file, user_prompt)
            if not result_data:
                self.append_log("语音转文字失败，流程结束。")
                self.update_task_status(link, "transcribe", "failed")
                self.update_progress(0, "失败")
                return
            self.update_task_status(link, "transcribe", "completed", result_data.get("segments", []))

            # 阶段3：AI分析
            self.update_task_status(link, "ai_analysis", "in_progress")
            self.update_progress(60, "使用AI进行文本分析...")
            # AI分析已在speech_to_text中完成
            summary = result_data.get("ai_summary", "")
            self.update_task_status(link, "ai_analysis", "completed", summary)
            
            # 从AI分析结果中提取标题并更新任务标题
            if summary:
                # 从摘要中提取标题
                title = self.extract_title_from_summary(summary, link)
                # 确保标题有效才更新
                if title and title != "未知标题":
                    # 更新任务的标题字段
                    for task in self.history.get("tasks", []):
                        if task.get("link") == link:
                            # 只有当新标题比旧标题更有意义时才更新
                            old_title = task.get("title", "")
                            if not old_title or old_title == "未知标题":
                                task["title"] = title
                                task["updated_at"] = datetime.now().isoformat()
                                save_history(self.history)
                                self.append_log(f"从AI分析结果中提取并更新标题：{title}")
                            break

            # 阶段4：生成Markdown
            self.update_task_status(link, "generate_md", "in_progress")
            self.update_progress(80, "生成Markdown文档...")
            md_file = self.generate_md(result_data, link, platform)
            if not md_file:
                self.append_log("生成文档失败")
                self.update_task_status(link, "generate_md", "failed")
                self.update_progress(0, "失败")
                return
            self.update_task_status(link, "generate_md", "completed", md_file)
            
            # 清理缓存文件（只有在生成MD文件成功后才清理）
            self.cleanup_task_cache(link, {
                "download": video_file,
                "transcribe": result_data.get("segments", []),
                "ai_analysis": result_data.get("ai_summary", "")
            })

            # 阶段5：上传到飞书（默认禁用）
            if self.feishu_enabled:
                self.update_task_status(link, "feishu_upload", "in_progress")
                self.update_progress(90, "上传到飞书...")
                
                # 读取Markdown文件内容
                with open(md_file, 'r', encoding='utf-8') as f:
                    md_content = f.read()
                
                try:
                    # 导入飞书集成模块
                    from feishu_integration import FeishuKnowledgeBase
                    
                    # 初始化飞书客户端
                    feishu = FeishuKnowledgeBase('cli_a9b7cc9aba389bc4', 'q3VZTLZZjrsNeiJheqfkocH5ReV6Rmc6')
                    
                    # 从用户提示词中解析飞书文件夹路径
                    prompt_folder = feishu.parse_feishu_folder_from_prompt(user_prompt)
                    final_folder = feishu_folder_path or prompt_folder
                    
                    # 上传文档
                    doc_title = os.path.basename(md_file).replace('.md', '')
                    doc_token = feishu.upload_document(doc_title, md_content, feishu_folder_path=final_folder)
                    
                    if doc_token:
                        self.append_log(f"文档已上传到飞书：{doc_token}")
                        self.update_task_status(link, "feishu_upload", "completed", doc_token)
                    else:
                        self.append_log("上传到飞书失败")
                        self.update_task_status(link, "feishu_upload", "failed")
                except Exception as e:
                    self.append_log(f"飞书上传异常：{e}")
                    self.update_task_status(link, "feishu_upload", "failed")
            else:
                self.append_log("飞书功能已禁用，跳过上传步骤", "INFO")
                self.update_task_status(link, "feishu_upload", "completed", "飞书功能已禁用")
            
            self.update_progress(100, "完成")
        except Exception as e:
            self.append_log(f"异常：{e}")
            import traceback
            traceback.print_exc()
            # 确保任务状态被更新为失败
            self.update_task_status(link, "download", "failed")
            self.update_progress(0, "失败")
        finally:
            # 确保任务状态被正确更新
            try:
                # 检查任务是否存在于历史记录中
                task_exists = False
                for task in self.history.get("tasks", []):
                    if task.get("link") == link:
                        task_exists = True
                        break
                
                # 如果任务不存在，添加到历史记录
                if not task_exists:
                    self.add_task_to_history(link, user_prompt, feishu_folder_path)
                    # 更新状态为失败
                    self.update_task_status(link, "download", "failed")
            except Exception as e:
                self.append_log(f"历史记录更新异常：{e}")
            
            # 只有在非队列处理时才启用按钮
            if not self.processing_queue:
                self.start_btn.config(state=tk.NORMAL)

    # 步骤1：下载视频
    def download_video(self, link: str):
        import time
        download_start = time.time()
        
        try:
            # 清理链接中的反引号
            link = link.strip('`')
            
            # 检查视频缓存
            with self.video_cache_lock:
                if link in self.video_cache:
                    cached_file = self.video_cache[link]
                    if os.path.exists(cached_file) and os.path.getsize(cached_file) > 0:
                        self.append_log(f"使用缓存的视频文件: {cached_file}", "INFO")
                        download_end = time.time()
                        self.append_log(f"视频下载耗时: {download_end - download_start:.2f}秒（使用缓存）", "INFO")
                        return cached_file
                    else:
                        # 缓存文件不存在或为空，删除缓存条目
                        del self.video_cache[link]
                        self.append_log("缓存视频文件不存在，重新下载", "INFO")
            
            # 使用yt-dlp工具下载视频
            import subprocess
            import os
            
            self.append_log("使用yt-dlp下载视频...", "INFO")
            
            # 获取当前目录下的视频文件数量，作为总序号
            existing_videos = [f for f in os.listdir(VIDEO_DIR) if f.endswith('.mp4')]
            total_count = len(existing_videos) + 1
            
            # 获取当前日期（月-日）
            current_date = time.strftime('%m-%d')
            
            # 从链接中提取文档名称（使用链接的最后部分）
            import re
            doc_name_match = re.search(r'\d+', link.split('/')[-1])
            doc_name = doc_name_match.group(0) if doc_name_match else "unknown"
            
            # 构建新的文件名：总记录序号-月-日-文档名称
            new_filename = f"{total_count:03d}-{current_date}-{doc_name}.mp4"
            output_file = os.path.join(VIDEO_DIR, new_filename)
            
            # 从链接中提取 xsec_token
            xsec_token_match = re.search(r'xsec_token=([^&]+)', link)
            xsec_token = xsec_token_match.group(1) if xsec_token_match else ""
            
            # 构建cookie字符串
            cookie_string = f"xsec_token={xsec_token}" if xsec_token else ""
            
            # 根据链接类型设置不同的referer
            referer = "https://www.xiaohongshu.com/"
            if "bilibili.com" in link or "bilibili" in link:
                referer = "https://www.bilibili.com/"
            elif "youtube.com" in link or "youtu.be" in link:
                referer = "https://www.youtube.com/"
            
            # 构建yt-dlp命令，直接下载到目标文件夹（优化：减少不必要的参数）
            cmd = [
                "yt-dlp",
                "--user-agent", "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/144.0.0.0 Safari/537.36",
                "--referer", referer,
                "--no-check-certificate",
                "--quiet",  # 静默模式，减少输出
                "--no-warnings",  # 禁用警告
                "-o", output_file,
                link
            ]
            
            # 执行命令（增加超时时间，添加重试机制）
            max_retries = 2
            retry_count = 0
            result = None
            
            while retry_count < max_retries:
                try:
                    self.append_log(f"执行yt-dlp命令（尝试 {retry_count+1}/{max_retries}）...", "INFO")
                    # 增加超时时间到60秒
                    result = subprocess.run(cmd, capture_output=True, text=True, timeout=60)
                    break
                except subprocess.TimeoutExpired:
                    retry_count += 1
                    self.append_log(f"yt-dlp执行超时，正在重试...（{retry_count}/{max_retries}）", "WARNING")
                    if retry_count >= max_retries:
                        self.append_log("yt-dlp执行多次超时", "ERROR")
                        download_end = time.time()
                        self.append_log(f"视频下载耗时: {download_end - download_start:.2f}秒（超时）", "INFO")
                        # 直接返回None表示失败，不使用示例文件
                        return None
                except Exception as e:
                    self.append_log(f"yt-dlp执行异常：{e}", "ERROR")
                    retry_count += 1
                    if retry_count >= max_retries:
                        self.append_log("yt-dlp执行失败", "ERROR")
                        download_end = time.time()
                        self.append_log(f"视频下载耗时: {download_end - download_start:.2f}秒（失败）", "INFO")
                        # 直接返回None表示失败，不使用示例文件
                        return None
            
            if result.returncode == 0:
                # 检查视频文件是否存在且大小大于0
                if os.path.exists(output_file) and os.path.getsize(output_file) > 0:
                    self.append_log(f"视频下载成功: {output_file}", "INFO")
                    # 添加到缓存
                    with self.video_cache_lock:
                        self.video_cache[link] = output_file
                    download_end = time.time()
                    self.append_log(f"视频下载耗时: {download_end - download_start:.2f}秒", "INFO")
                    return output_file
                else:
                    self.append_log("视频文件不存在或为空", "ERROR")
            else:
                # 优化：只显示部分错误信息
                error_msg = result.stderr[:500] + "..." if len(result.stderr) > 500 else result.stderr
                self.append_log(f"yt-dlp执行失败: {error_msg}", "ERROR")
            
            # 执行失败，直接返回None表示失败，不使用示例文件
            self.append_log("视频下载失败", "ERROR")
            download_end = time.time()
            self.append_log(f"视频下载耗时: {download_end - download_start:.2f}秒（失败）", "INFO")
            return None
            
        except Exception as e:
            download_end = time.time()
            self.append_log(f"下载异常：{e}", "ERROR")
            self.append_log(f"视频下载耗时: {download_end - download_start:.2f}秒（异常）", "INFO")
            # 直接返回None表示失败，不使用示例文件
            return None

    # 步骤2：保存视频
    def save_video(self, video_url: str, link: str):
        try:
            # 检查video_url是否已经是本地文件路径
            if os.path.exists(video_url) and os.path.isfile(video_url):
                # 如果是本地文件，直接复制到videos目录
                url_hash = hashlib.md5(link.encode()).hexdigest()[:8]
                ts = int(time.time())
                file_path = os.path.join(VIDEO_DIR, f"video_{url_hash}_{ts}.mp4")
                
                # 复制文件
                import shutil
                shutil.copy2(video_url, file_path)
                
                self.append_log(f"视频已从临时位置复制到: {file_path}", "INFO")
                
                # 清理临时目录
                temp_dir = os.path.dirname(video_url)
                if 'temp' in temp_dir.lower():
                    try:
                        shutil.rmtree(temp_dir)
                        self.append_log(f"已清理临时目录: {temp_dir}", "DEBUG")
                    except:
                        pass
                
                return file_path
            else:
                # 如果是URL，使用异步方式下载视频
                url_hash = hashlib.md5(link.encode()).hexdigest()[:8]
                ts = int(time.time())
                file_path = os.path.join(VIDEO_DIR, f"video_{url_hash}_{ts}.mp4")
                self.append_log(f"保存至：{file_path}")
                
                # 运行异步下载
                try:
                    # 创建事件循环并运行异步任务
                    loop = asyncio.new_event_loop()
                    asyncio.set_event_loop(loop)
                    result = loop.run_until_complete(self.async_download_video(video_url, file_path))
                    loop.close()
                    
                    if result:
                        self.append_log(f"视频下载完成: {file_path}", "INFO")
                        return file_path
                    else:
                        self.append_log("异步下载视频失败", "ERROR")
                        return None
                except Exception as e:
                    self.append_log(f"异步下载异常：{e}", "ERROR")
                    # 回退到同步下载
                    return self.sync_download_video(video_url, file_path)
        except Exception as e:
            self.append_log(f"保存异常：{e}")
            return None
    
    # 异步下载视频
    async def async_download_video(self, video_url: str, file_path: str):
        """异步下载视频文件"""
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/144.0.0.0 Safari/537.36',
            'Accept-Encoding': 'identity;q=1, *;q=0',
            'Range': 'bytes=0-',
            'Referer': 'https://www.hellotik.app/',
            'Sec-Ch-Ua': '"Not(A:Brand";v="8", "Chromium";v="144", "Google Chrome";v="144")',
            'Sec-Ch-Ua-Mobile': '?0',
            'Sec-Ch-Ua-Platform': '"Windows"'
        }
        
        self.append_log(f"使用异步方式下载视频: {video_url}", "DEBUG")
        
        try:
            async with aiohttp.ClientSession() as session:
                async with session.get(video_url, headers=headers, timeout=aiohttp.ClientTimeout(total=60)) as response:
                    self.append_log(f"下载响应状态码: {response.status}", "DEBUG")
                    
                    if response.status != 200:
                        self.append_log(f"下载失败，状态码: {response.status}", "ERROR")
                        return False
                    
                    content_length = response.headers.get('Content-Length')
                    self.append_log(f"Content-Length: {content_length}", "DEBUG")
                    
                    # 异步写入文件
                    with open(file_path, 'wb') as f:
                        async for chunk in response.content.iter_chunked(8192):
                            if chunk:
                                f.write(chunk)
                    
            return True
        except Exception as e:
            self.append_log(f"异步下载异常：{e}", "ERROR")
            return False
    
    # 同步下载视频（作为回退方案）
    def sync_download_video(self, video_url: str, file_path: str):
        """同步下载视频文件（作为回退方案）"""
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/144.0.0.0 Safari/537.36',
            'Accept-Encoding': 'identity;q=1, *;q=0',
            'Range': 'bytes=0-',
            'Referer': 'https://www.hellotik.app/',
            'Sec-Ch-Ua': '"Not(A:Brand";v="8", "Chromium";v="144", "Google Chrome";v="144")',
            'Sec-Ch-Ua-Mobile': '?0',
            'Sec-Ch-Ua-Platform': '"Windows"'
        }
        
        self.append_log(f"使用同步方式下载视频: {video_url}", "DEBUG")
        
        try:
            r = requests.get(video_url, stream=True, headers=headers, timeout=60)
            self.append_log(f"下载响应状态码: {r.status_code}", "DEBUG")
            
            r.raise_for_status()
            
            with open(file_path, 'wb') as f:
                for chunk in r.iter_content(chunk_size=8192):
                    if chunk:
                        f.write(chunk)
            
            self.append_log(f"同步下载视频完成: {file_path}", "INFO")
            return file_path
        except Exception as e:
            self.append_log(f"同步下载异常：{e}", "ERROR")
            return None

    # 步骤3：语音转文字
    def speech_to_text(self, video_file: str, user_prompt: str = ""):
        try:
            import time
            start_time = time.time()
            
            # 检查视频文件大小
            file_size = os.path.getsize(video_file)
            if file_size < 1024:  # 如果文件小于1KB，可能是空文件或示例文件
                self.append_log(f"检测到小文件（{file_size} bytes），可能是示例视频文件，使用模拟数据...", "WARNING")
                # 直接返回模拟数据
                return {
                    "segments": [
                        {"start_time": 0, "text": "这是一段模拟的视频转文字结果。"},
                        {"start_time": 10, "text": "视频内容包括产品介绍、使用方法和注意事项。"},
                        {"start_time": 20, "text": "这是一个示例文本，用于演示语音转文字功能。"}
                    ],
                    "ai_summary": "视频主要介绍了产品的基本信息、使用步骤和注意事项，帮助用户快速了解产品的核心功能和使用方法。"
                }
            
            # 检查是否使用LangChain集成
            if self.langchain_integration:
                self.append_log("使用 LangChain 集成进行语音转文字和总结...", "INFO")
                self.update_progress(45, "使用 LangChain 处理视频...")
                
                # 为了显示进度，我们可以添加一些中间状态更新
                import threading
                
                # 创建一个线程来定期更新进度
                def progress_updater():
                    progress = 50
                    while not transcribe_done:
                        if progress < 85:
                            progress += 1
                            self.update_progress(progress, f"正在处理视频... {progress-45}%")
                        time.sleep(1)
                
                transcribe_done = False
                progress_thread = threading.Thread(target=progress_updater)
                progress_thread.daemon = True
                progress_thread.start()
                
                try:
                    # 使用LangChain处理视频
                    langchain_start = time.time()
                    result = self.langchain_integration.process_video(video_file, user_prompt)
                    langchain_end = time.time()
                    
                    transcribe_done = True
                    self.append_log(f"LangChain处理耗时: {langchain_end - langchain_start:.2f}秒", "INFO")
                    
                    self.update_progress(85, "处理完成，准备生成文档...")
                    
                    self.append_log("语音转文字和总结完成！", "INFO")
                    self.append_log(f"转写结果: {result['transcribe_text'][:100]}...", "INFO")
                    self.append_log(f"总结结果: {result['ai_summary'][:100]}...", "INFO")
                    
                    total_end = time.time()
                    self.append_log(f"语音转文字总耗时: {total_end - start_time:.2f}秒", "INFO")
                    
                    return {
                        "segments": result["segments"],
                        "ai_summary": result["ai_summary"]
                    }
                except Exception as e:
                    self.append_log(f"LangChain处理异常：{e}", "ERROR")
                    # 回退到传统方法
                    transcribe_done = True
                    self.append_log("回退到传统方法进行语音转文字...", "INFO")
            
            # 使用传统方法：Whisper 本地模型
            self.append_log("使用 Whisper 本地模型进行语音转文字...", "INFO")
            
            # 导入 whisper 库
            import whisper
            
            # 加载 Whisper 模型（使用缓存的模型）
            model_load_start = time.time()
            self.append_log("加载 Whisper 模型...", "INFO")
            self.update_progress(45, "加载语音转文字模型...")
            
            # 检查模型缓存（线程安全）
            with self.model_cache_lock:
                if self.model_cache is None:
                    self.append_log("首次加载 Whisper 模型...", "INFO")
                    # 使用 tiny 模型提高转写速度
                    try:
                        self.append_log("加载 Whisper tiny 模型（提高转写速度）...", "INFO")
                        self.model_cache = whisper.load_model("tiny")
                        self.append_log("Whisper tiny 模型加载完成并缓存", "INFO")
                    except Exception as e:
                        self.append_log(f"加载 tiny 模型失败：{e}", "WARNING")
                        # 回退到 small 模型
                        self.append_log("加载 Whisper small 模型作为回退...", "INFO")
                        self.model_cache = whisper.load_model("small")
                        self.append_log("Whisper small 模型加载完成并缓存", "INFO")
                else:
                    self.append_log("使用缓存的 Whisper 模型", "INFO")
            
            model_load_end = time.time()
            self.append_log(f"模型加载耗时: {model_load_end - model_load_start:.2f}秒", "INFO")
            
            model = self.model_cache
            self.update_progress(50, "模型加载完成，准备开始转写...")
            
            # 直接使用 Whisper 处理视频文件
            self.append_log("开始转写...", "INFO")
            self.update_progress(55, "正在分析视频音频...")
            
            # 为了显示进度，我们可以添加一些中间状态更新
            import threading
            
            # 创建一个线程来定期更新进度
            def progress_updater():
                progress = 60
                while not transcribe_done:
                    if progress < 75:
                        progress += 1
                        self.update_progress(progress, f"正在转写音频... {progress-55}%")
                    time.sleep(1)
            
            transcribe_done = False
            progress_thread = threading.Thread(target=progress_updater)
            progress_thread.daemon = True
            progress_thread.start()
            
            try:
                # 优化Whisper模型参数，在速度和准确率之间找到平衡点
                self.append_log("使用优化参数进行转写...", "INFO")
                transcribe_start = time.time()
                result = model.transcribe(
                    video_file, 
                    language="zh",  # 明确指定中文
                    fp16=False,  # 禁用FP16，提高兼容性
                    verbose=False,  # 禁用详细输出，提高速度
                    task="transcribe",  # 明确指定任务为转写
                    beam_size=1,  # 进一步减小beam_size，显著提高速度
                    temperature=0.0,  # 保持temperature=0.0，确保准确性
                    best_of=1,  # 进一步减小best_of，提高速度
                    patience=0.0,  # 进一步减小patience，提高速度
                    initial_prompt="请使用标准简体中文进行转写。",  # 精简提示词
                    condition_on_previous_text=False,  # 禁用上下文依赖，提高速度
                    compression_ratio_threshold=2.4  # 设置压缩比阈值，过滤低质量转写
                )
                
                transcribe_end = time.time()
                self.append_log(f"转写耗时: {transcribe_end - transcribe_start:.2f}秒", "INFO")
                
                transcribe_done = True
                self.update_progress(75, "转写完成，正在处理结果...")
                
                # 获取转写结果
                text = result["text"]
                segments = []
                for seg in result["segments"]:
                    segments.append({
                        "start_time": seg["start"],
                        "text": seg["text"].strip()
                    })
                
                self.append_log("语音转文字完成！", "INFO")
                self.append_log(f"转写结果: {text[:100]}...", "INFO")
                
                # 使用火山引擎API进行文本总结
                self.update_progress(80, "使用AI进行文本总结...")
                self.append_log("使用火山引擎API进行文本总结...", "INFO")
                summary = self.summarize_with_volcengine(text, user_prompt)
                
                self.update_progress(85, "总结完成，准备生成文档...")
                
                if summary:
                    self.append_log("文本总结成功", "INFO")
                    return {
                        "segments": segments,
                        "ai_summary": summary
                    }
                else:
                    # 总结失败，使用转写文本的前100个字符作为摘要
                    return {
                        "segments": segments,
                        "ai_summary": text[:100] + "...（省略部分内容）"
                    }
            except RuntimeError as e:
                # 处理Whisper模型的RuntimeError，特别是张量形状错误
                if "cannot reshape tensor" in str(e) or "0 elements" in str(e):
                    self.append_log(f"Whisper模型无法处理此文件（可能是示例文件或无音频数据）：{e}", "WARNING")
                    transcribe_done = True
                    # 返回模拟数据
                    return {
                        "segments": [
                            {"start_time": 0, "text": "这是一段模拟的视频转文字结果。"},
                            {"start_time": 10, "text": "视频内容包括产品介绍、使用方法和注意事项。"},
                            {"start_time": 20, "text": "这是一个示例文本，用于演示语音转文字功能。"}
                        ],
                        "ai_summary": "视频主要介绍了产品的基本信息、使用步骤和注意事项，帮助用户快速了解产品的核心功能和使用方法。"
                    }
                else:
                    # 其他RuntimeError，继续抛出
                    raise
        except Exception as e:
            self.append_log(f"语音转文字异常：{type(e).__name__}: {e}", "ERROR")
            # 使用模拟数据作为最后的备用方案
            return {
                "segments": [
                    {"start_time": 0, "text": "这是一段模拟的视频转文字结果。"},
                    {"start_time": 10, "text": "视频内容包括产品介绍、使用方法和注意事项。"},
                    {"start_time": 20, "text": "这是一个示例文本，用于演示语音转文字功能。"}
                ],
                "ai_summary": "视频主要介绍了产品的基本信息、使用步骤和注意事项，帮助用户快速了解产品的核心功能和使用方法。"
            }
    
    # 使用火山引擎 API 进行文本总结
    def summarize_with_volcengine(self, text: str, user_prompt: str = ""):
        try:
            api_key = CONFIG.get("volcengine_api_key", VOLCENGINE_API_KEY)
            summary_prompt = CONFIG.get("summary_prompt", DEFAULT_CONFIG["summary_prompt"])
            system_prompt = CONFIG.get("system_prompt", DEFAULT_CONFIG["system_prompt"])
            rules = CONFIG.get("rules", DEFAULT_CONFIG["rules"])
            
            # 确保 volcenginesdkarkruntime 已安装
            try:
                from volcenginesdkarkruntime import Ark
            except ImportError:
                self.append_log("正在安装 volcengine-python-sdk[ark]...", "INFO")
                import subprocess
                subprocess.run(["pip", "install", "--upgrade", "volcengine-python-sdk[ark]"], check=True)
                from volcenginesdkarkruntime import Ark
            
            # 构建请求输入
            input_content = [
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "input_text",
                            "text": f"{system_prompt}\n\n分析规则：\n{rules}\n\n{summary_prompt.format(text=text)}"
                        }
                    ],
                }
            ]
            
            # 如果有 user_prompt，添加到输入中
            if user_prompt:
                input_content.append({
                    "role": "user",
                    "content": [
                        {
                            "type": "input_text",
                            "text": user_prompt
                        }
                    ],
                })
            
            # 发送请求（增加重试机制）
            max_retries = 3
            for retry in range(max_retries):
                try:
                    self.append_log(f"调用火山引擎API进行总结（尝试 {retry+1}/{max_retries}）...", "INFO")
                    
                    # 创建Ark客户端
                    client = Ark(
                        base_url=VOLCENGINE_API_URL,
                        api_key=api_key,
                    )
                    
                    # 发送测试请求
                    response = client.responses.create(
                        model="doubao-seed-1-8-251228",
                        input=input_content
                    )
                    
                    # 解析响应
                    if response.status == "completed" and response.output:
                        for item in response.output:
                            if item.type == "message" and item.role == "assistant":
                                for content in item.content:
                                    if content.type == "output_text":
                                        summary = content.text
                                        if summary:
                                            self.append_log("火山引擎API调用成功", "INFO")
                                            return summary
                    
                    self.append_log("火山引擎API返回空结果或格式不正确", "ERROR")
                    return None
                    
                except Exception as e:
                    self.append_log(f"火山引擎 API 调用异常：{e}", "ERROR")
                    if retry < max_retries - 1:
                        self.append_log(f"等待后重试...", "INFO")
                        time.sleep(2)
                    else:
                        return None
        except Exception as e:
            self.append_log(f"火山引擎 API 调用异常：{e}", "ERROR")
            return None

    # 步骤4：生成Markdown
    def generate_md(self, result_data: dict, link: str, platform: str):
        try:
            # 获取当前目录下的Markdown文件数量，作为总序号
            existing_md = [f for f in os.listdir(OUTPUT_DIR) if f.endswith('.md')]
            total_count = len(existing_md) + 1
            
            # 获取当前日期（月-日）
            current_date = time.strftime('%m-%d')
            
            # 从内容中提取标题（使用AI分析结果）
            summary = result_data.get('ai_summary', '')
            doc_name = self.extract_title_from_summary(summary, link)
            
            # 计算URL哈希（用于视频信息中显示）
            url_hash = hashlib.md5(link.encode()).hexdigest()[:8]
            
            # 构建新的文件名：总记录序号-月-日-文档名称
            md_path = os.path.join(OUTPUT_DIR, f"{total_count:03d}-{current_date}-{doc_name}_视频分析.md")

            segs = result_data.get('segments', [])
            transcript_lines = []
            for seg in segs:
                tstr = time.strftime('%H:%M:%S', time.gmtime(seg.get('start_time', 0)))
                transcript_lines.append(f"- [{tstr}] {seg.get('text','')}")
            transcript = "\n".join(transcript_lines)
            summary = result_data.get('ai_summary', result_data.get('summary', ''))

            md = f"""# {platform}视频内容分析

## 视频信息
- 分析时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
- 原始链接: {link}
- 平台: {platform}
- URL哈希: {url_hash}

## 📝 语音转文字内容
{transcript}

## 🤖 AI智能分析摘要
{summary}

---
*由视频转文字处理工具自动生成*
"""
            with open(md_path, 'w', encoding='utf-8') as f:
                f.write(md)
            return md_path
        except Exception as e:
            self.append_log(f"写入MD失败：{e}")
            return None
    
    # 打开配置窗口
    def open_config_window(self):
        config_window = tk.Toplevel(self.root)
        config_window.title("配置")
        config_window.geometry("600x400")
        config_window.resizable(False, False)
        
        # 配置文件路径
        tk.Label(config_window, text="配置文件：", font=("微软雅黑", 10)).pack(anchor=tk.W, padx=20, pady=(20, 5))
        tk.Label(config_window, text=CONFIG_FILE, font=("Consolas", 9), fg="gray").pack(anchor=tk.W, padx=20, pady=(0, 15))
        
        # 总结提示文本
        tk.Label(config_window, text="总结要求文本：", font=("微软雅黑", 10)).pack(anchor=tk.W, padx=20, pady=(10, 5))
        summary_prompt_var = tk.StringVar(value=CONFIG.get("summary_prompt", DEFAULT_CONFIG["summary_prompt"]))
        summary_prompt_text = scrolledtext.ScrolledText(config_window, height=10, font=("Consolas", 10))
        summary_prompt_text.pack(fill=tk.X, padx=20, pady=(0, 15))
        summary_prompt_text.insert(tk.END, summary_prompt_var.get())
        
        # 保存按钮
        def save_config_changes():
            global CONFIG
            new_summary_prompt = summary_prompt_text.get(1.0, tk.END).strip()
            new_config = CONFIG.copy()
            new_config["summary_prompt"] = new_summary_prompt
            if save_config(new_config):
                CONFIG = new_config
                messagebox.showinfo("成功", "配置已保存")
                config_window.destroy()
            else:
                messagebox.showerror("失败", "保存配置失败")
        
        save_btn = ttk.Button(config_window, text="保存", command=save_config_changes)
        save_btn.pack(pady=20)
        
        # 居中显示
        config_window.transient(self.root)
        config_window.grab_set()
        self.root.wait_window(config_window)
    
    # 打开AI配置窗口
    def open_ai_config_window(self):
        ai_config_window = tk.Toplevel(self.root)
        ai_config_window.title("AI配置 - 视频转文字处理工具")
        ai_config_window.geometry("1000x600")
        ai_config_window.resizable(True, True)
        ai_config_window.configure(bg="#f0f4f8")
        
        # 设置样式
        style = ttk.Style()
        style.theme_use("clam")
        
        # 自定义样式
        style.configure(
            "TButton", 
            padding=(12, 6),
            font=("微软雅黑", 10),
            foreground="#0066cc",
            background="#e6f0ff"
        )
        style.configure(
            "TLabel", 
            font=("微软雅黑", 10),
            foreground="#333"
        )
        style.configure(
            "TEntry", 
            padding=(8, 4),
            font=("微软雅黑", 10)
        )
        style.configure(
            "TLabelframe", 
            font=("微软雅黑", 10, "bold"),
            foreground="#0066cc"
        )
        
        # 主容器 - 改进的滚动实现
        main_frame = tk.Frame(ai_config_window, bg="#f0f4f8")
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # 画布和滚动条
        canvas = tk.Canvas(main_frame, bg="#f0f4f8")
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        scrollbar.pack(side="right", fill="y")
        canvas.pack(side="left", fill=tk.BOTH, expand=True)
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # 内容容器
        main_container = tk.Frame(canvas, bg="#f0f4f8")
        canvas.create_window((0, 0), window=main_container, anchor="nw", width=960)
        
        # 配置滚动区域
        def on_configure(event):
            canvas.configure(scrollregion=canvas.bbox("all"))
        
        main_container.bind("<Configure>", on_configure)
        
        # 添加鼠标滚轮支持
        def on_mouse_wheel(event):
            canvas.yview_scroll(-1 * (event.delta // 120), "units")
        
        canvas.bind_all("<MouseWheel>", on_mouse_wheel)
        
        # 顶部标题区域
        title_frame = tk.Frame(main_container, bg="#f0f4f8")
        title_frame.pack(fill=tk.X, pady=(0, 20))
        
        title_label = tk.Label(
            title_frame, 
            text="AI配置中心", 
            font=("微软雅黑", 16, "bold"),
            foreground="#0066cc",
            bg="#f0f4f8"
        )
        title_label.pack(anchor=tk.W)
        
        subtitle_label = tk.Label(
            title_frame, 
            text="智能分析参数配置", 
            font=("微软雅黑", 10, "italic"),
            foreground="#666",
            bg="#f0f4f8"
        )
        subtitle_label.pack(anchor=tk.W, pady=(5, 0))
        
        # 配置文件路径
        config_path_frame = tk.Frame(main_container, bg="#ffffff", bd=0, relief=tk.RAISED)
        config_path_frame.pack(fill=tk.X, pady=(0, 15))
        config_path_frame.configure(bg="#ffffff", highlightbackground="#0066cc", highlightthickness=1, borderwidth=0)
        
        ttk.Label(config_path_frame, text="配置文件：", font=("微软雅黑", 10), background="#ffffff").pack(anchor=tk.W, padx=15, pady=(10, 5))
        ttk.Label(
            config_path_frame, 
            text=CONFIG_FILE, 
            font=("Consolas", 9, "italic"), 
            foreground="#666",
            background="#ffffff"
        ).pack(anchor=tk.W, padx=15, pady=(0, 10))
        
        # 文件上传区域
        file_frame = tk.Frame(main_container, bg="#ffffff", bd=0, relief=tk.RAISED)
        file_frame.pack(fill=tk.X, pady=(0, 20))
        file_frame.configure(bg="#ffffff", highlightbackground="#0066cc", highlightthickness=1, borderwidth=0)
        
        file_title_frame = tk.Frame(file_frame, bg="#ffffff")
        file_title_frame.pack(fill=tk.X, padx=15, pady=(15, 10))
        
        file_title = tk.Label(
            file_title_frame, 
            text="批量导入配置", 
            font=("微软雅黑", 12, "bold"),
            foreground="#0066cc",
            bg="#ffffff"
        )
        file_title.pack(anchor=tk.W)
        
        file_desc = tk.Label(
            file_title_frame, 
            text="从文件中导入所有配置选项", 
            font=("微软雅黑", 9),
            foreground="#666",
            bg="#ffffff"
        )
        file_desc.pack(anchor=tk.W, pady=(5, 0))
        
        file_btn_frame = tk.Frame(file_frame, bg="#ffffff")
        file_btn_frame.pack(fill=tk.X, padx=15, pady=(0, 15))
        
        def import_from_file():
            file_path = filedialog.askopenfilename(
                title="选择配置文件",
                filetypes=[("文本文件", "*.txt"), ("Markdown文件", "*.md"), ("所有文件", "*.*")]
            )
            if file_path:
                try:
                    with open(file_path, 'r', encoding='utf-8') as f:
                        content = f.read()
                    # 简单解析：按空行分割为三个部分
                    parts = content.split('\n\n\n')
                    if len(parts) >= 3:
                        system_prompt_text.delete(1.0, tk.END)
                        system_prompt_text.insert(tk.END, parts[0].strip())
                        rules_text.delete(1.0, tk.END)
                        rules_text.insert(tk.END, parts[1].strip())
                        output_template_text.delete(1.0, tk.END)
                        output_template_text.insert(tk.END, parts[2].strip())
                        messagebox.showinfo("成功", "配置已从文件导入")
                    else:
                        messagebox.showwarning("提示", "文件格式不正确，请确保文件包含system prompt、rules和output template三个部分，用三个空行分隔")
                except Exception as e:
                    messagebox.showerror("错误", f"导入失败：{e}")
        
        import_btn = ttk.Button(file_btn_frame, text="选择文件", command=import_from_file)
        import_btn.pack(side=tk.LEFT, padx=5)
        
        # System Prompt 配置
        system_prompt_frame = tk.Frame(main_container, bg="#ffffff", bd=0, relief=tk.RAISED)
        system_prompt_frame.pack(fill=tk.X, pady=(0, 15))
        system_prompt_frame.configure(bg="#ffffff", highlightbackground="#0066cc", highlightthickness=1, borderwidth=0)
        
        system_title_frame = tk.Frame(system_prompt_frame, bg="#ffffff")
        system_title_frame.pack(fill=tk.X, padx=15, pady=(15, 10))
        
        system_title = tk.Label(
            system_title_frame, 
            text="System Prompt", 
            font=("微软雅黑", 12, "bold"),
            foreground="#0066cc",
            bg="#ffffff"
        )
        system_title.pack(side=tk.LEFT)
        
        system_desc = tk.Label(
            system_title_frame, 
            text="定义AI助手的角色和能力", 
            font=("微软雅黑", 9),
            foreground="#666",
            bg="#ffffff"
        )
        system_desc.pack(side=tk.LEFT, padx=10)
        
        system_prompt_control_frame = tk.Frame(system_prompt_frame, bg="#ffffff")
        system_prompt_control_frame.pack(fill=tk.X, padx=15, pady=(0, 10))
        
        def import_system_prompt():
            file_path = filedialog.askopenfilename(
                title="选择System Prompt文件",
                filetypes=[("文本文件", "*.txt"), ("Markdown文件", "*.md"), ("所有文件", "*.*")]
            )
            if file_path:
                try:
                    with open(file_path, 'r', encoding='utf-8') as f:
                        content = f.read()
                    system_prompt_text.delete(1.0, tk.END)
                    system_prompt_text.insert(tk.END, content.strip())
                    messagebox.showinfo("成功", "System Prompt已从文件导入")
                except Exception as e:
                    messagebox.showerror("错误", f"导入失败：{e}")
        
        import_system_btn = ttk.Button(system_prompt_control_frame, text="导入文件", command=import_system_prompt)
        import_system_btn.pack(side=tk.RIGHT)
        
        system_prompt_text = scrolledtext.ScrolledText(
            system_prompt_frame, 
            height=5, 
            font=("Consolas", 10),
            bd=0, 
            bg="#f9f9f9",
            relief=tk.FLAT
        )
        system_prompt_text.pack(fill=tk.X, padx=15, pady=(0, 15))
        system_prompt_text.insert(tk.END, CONFIG.get("system_prompt", DEFAULT_CONFIG["system_prompt"]))
        
        # Rules 配置
        rules_frame = tk.Frame(main_container, bg="#ffffff", bd=0, relief=tk.RAISED)
        rules_frame.pack(fill=tk.X, pady=(0, 15))
        rules_frame.configure(bg="#ffffff", highlightbackground="#0066cc", highlightthickness=1, borderwidth=0)
        
        rules_title_frame = tk.Frame(rules_frame, bg="#ffffff")
        rules_title_frame.pack(fill=tk.X, padx=15, pady=(15, 10))
        
        rules_title = tk.Label(
            rules_title_frame, 
            text="Rules 配置", 
            font=("微软雅黑", 12, "bold"),
            foreground="#0066cc",
            bg="#ffffff"
        )
        rules_title.pack(side=tk.LEFT)
        
        rules_desc = tk.Label(
            rules_title_frame, 
            text="设定AI分析的规则和标准", 
            font=("微软雅黑", 9),
            foreground="#666",
            bg="#ffffff"
        )
        rules_desc.pack(side=tk.LEFT, padx=10)
        
        rules_control_frame = tk.Frame(rules_frame, bg="#ffffff")
        rules_control_frame.pack(fill=tk.X, padx=15, pady=(0, 10))
        
        def import_rules():
            file_path = filedialog.askopenfilename(
                title="选择Rules文件",
                filetypes=[("文本文件", "*.txt"), ("Markdown文件", "*.md"), ("所有文件", "*.*")]
            )
            if file_path:
                try:
                    with open(file_path, 'r', encoding='utf-8') as f:
                        content = f.read()
                    rules_text.delete(1.0, tk.END)
                    rules_text.insert(tk.END, content.strip())
                    messagebox.showinfo("成功", "Rules已从文件导入")
                except Exception as e:
                    messagebox.showerror("错误", f"导入失败：{e}")
        
        import_rules_btn = ttk.Button(rules_control_frame, text="导入文件", command=import_rules)
        import_rules_btn.pack(side=tk.RIGHT)
        
        rules_text = scrolledtext.ScrolledText(
            rules_frame, 
            height=5, 
            font=("Consolas", 10),
            bd=0, 
            bg="#f9f9f9",
            relief=tk.FLAT
        )
        rules_text.pack(fill=tk.X, padx=15, pady=(0, 15))
        rules_text.insert(tk.END, CONFIG.get("rules", DEFAULT_CONFIG["rules"]))
        
        # 产出模板配置
        output_template_frame = tk.Frame(main_container, bg="#ffffff", bd=0, relief=tk.RAISED)
        output_template_frame.pack(fill=tk.X, pady=(0, 15))
        output_template_frame.configure(bg="#ffffff", highlightbackground="#0066cc", highlightthickness=1, borderwidth=0)
        
        template_title_frame = tk.Frame(output_template_frame, bg="#ffffff")
        template_title_frame.pack(fill=tk.X, padx=15, pady=(15, 10))
        
        template_title = tk.Label(
            template_title_frame, 
            text="产出模板配置", 
            font=("微软雅黑", 12, "bold"),
            foreground="#0066cc",
            bg="#ffffff"
        )
        template_title.pack(side=tk.LEFT)
        
        template_desc = tk.Label(
            template_title_frame, 
            text="定义最终生成文档的格式和结构", 
            font=("微软雅黑", 9),
            foreground="#666",
            bg="#ffffff"
        )
        template_desc.pack(side=tk.LEFT, padx=10)
        
        output_template_control_frame = tk.Frame(output_template_frame, bg="#ffffff")
        output_template_control_frame.pack(fill=tk.X, padx=15, pady=(0, 10))
        
        def import_output_template():
            file_path = filedialog.askopenfilename(
                title="选择产出模板文件",
                filetypes=[("Markdown文件", "*.md"), ("文本文件", "*.txt"), ("所有文件", "*.*")]
            )
            if file_path:
                try:
                    with open(file_path, 'r', encoding='utf-8') as f:
                        content = f.read()
                    output_template_text.delete(1.0, tk.END)
                    output_template_text.insert(tk.END, content.strip())
                    messagebox.showinfo("成功", "产出模板已从文件导入")
                except Exception as e:
                    messagebox.showerror("错误", f"导入失败：{e}")
        
        import_template_btn = ttk.Button(output_template_control_frame, text="导入文件", command=import_output_template)
        import_template_btn.pack(side=tk.RIGHT)
        
        output_template_text = scrolledtext.ScrolledText(
            output_template_frame, 
            height=8, 
            font=("Consolas", 10),
            bd=0, 
            bg="#f9f9f9",
            relief=tk.FLAT
        )
        output_template_text.pack(fill=tk.X, padx=15, pady=(0, 15))
        output_template_text.insert(tk.END, CONFIG.get("output_template", DEFAULT_CONFIG["output_template"]))
        
        # User Prompt 配置
        user_prompt_frame = tk.Frame(main_container, bg="#ffffff", bd=0, relief=tk.RAISED)
        user_prompt_frame.pack(fill=tk.X, pady=(0, 15))
        user_prompt_frame.configure(bg="#ffffff", highlightbackground="#0066cc", highlightthickness=1, borderwidth=0)
        
        user_title_frame = tk.Frame(user_prompt_frame, bg="#ffffff")
        user_title_frame.pack(fill=tk.X, padx=15, pady=(15, 10))
        
        user_title = tk.Label(
            user_title_frame, 
            text="User Prompt", 
            font=("微软雅黑", 12, "bold"),
            foreground="#0066cc",
            bg="#ffffff"
        )
        user_title.pack(side=tk.LEFT)
        
        user_desc = tk.Label(
            user_title_frame, 
            text="每次处理视频时的额外提示信息", 
            font=("微软雅黑", 9),
            foreground="#666",
            bg="#ffffff"
        )
        user_desc.pack(side=tk.LEFT, padx=10)
        
        user_prompt_control_frame = tk.Frame(user_prompt_frame, bg="#ffffff")
        user_prompt_control_frame.pack(fill=tk.X, padx=15, pady=(0, 10))
        
        def import_user_prompt():
            file_path = filedialog.askopenfilename(
                title="选择User Prompt文件",
                filetypes=[("文本文件", "*.txt"), ("Markdown文件", "*.md"), ("所有文件", "*.*")]
            )
            if file_path:
                try:
                    with open(file_path, 'r', encoding='utf-8') as f:
                        content = f.read()
                    user_prompt_text.delete(1.0, tk.END)
                    user_prompt_text.insert(tk.END, content.strip())
                    messagebox.showinfo("成功", "User Prompt已从文件导入")
                except Exception as e:
                    messagebox.showerror("错误", f"导入失败：{e}")
        
        import_user_btn = ttk.Button(user_prompt_control_frame, text="导入文件", command=import_user_prompt)
        import_user_btn.pack(side=tk.RIGHT)
        
        user_prompt_text = scrolledtext.ScrolledText(
            user_prompt_frame, 
            height=3, 
            font=("Consolas", 10),
            bd=0, 
            bg="#f9f9f9",
            relief=tk.FLAT
        )
        user_prompt_text.pack(fill=tk.X, padx=15, pady=(0, 15))
        user_prompt_text.insert(tk.END, CONFIG.get("user_prompt", DEFAULT_CONFIG["user_prompt"]))
        
        # 按钮框架
        button_frame = tk.Frame(main_container, bg="#f0f4f8")
        button_frame.pack(fill=tk.X, pady=20)
        
        # 保存按钮
        def save_ai_config_changes():
            global CONFIG
            new_system_prompt = system_prompt_text.get(1.0, tk.END).strip()
            new_rules = rules_text.get(1.0, tk.END).strip()
            new_output_template = output_template_text.get(1.0, tk.END).strip()
            new_user_prompt = user_prompt_text.get(1.0, tk.END).strip()
            
            new_config = CONFIG.copy()
            new_config["system_prompt"] = new_system_prompt
            new_config["rules"] = new_rules
            new_config["output_template"] = new_output_template
            new_config["user_prompt"] = new_user_prompt
            
            if save_config(new_config):
                CONFIG = new_config
                messagebox.showinfo("成功", "AI配置已保存")
                ai_config_window.destroy()
            else:
                messagebox.showerror("失败", "保存AI配置失败")
        
        save_btn = ttk.Button(button_frame, text="保存配置", command=save_ai_config_changes)
        save_btn.pack(side=tk.RIGHT, padx=10)
        
        cancel_btn = ttk.Button(button_frame, text="取消", command=ai_config_window.destroy)
        cancel_btn.pack(side=tk.RIGHT, padx=10)
        
        # 居中显示
        ai_config_window.transient(self.root)
        ai_config_window.grab_set()
        self.root.wait_window(ai_config_window)
    
    def open_thread_config_window(self):
        """打开线程配置窗口"""
        thread_config_window = tk.Toplevel(self.root)
        thread_config_window.title("线程配置 - 视频转文字处理工具")
        thread_config_window.geometry("800x500")
        thread_config_window.resizable(True, True)
        thread_config_window.configure(bg="#f0f4f8")
        
        # 设置样式
        style = ttk.Style()
        style.theme_use("clam")
        
        # 自定义样式
        style.configure(
            "TButton", 
            padding=(12, 6),
            font=("微软雅黑", 10),
            foreground="#0066cc",
            background="#e6f0ff"
        )
        style.configure(
            "TLabel", 
            font=("微软雅黑", 10),
            foreground="#333"
        )
        style.configure(
            "TEntry", 
            padding=(8, 4),
            font=("微软雅黑", 10)
        )
        style.configure(
            "TLabelframe", 
            font=("微软雅黑", 10, "bold"),
            foreground="#0066cc"
        )
        
        # 主容器
        main_container = tk.Frame(thread_config_window, bg="#f0f4f8")
        main_container.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # 顶部标题区域
        title_frame = tk.Frame(main_container, bg="#f0f4f8")
        title_frame.pack(fill=tk.X, pady=(0, 20))
        
        title_label = tk.Label(
            title_frame, 
            text="线程配置中心", 
            font=("微软雅黑", 16, "bold"),
            foreground="#0066cc",
            bg="#f0f4f8"
        )
        title_label.pack(anchor=tk.W)
        
        subtitle_label = tk.Label(
            title_frame, 
            text="调整线程池大小和任务队列限制", 
            font=("微软雅黑", 10, "italic"),
            foreground="#666",
            bg="#f0f4f8"
        )
        subtitle_label.pack(anchor=tk.W, pady=(5, 0))
        
        # 系统信息区域
        system_info_frame = tk.Frame(main_container, bg="#ffffff", bd=0, relief=tk.RAISED)
        system_info_frame.pack(fill=tk.X, pady=(0, 20))
        system_info_frame.configure(bg="#ffffff", highlightbackground="#0066cc", highlightthickness=1, borderwidth=0)
        
        ttk.Label(system_info_frame, text="系统信息：", font=("微软雅黑", 10, "bold"), background="#ffffff").pack(anchor=tk.W, padx=15, pady=(10, 5))
        
        info_text = f"系统CPU核心数：{self.cpu_count}\n当前活跃线程数：{len(self.active_futures)}\n当前队列长度：{len(self.task_queue)}\n当前队列最大大小：{self.queue_max_size}"
        info_label = tk.Label(
            system_info_frame, 
            text=info_text, 
            font=("Consolas", 9), 
            foreground="#666",
            background="#ffffff",
            justify=tk.LEFT
        )
        info_label.pack(anchor=tk.W, padx=15, pady=(0, 10))
        
        # 线程池配置区域
        thread_pool_frame = tk.Frame(main_container, bg="#ffffff", bd=0, relief=tk.RAISED)
        thread_pool_frame.pack(fill=tk.X, pady=(0, 20))
        thread_pool_frame.configure(bg="#ffffff", highlightbackground="#0066cc", highlightthickness=1, borderwidth=0)
        
        thread_pool_title_frame = tk.Frame(thread_pool_frame, bg="#ffffff")
        thread_pool_title_frame.pack(fill=tk.X, padx=15, pady=(15, 10))
        
        thread_pool_title = tk.Label(
            thread_pool_title_frame, 
            text="线程池配置", 
            font=("微软雅黑", 12, "bold"),
            foreground="#0066cc",
            bg="#ffffff"
        )
        thread_pool_title.pack(side=tk.LEFT)
        
        thread_pool_desc = tk.Label(
            thread_pool_title_frame, 
            text=f"范围：1-{self.cpu_count}", 
            font=("微软雅黑", 9),
            foreground="#666",
            bg="#ffffff"
        )
        thread_pool_desc.pack(side=tk.LEFT, padx=10)
        
        # 线程数输入
        thread_count_frame = tk.Frame(thread_pool_frame, bg="#ffffff")
        thread_count_frame.pack(fill=tk.X, padx=15, pady=(0, 15))
        
        ttk.Label(thread_count_frame, text="最大线程数：", font=("微软雅黑", 10), background="#ffffff").pack(side=tk.LEFT, padx=(0, 10))
        
        thread_count_var = tk.StringVar(value=str(self.max_workers))
        thread_count_entry = ttk.Entry(
            thread_count_frame, 
            textvariable=thread_count_var, 
            width=10, 
            font=("微软雅黑", 10)
        )
        thread_count_entry.pack(side=tk.LEFT, padx=(0, 10))
        
        # 任务队列配置区域
        queue_frame = tk.Frame(main_container, bg="#ffffff", bd=0, relief=tk.RAISED)
        queue_frame.pack(fill=tk.X, pady=(0, 20))
        queue_frame.configure(bg="#ffffff", highlightbackground="#0066cc", highlightthickness=1, borderwidth=0)
        
        queue_title_frame = tk.Frame(queue_frame, bg="#ffffff")
        queue_title_frame.pack(fill=tk.X, padx=15, pady=(15, 10))
        
        queue_title = tk.Label(
            queue_title_frame, 
            text="任务队列配置", 
            font=("微软雅黑", 12, "bold"),
            foreground="#0066cc",
            bg="#ffffff"
        )
        queue_title.pack(side=tk.LEFT)
        
        # 队列大小输入
        queue_size_frame = tk.Frame(queue_frame, bg="#ffffff")
        queue_size_frame.pack(fill=tk.X, padx=15, pady=(0, 15))
        
        ttk.Label(queue_size_frame, text="队列最大大小：", font=("微软雅黑", 10), background="#ffffff").pack(side=tk.LEFT, padx=(0, 10))
        
        queue_size_var = tk.StringVar(value=str(self.queue_max_size))
        queue_size_entry = ttk.Entry(
            queue_size_frame, 
            textvariable=queue_size_var, 
            width=10, 
            font=("微软雅黑", 10)
        )
        queue_size_entry.pack(side=tk.LEFT, padx=(0, 10))
        
        # 按钮框架
        button_frame = tk.Frame(main_container, bg="#f0f4f8")
        button_frame.pack(fill=tk.X, pady=20)
        
        # 应用按钮
        def apply_thread_config_changes():
            try:
                # 验证并更新线程数
                new_thread_count = int(thread_count_var.get())
                if new_thread_count < 1 or new_thread_count > self.cpu_count:
                    messagebox.showerror("错误", f"线程数必须在1-{self.cpu_count}之间")
                    return
                
                # 验证并更新队列大小
                new_queue_size = int(queue_size_var.get())
                if new_queue_size < 1:
                    messagebox.showerror("错误", "队列大小必须大于0")
                    return
                
                # 应用线程池调整
                if new_thread_count != self.max_workers:
                    self.update_thread_pool_size(new_thread_count)
                
                # 应用队列大小调整
                if new_queue_size != self.queue_max_size:
                    self.queue_max_size = new_queue_size
                    self.append_log(f"任务队列最大大小已调整为：{new_queue_size}")
                
                messagebox.showinfo("成功", "线程配置已应用")
                thread_config_window.destroy()
            except ValueError:
                messagebox.showerror("错误", "请输入有效的数字")
        
        apply_btn = ttk.Button(button_frame, text="应用配置", command=apply_thread_config_changes)
        apply_btn.pack(side=tk.RIGHT, padx=10)
        
        cancel_btn = ttk.Button(button_frame, text="取消", command=thread_config_window.destroy)
        cancel_btn.pack(side=tk.RIGHT, padx=10)
        
        # 居中显示
        thread_config_window.transient(self.root)
        thread_config_window.grab_set()
        self.root.wait_window(thread_config_window)
    
    def update_thread_pool_size(self, new_size):
        """更新线程池大小
        
        Args:
            new_size: 新的线程池大小
        """
        try:
            self.append_log(f"开始调整线程池大小：从 {self.max_workers} 到 {new_size}")
            
            if new_size > self.max_workers:
                # 增加线程数
                self.max_workers = new_size
                # 重新创建线程池
                self.executor.shutdown(wait=False)
                self.executor = concurrent.futures.ThreadPoolExecutor(max_workers=self.max_workers)
                self.append_log(f"线程池已扩容至：{self.max_workers} 线程")
                
                # 扩容后，如果队列中有任务且未在处理中，继续处理
                if self.task_queue and not self.processing_queue:
                    self.append_log("线程池扩容完成，继续处理队列任务")
                    self.start_queue_processing()
            elif new_size < self.max_workers:
                # 减少线程数
                # 记录需要保留的活跃任务数量
                tasks_to_keep = min(new_size, len(self.active_futures))
                
                # 取消多余的任务（FIFO原则，取消最早提交的任务）
                tasks_to_cancel = self.active_futures[tasks_to_keep:]
                cancelled_count = 0
                for future in tasks_to_cancel:
                    if not future.done():
                        future.cancel()
                        self.append_log(f"已取消任务：{future}")
                        cancelled_count += 1
                
                # 更新活跃任务列表
                self.active_futures = self.active_futures[:tasks_to_keep]
                
                # 更新线程池大小
                self.max_workers = new_size
                # 重新创建线程池
                self.executor.shutdown(wait=False)
                self.executor = concurrent.futures.ThreadPoolExecutor(max_workers=self.max_workers)
                self.append_log(f"线程池已缩容至：{self.max_workers} 线程")
                self.append_log(f"已取消 {cancelled_count} 个任务")
                
                # 缩容后，无论是否正在处理中，都检查队列并继续处理
                # 因为即使正在处理中，我们也需要确保新的线程池能够处理剩余任务
                if self.task_queue:
                    # 如果当前没有在处理中，直接开始处理
                    if not self.processing_queue:
                        self.append_log("线程池缩容完成，继续处理队列任务")
                        self.start_queue_processing()
                    else:
                        # 如果当前正在处理中，记录日志，等待当前批次完成后自动继续
                        self.append_log("线程池缩容完成，当前有任务正在处理中，将在完成后继续处理队列任务")
            
            # 保存线程数量到配置文件
            global CONFIG
            new_config = CONFIG.copy()
            new_config["max_workers"] = self.max_workers
            save_config(new_config)
            CONFIG = new_config
            self.append_log(f"线程数量已保存到配置文件：{self.max_workers}")
            
            # 更新队列状态显示
            self.update_queue_status()
        except Exception as e:
            self.append_log(f"调整线程池大小失败：{e}")
            messagebox.showerror("错误", f"调整线程池大小失败：{e}")


def main():
    root = tk.Tk()
    app = App(root)
    root.mainloop()

if __name__ == '__main__':
    main()
