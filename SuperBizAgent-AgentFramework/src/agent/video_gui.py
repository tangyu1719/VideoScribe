#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
视频转文字处理工具 - 桌面GUI（Tkinter）
- 只需输入链接并选择平台
- 第一步：根据平台调用下载API获取视频URL并保存至本地
- 第二步：调用语音转文字API并生成Markdown文档
- 输出目录：./output  视频目录：./videos
注意：若第三方网站API不可用，将在日志中提示；程序不中断。
"""

# 忽略NumPy版本警告
import warnings
warnings.filterwarnings("ignore", message="A NumPy version >=1.23.5 and <2.3.0 is required for this version of SciPy")

# ==================== 全局镜像配置 ====================
# 所有下载优先使用镜像源
MIRROR_CONFIG = {
    "enabled": True,  # 是否启用镜像
    "github_mirror": "https://ghproxy.com",  # GitHub镜像
    "pypi_mirror": "https://pypi.tuna.tsinghua.edu.cn/simple",  # PyPI清华镜像
    "npm_mirror": "https://registry.npmmirror.com",  # npm淘宝镜像
}

# ==================== 全局模型缓存配置 ====================
# 统一模型缓存路径，确保多个项目共用同一套模型
import os
# Whisper模型缓存
WHISPER_CACHE_DIR = os.path.expanduser("~/.cache/whisper")
os.environ["WHISPER_CACHE_DIR"] = WHISPER_CACHE_DIR
# HuggingFace模型缓存
HF_HOME = os.path.expanduser("~/.cache/huggingface")
os.environ["HF_HOME"] = HF_HOME
# 避免 Transformers 后台线程访问 huggingface.co 时在国内网络下 ConnectTimeout(10060)；本地有缓存即可离线
# 若需在线拉模型，启动前设置环境变量 HF_HUB_OFFLINE=0
_hf_off = os.environ.get("HF_HUB_OFFLINE", "").strip().lower()
if _hf_off not in ("0", "false", "no"):
    os.environ.setdefault("HF_HUB_OFFLINE", "1")
    os.environ.setdefault("TRANSFORMERS_OFFLINE", "1")
    os.environ.setdefault("HF_DATASETS_OFFLINE", "1")
# Torch模型缓存
TORCH_HOME = os.path.expanduser("~/.cache/torch")
os.environ["TORCH_HOME"] = TORCH_HOME

import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox, filedialog
import threading
import queue
import concurrent.futures
import requests
import json
import os
import time
import hashlib
from datetime import datetime
import multiprocessing
import asyncio
import aiohttp

# LangChain集成已移除
try:
    # Cursor 风格工具（Search / GitHub / Playwright / 读写编辑 / 终端 / 预览 / 联网搜索）
    from cursor_tools.tools import (
        build_cursor_like_tools,
        plan_tool_calls_with_llm,
        execute_tool_calls,
        format_tool_results_for_context,
    )
    CURSOR_TOOLS_AVAILABLE = True
except Exception:
    CURSOR_TOOLS_AVAILABLE = False

try:
    from langchain_standard_runtime import StandardLangChainRuntime, LLMEndpointConfig
    LANGCHAIN_STANDARD_AVAILABLE = True
except Exception:
    LANGCHAIN_STANDARD_AVAILABLE = False

# RAG知识库集成 - 使用快速版kb_manager_fast（BGE-Large 1024维）
try:
    from kb_manager_fast import get_fast_knowledge_base as get_knowledge_base
    KB_MANAGER_AVAILABLE = True
    print("使用快速知识库管理器（BGE-Large 1024维）")
except ImportError as e:
    KB_MANAGER_AVAILABLE = False
    print(f"警告：知识库管理器未安装: {e}")

# 新的RAG工具系统（意图识别、元数据管理）
try:
    from rag_tools import RAGTool, IntentRecognizer, DocumentMetadata, get_metadata_manager, QueryRewriter
    from metadata_dialog import show_metadata_dialog
    from retrieved_chunks_view import create_retrieved_chunks_view
    RAG_TOOLS_AVAILABLE = True
    print("使用新的RAG工具系统（意图识别+元数据）")
except ImportError as e:
    RAG_TOOLS_AVAILABLE = False
    print(f"警告：RAG工具系统未安装: {e}")

# ReAct思考链组件
try:
    from react_thought_chain import create_react_thought_chain, ThoughtStep
    REACT_CHAIN_AVAILABLE = True
    print("使用ReAct思考链组件")
except ImportError as e:
    REACT_CHAIN_AVAILABLE = False
    print(f"警告：ReAct思考链组件未安装: {e}")

# 向后兼容：旧的RAG知识库
try:
    from rag_knowledge_base import RAGKnowledgeBase
    RAG_AVAILABLE = True
except ImportError:
    RAG_AVAILABLE = False
    print("警告：RAG知识库模块未安装，知识库功能将不可用")

# 运维Agent集成
try:
    from ops_agent import create_ops_agent, OpsAgent
    OPS_AGENT_AVAILABLE = True
    print("使用运维Agent模块")
except ImportError as e:
    OPS_AGENT_AVAILABLE = False
    print(f"警告：运维Agent模块未安装: {e}")

# AI问答系统集成
try:
    from chat_gui import ChatGUI
    CHAT_GUI_AVAILABLE = True
except ImportError:
    CHAT_GUI_AVAILABLE = False
    print("警告：AI问答系统模块未安装")

# 新的AI问答页面（带任务管理）
try:
    from ai_chat_page import AIChatPage
    AI_CHAT_PAGE_AVAILABLE = True
except ImportError:
    AI_CHAT_PAGE_AVAILABLE = False
    print("警告：AI问答页面模块未安装")

# AI API配置管理模块
try:
    from ai_api_config_gui import open_ai_api_config_window, AIAPIConfigManager
    AI_API_CONFIG_AVAILABLE = True
except ImportError:
    AI_API_CONFIG_AVAILABLE = False
    print("警告：AI API配置模块未安装")

# 多模态文档处理模块
try:
    from multimodal_gui import MultimodalProcessingPage
    MULTIMODAL_AVAILABLE = True
except ImportError:
    MULTIMODAL_AVAILABLE = False
    print("警告：多模态文档处理模块未安装")

APP_TITLE = "视频转文字处理工具 (GUI)"

# 主窗口视觉主题：浅灰底、细线白卡片、低饱和强调色（Tkinter 可维护范围内尽量「轻」）
UI_BG = "#f4f4f5"
UI_CARD = "#ffffff"
UI_BORDER = "#e4e4e7"
UI_ACCENT = "#2563eb"
UI_ACCENT_SOFT = "#eef2ff"
UI_TEXT = "#18181b"
UI_TEXT_MUTED = "#52525b"
UI_TEXT_LIGHT = "#a1a1aa"
UI_LOG_BG = "#fafafa"
UI_FONT = ("Microsoft YaHei UI", 10)
UI_FONT_BOLD = ("Microsoft YaHei UI", 10, "bold")
UI_FONT_TITLE = ("Microsoft YaHei UI", 20, "bold")
UI_FONT_NAV = ("Microsoft YaHei UI", 11, "bold")

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
VIDEO_DIR = os.path.join(BASE_DIR, "videos")
OUTPUT_DIR = os.path.join(BASE_DIR, "output")
CONFIG_FILE = os.path.join(BASE_DIR, "config.json")
HISTORY_FILE = os.path.join(BASE_DIR, "history.json")

# 统一Whisper模型缓存路径 - 使用用户目录下的统一缓存
WHISPER_CACHE_DIR = os.path.expanduser("~/.cache/whisper")
os.environ["WHISPER_CACHE_DIR"] = WHISPER_CACHE_DIR

# 设置 ffmpeg 路径（统一 demo_wendanghua/ffmpeg/bin）
from ffmpeg_path import resolve_ffmpeg_bin_dir

FFMPEG_DIR = resolve_ffmpeg_bin_dir()
if FFMPEG_DIR:
    os.environ["PATH"] = FFMPEG_DIR + os.pathsep + os.environ.get("PATH", "")
    print(f"已添加本地 ffmpeg 路径: {FFMPEG_DIR}")

# 火山引擎 API 配置（正确的API Key）
VOLCENGINE_API_KEY = "ebc08852-e7ae-4e64-b71c-79cfcce9d251"
VOLCENGINE_API_URL = "https://ark.cn-beijing.volces.com/api/v3"

# AI 对话专用 API 配置（正确的API Key）
AI_CHAT_API_KEY = "ebc08852-e7ae-4e64-b71c-79cfcce9d251"
AI_CHAT_API_URL = "https://ark.cn-beijing.volces.com/api/v3"
AI_CHAT_MODEL = "ep-20260411182220-jv5qt"  # Doubao-Seed-2.0-mini 主接入点
AI_CHAT_MODEL_BACKUP = "ep-20260320202115-9jqfp"  # Doubao-Seed-2.0-mini 备用接入点

# 百度 OCR API 配置（与链接分析保持一致）
BAIDU_OCR_APP_ID = '122094788'
BAIDU_OCR_API_KEY = 'KZOpVw7PGLRiBdsqRnuLFVY7'
BAIDU_OCR_SECRET_KEY = 'L1pdbtb4IZZv67ofXnsxDNAhELGN2UXs'

# 默认配置
DEFAULT_CONFIG = {
    "summary_prompt": "请对以下文本进行总结，提取关键知识点，整理成结构化的格式。\n要求：\n1. 第一行必须是一个简洁的中文标题（不超过20个字符，不要包含#号）\n2. 后续内容按逻辑分段整理\n{text}",
    "volcengine_api_key": VOLCENGINE_API_KEY,
    "system_prompt": "你是一个专业的视频内容分析助手，擅长从视频转写内容中提取关键信息并进行结构化分析。你的输出格式要求：\n1. 第一行是简洁的中文标题（不超过20字符，不要包含#号，不要包含markdown语法标记）\n2. 后续是结构化的分析内容",
    "rules": "1. 第一行必须是简洁的中文标题（不超过20字符，不要包含#号）\n2. 提取视频中的关键知识点和核心信息\n3. 保持客观中立的分析态度\n4. 结构化呈现分析结果\n5. 重点关注视频中的技术讲解和实用信息",
    "file_naming_rule": "总记录序号-月-日-文档名称（文档名称从AI生成的第一行标题中提取）",  # 文件名命名规则
    "output_template": "# {platform}视频分析\n\n## 视频信息\n- 分析时间: {datetime}\n- 原始链接: {link}\n- 平台: {platform}\n\n## 语音转文字内容\n{transcript}\n\n## AI分析摘要\n{summary}",
    "user_prompt": "",
    "ai_chat_model": "ep-20260411182220-jv5qt",
    "ai_chat_model_backup": "ep-20260320202115-9jqfp",
    # 飞书：同步开关持久化在 config.json；凭证与默认目录在「AI配置」中填写
    "feishu_sync_enabled": False,
    "feishu_app_id": "",
    "feishu_app_secret": "",
    "feishu_default_folder_path": "",
    # 云空间文件夹 fldcn，或 …/drive/folder/fldcn… 完整 URL；纯中文路径需配合此项
    "feishu_folder_token": "",
    # 知识库：MD 先导入云文档再迁入 wiki（需 wiki:wiki 等权限，应用需加入知识库管理员）
    "feishu_wiki_sync_enabled": False,
    "feishu_wiki_space_name": "就业知识库",
    "feishu_wiki_space_id": "",
    "feishu_wiki_anchor_node_token": "",
    "feishu_wiki_path_ensure": "就业技术文档集/AI相关",
    # 标准 LangChain 架构开关（模型+Agent+Memory+RAG Tool）
    "langchain_standard_enabled": True,
    "llm_provider": "ark",
    "llm_base_url": AI_CHAT_API_URL,
}

for d in (VIDEO_DIR, OUTPUT_DIR):
    if not os.path.exists(d):
        os.makedirs(d)

# 加载配置文件（与 DEFAULT_CONFIG 合并；若 MariaDB 有 video_agent_config 则覆盖同名键）
def load_config():
    base = dict(DEFAULT_CONFIG)
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                user = json.load(f)
            if isinstance(user, dict):
                base.update(user)
        except Exception as e:
            print(f"加载配置文件失败：{e}")
    try:
        from video_config_mariadb import load_from_mariadb

        db_cfg = load_from_mariadb()
        if db_cfg:
            base.update(db_cfg)
            print("[配置] 已合并 MariaDB video_agent_config（库中字段覆盖 config.json）")
    except Exception as e:
        print(f"[配置] MariaDB 合并跳过：{e}")
    return base

# 保存配置文件（config.json + MariaDB 双写，与 ai_api_config_gui / db 模块一致）
def save_config(config):
    ok_file = False
    try:
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
        ok_file = True
    except Exception as e:
        print(f"保存配置文件失败：{e}")
    try:
        from video_config_mariadb import is_available, save_to_mariadb

        if is_available():
            if save_to_mariadb(config):
                print("[配置] 已同步至 MariaDB 表 video_agent_config")
            else:
                print("[配置] 警告：MariaDB 写入失败，已仅更新 config.json；请检查 db 模块与数据库")
    except Exception as e:
        print(f"[配置] MariaDB 同步异常（已仍尝试保留 config.json）：{e}")
    return ok_file

# 加载历史记录
def load_history():
    if os.path.exists(HISTORY_FILE):
        try:
            with open(HISTORY_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            print(f"加载历史记录失败：{e}")
    return {"tasks": []}

# 保存历史记录
def save_history(history):
    try:
        with open(HISTORY_FILE, 'w', encoding='utf-8') as f:
            json.dump(history, f, ensure_ascii=False, indent=2)
        return True
    except Exception as e:
        print(f"保存历史记录失败：{e}")
        return False


# ==================== 豆包AI风格聊天界面 ====================
class DoubaoChatPage(tk.Frame):
    """豆包AI风格聊天页面 - 模仿网页版豆包AI界面"""

    # AI配置默认值（深拷贝自视频分析配置，但独立管理）
    DEFAULT_AI_CONFIG = {
        "thinking_system_prompt": "你是一个善于分析的AI助手。请简要分析用户问题的关键点，列出3-5个思考步骤。用简洁的要点形式回答。",
        "response_system_prompt": "你是一个专业的AI助手，擅长回答各种问题。请根据用户的问题提供准确、有用的回答。",
        "temperature": 0.7,
        "max_tokens": 4096,
        "top_p": 0.9,
        "api_key": AI_CHAT_API_KEY,
        "model": AI_CHAT_MODEL,
        "model_backup": AI_CHAT_MODEL_BACKUP
    }

    def __init__(self, parent, rag_kb=None, rag_tool=None, **kwargs):
        super().__init__(parent, bg="#f5f5f5", **kwargs)
        self.sessions = []
        self.current_session_id = None
        self.messages = []
        self.pending_images = []
        self.ai_config = self._load_ai_config()
        self._rag_kb = rag_kb  # 接收传入的RAG知识库
        self._rag_tool = rag_tool  # 接收传入的RAG工具
        self._cursor_tools = build_cursor_like_tools(BASE_DIR) if CURSOR_TOOLS_AVAILABLE else []
        self._standard_runtime = None
        self._init_standard_langchain_runtime()

        # 加载会话历史
        self._load_sessions()

        # 创建UI
        self._create_ui()

        # 如果没有会话，创建新会话
        if not self.sessions:
            self._create_new_session()
        else:
            self._load_session(self.sessions[0]["id"])

    def _init_standard_langchain_runtime(self):
        """初始化标准 LangChain 运行时（可开关）。"""
        try:
            enabled = bool(CONFIG.get("langchain_standard_enabled", True))
            if not enabled or not LANGCHAIN_STANDARD_AVAILABLE:
                self._standard_runtime = None
                return

            cfg = LLMEndpointConfig(
                provider=str(CONFIG.get("llm_provider", "ark")),
                api_key=self.ai_config.get("api_key", AI_CHAT_API_KEY),
                base_url=str(CONFIG.get("llm_base_url", AI_CHAT_API_URL)),
                model=self.ai_config.get("model", AI_CHAT_MODEL),
                temperature=float(self.ai_config.get("temperature", 0.7)),
                max_tokens=int(self.ai_config.get("max_tokens", 4096)),
            )

            self._standard_runtime = StandardLangChainRuntime(
                base_dir=BASE_DIR,
                llm_config=cfg,
                rag_tool=self._rag_tool,
                rag_kb=self._rag_kb,
                logger=self._log_to_file,
            )
            if self._standard_runtime and self._standard_runtime.ready:
                self._log_to_file("标准 LangChain 运行时已启用")
            else:
                self._log_to_file("标准 LangChain 运行时不可用，继续使用旧链路")
        except Exception as e:
            self._standard_runtime = None
            self._log_to_file(f"标准 LangChain 运行时初始化失败: {e}")

    def _load_ai_config(self):
        """加载AI配置"""
        try:
            config_file = os.path.join(BASE_DIR, "ai_chat_config.json")
            if os.path.exists(config_file):
                with open(config_file, 'r', encoding='utf-8') as f:
                    loaded_config = json.load(f)
                    # 合并默认配置和加载的配置
                    config = self.DEFAULT_AI_CONFIG.copy()
                    config.update(loaded_config)
                    return config
        except Exception as e:
            print(f"加载AI配置失败: {e}")
        return self.DEFAULT_AI_CONFIG.copy()

    def _save_ai_config(self):
        """保存AI配置"""
        try:
            config_file = os.path.join(BASE_DIR, "ai_chat_config.json")
            with open(config_file, 'w', encoding='utf-8') as f:
                json.dump(self.ai_config, f, ensure_ascii=False, indent=2)
            print("AI配置已保存")
        except Exception as e:
            print(f"保存AI配置失败: {e}")

    def _create_ui(self):
        """创建豆包AI风格界面"""
        # 主分割面板
        main_paned = tk.PanedWindow(self, orient=tk.HORIZONTAL, bg="#f5f5f5", sashwidth=1)
        main_paned.pack(fill=tk.BOTH, expand=True)

        # 左侧边栏
        self.sidebar = self._create_sidebar(main_paned)
        main_paned.add(self.sidebar, width=260, minsize=200)

        # 右侧聊天区域
        self.chat_area = self._create_chat_area(main_paned)
        main_paned.add(self.chat_area, width=940, minsize=400)

    def _create_sidebar(self, parent):
        """创建左侧边栏 - 豆包风格"""
        sidebar = tk.Frame(parent, bg="#ffffff", width=260)
        sidebar.pack_propagate(False)

        # 顶部标题区域
        header = tk.Frame(sidebar, bg="#ffffff", height=60)
        header.pack(fill=tk.X, padx=15, pady=10)
        header.pack_propagate(False)

        # AI助手标题
        title_label = tk.Label(
            header,
            text="🤖 AI助手",
            font=("微软雅黑", 16, "bold"),
            bg="#ffffff",
            fg="#1a1a1a"
        )
        title_label.pack(side=tk.LEFT)

        # 新对话按钮
        new_chat_btn = tk.Label(
            header,
            text="＋",
            font=("微软雅黑", 20),
            bg="#ffffff",
            fg="#666666",
            cursor="hand2"
        )
        new_chat_btn.pack(side=tk.RIGHT, padx=(0, 10))
        new_chat_btn.bind("<Button-1>", lambda e: self._create_new_session())

        # 设置按钮（COZE 风格）
        settings_btn = tk.Label(
            header,
            text="⚙️",
            font=("微软雅黑", 14),
            bg="#ffffff",
            fg="#666666",
            cursor="hand2"
        )
        settings_btn.pack(side=tk.RIGHT)
        settings_btn.bind("<Button-1>", lambda e: self._open_settings())
        
        # 知识库管理按钮（打开独立窗口）
        if RAG_AVAILABLE:
            kb_manage_btn = tk.Label(
                header,
                text="📚",
                font=("微软雅黑", 14),
                bg="#ffffff",
                fg="#666666",
                cursor="hand2"
            )
            kb_manage_btn.pack(side=tk.RIGHT, padx=(0, 10))
            kb_manage_btn.bind("<Button-1>", lambda e: self._open_rag_manager())

        # 分隔线
        separator = tk.Frame(sidebar, bg="#e8e8e8", height=1)
        separator.pack(fill=tk.X, padx=15)

        # 知识库管理区域（添加文件/文件夹按钮）
        if RAG_AVAILABLE:
            kb_frame = tk.Frame(sidebar, bg="#ffffff", padx=15, pady=10)
            kb_frame.pack(fill=tk.X)
            
            tk.Label(
                kb_frame,
                text="📚 知识库",
                font=("微软雅黑", 11, "bold"),
                bg="#ffffff",
                fg="#333333"
            ).pack(anchor="w", pady=(0, 5))
            
            btn_frame = tk.Frame(kb_frame, bg="#ffffff")
            btn_frame.pack(fill=tk.X)
            
            # 添加文件按钮
            add_file_btn = tk.Button(
                btn_frame,
                text="📄 添加文件",
                command=self._add_file_to_kb,
                bg="#f0f0f0",
                fg="#333333",
                font=("微软雅黑", 9),
                relief=tk.FLAT,
                cursor="hand2"
            )
            add_file_btn.pack(side=tk.LEFT, padx=(0, 5))
            
            # 添加文件夹按钮
            add_folder_btn = tk.Button(
                btn_frame,
                text="📁 添加文件夹",
                command=self._add_folder_to_kb,
                bg="#f0f0f0",
                fg="#333333",
                font=("微软雅黑", 9),
                relief=tk.FLAT,
                cursor="hand2"
            )
            add_folder_btn.pack(side=tk.LEFT)
            
            # 重建索引按钮
            rebuild_btn = tk.Button(
                btn_frame,
                text="🔄 重建索引",
                command=self._rebuild_kb_index,
                bg="#4e6ef2",
                fg="#ffffff",
                font=("微软雅黑", 9),
                relief=tk.FLAT,
                cursor="hand2"
            )
            rebuild_btn.pack(side=tk.LEFT, padx=(5, 0))
            
            # 分隔线
            separator2 = tk.Frame(sidebar, bg="#e8e8e8", height=1)
            separator2.pack(fill=tk.X, padx=15)

        # 会话列表区域
        list_frame = tk.Frame(sidebar, bg="#ffffff")
        list_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # 会话列表Canvas
        self.session_canvas = tk.Canvas(list_frame, bg="#ffffff", highlightthickness=0)
        self.session_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        scrollbar = ttk.Scrollbar(list_frame, orient="vertical", command=self.session_canvas.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.session_canvas.configure(yscrollcommand=scrollbar.set)

        self.session_list_frame = tk.Frame(self.session_canvas, bg="#ffffff")
        self.session_canvas.create_window((0, 0), window=self.session_list_frame, anchor="nw", width=220)

        self.session_list_frame.bind("<Configure>", lambda e: self.session_canvas.configure(
            scrollregion=self.session_canvas.bbox("all")))

        # 刷新会话列表
        self._refresh_session_list()

        return sidebar

    def _create_chat_area(self, parent):
        """创建右侧聊天区域"""
        chat_frame = tk.Frame(parent, bg="#f5f5f5")

        # 聊天消息区域
        msg_frame = tk.Frame(chat_frame, bg="#f5f5f5")
        msg_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=(20, 10))

        # 消息Canvas
        self.msg_canvas = tk.Canvas(msg_frame, bg="#f5f5f5", highlightthickness=0)
        self.msg_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        msg_scrollbar = ttk.Scrollbar(msg_frame, orient="vertical", command=self.msg_canvas.yview)
        msg_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.msg_canvas.configure(yscrollcommand=msg_scrollbar.set)

        self.messages_frame = tk.Frame(self.msg_canvas, bg="#f5f5f5")
        self.msg_canvas.create_window((0, 0), window=self.messages_frame, anchor="nw", width=900)

        self.messages_frame.bind("<Configure>", lambda e: self.msg_canvas.configure(
            scrollregion=self.msg_canvas.bbox("all")))

        # 底部输入区域
        self._create_input_area(chat_frame)

        return chat_frame

    def _create_input_area(self, parent):
        """创建底部输入区域 - 豆包风格"""
        input_frame = tk.Frame(parent, bg="#ffffff", height=120)
        input_frame.pack(fill=tk.X, side=tk.BOTTOM, padx=20, pady=20)
        input_frame.pack_propagate(False)

        # 输入框
        input_container = tk.Frame(input_frame, bg="#f5f5f5", height=50)
        input_container.pack(fill=tk.X, padx=15, pady=15)
        input_container.pack_propagate(False)

        self.input_box = tk.Text(
            input_container,
            font=("微软雅黑", 12),
            bg="#f5f5f5",
            fg="#333333",
            relief=tk.FLAT,
            height=2,
            wrap=tk.WORD
        )
        self.input_box.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(15, 10), pady=10)
        self.input_box.bind("<Return>", self._on_enter_pressed)
        self.input_box.bind("<Shift-Return>", self._on_shift_enter)

        # 右侧按钮区域
        btn_frame = tk.Frame(input_container, bg="#f5f5f5")
        btn_frame.pack(side=tk.RIGHT, padx=10)

        # 图片上传按钮
        img_btn = tk.Label(
            btn_frame,
            text="📎",
            font=("微软雅黑", 16),
            bg="#f5f5f5",
            fg="#666666",
            cursor="hand2"
        )
        img_btn.pack(side=tk.LEFT, padx=5)
        img_btn.bind("<Button-1>", lambda e: self._upload_image())

        # 发送按钮
        send_btn = tk.Label(
            btn_frame,
            text="➤",
            font=("微软雅黑", 16),
            bg="#4e6ef2",
            fg="#ffffff",
            cursor="hand2",
            width=3,
            height=1
        )
        send_btn.pack(side=tk.LEFT, padx=5)
        send_btn.bind("<Button-1>", lambda e: self._send_message())

        # 提示文字
        hint_label = tk.Label(
            input_frame,
            text="Enter发送，Shift+Enter换行",
            font=("微软雅黑", 9),
            bg="#ffffff",
            fg="#999999"
        )
        hint_label.pack(side=tk.BOTTOM, pady=(0, 8))

    def _create_selectable_text(self, parent, text, font, bg, fg, width=60, height=None):
        """创建可选择的文本组件（使用Text组件）"""
        # 计算需要的行数
        if height is None:
            lines = text.split('\n')
            height = max(1, len(lines))
            # 估算每行能显示的字符数
            avg_chars_per_line = 50
            for line in lines:
                needed_lines = max(1, (len(line) // avg_chars_per_line) + 1)
                height += needed_lines - 1
            height = min(height, 20)  # 最大20行
        
        text_widget = tk.Text(
            parent,
            font=font,
            bg=bg,
            fg=fg,
            wrap=tk.WORD,
            width=width,
            height=height,
            padx=10,
            pady=10,
            relief=tk.FLAT,
            highlightthickness=0,
            selectbackground="#b3d9ff",
            selectforeground="#000000"
        )
        text_widget.insert("1.0", text)
        text_widget.config(state=tk.DISABLED)  # 只读，但可选择
        
        # 绑定右键菜单
        text_widget.bind("<Button-3>", lambda e: self._show_text_context_menu(e, text_widget))
        
        return text_widget

    def _show_text_context_menu(self, event, text_widget):
        """显示文本右键菜单"""
        menu = tk.Menu(self, tearoff=0)
        menu.add_command(label="复制", command=lambda: self._copy_selected_text(text_widget))
        menu.add_command(label="全选", command=lambda: text_widget.tag_add("sel", "1.0", "end"))
        menu.post(event.x_root, event.y_root)

    def _copy_selected_text(self, text_widget):
        """复制选中的文本"""
        try:
            selected = text_widget.get("sel.first", "sel.last")
            self.clipboard_clear()
            self.clipboard_append(selected)
        except tk.TclError:
            pass  # 没有选中的文本

    def _create_message_bubble(self, msg, msg_index=None):
        """创建消息气泡（支持流式更新和文本选择）- 豆包风格"""
        is_user = msg.get("role") == "user"
        msg_id = msg.get("id") or msg_index or id(msg)

        # 消息容器
        bubble_frame = tk.Frame(self.messages_frame, bg="#f5f5f5")
        bubble_frame.pack(fill=tk.X, padx=20, pady=10)
        bubble_frame.msg_id = msg_id

        if is_user:
            # 用户消息 - 右对齐
            content_frame = tk.Frame(bubble_frame, bg="#f5f5f5")
            content_frame.pack(side=tk.RIGHT)

            # 头像
            avatar = tk.Label(
                content_frame,
                text="👤",
                font=("微软雅黑", 20),
                bg="#f5f5f5"
            )
            avatar.pack(side=tk.RIGHT, padx=(10, 0))

            # 气泡 - 使用可选择Text组件
            content = msg.get("content", "")
            lines = content.split('\n')
            height = max(1, len(lines))
            for line in lines:
                height += max(0, (len(line) // 45))  # 估算自动换行
            height = min(height, 20)
            
            bubble = tk.Text(
                content_frame,
                font=("微软雅黑", 12),
                bg="#4e6ef2",
                fg="#ffffff",
                wrap=tk.WORD,
                width=50,
                height=height,
                padx=15,
                pady=10,
                relief=tk.FLAT,
                highlightthickness=0,
                selectbackground="#80aaff",
                selectforeground="#ffffff"
            )
            bubble.insert("1.0", content)
            bubble.config(state=tk.DISABLED)
            bubble.pack(side=tk.RIGHT)
            bubble.bind("<Button-3>", lambda e: self._show_text_context_menu(e, bubble))
        else:
            # AI消息 - 左对齐
            content_frame = tk.Frame(bubble_frame, bg="#f5f5f5")
            content_frame.pack(side=tk.LEFT, fill=tk.X, expand=True)

            # 头像
            avatar = tk.Label(
                content_frame,
                text="🤖",
                font=("微软雅黑", 20),
                bg="#f5f5f5"
            )
            avatar.pack(side=tk.LEFT, padx=(0, 10), anchor="n")

            # AI内容容器
            ai_content_frame = tk.Frame(content_frame, bg="#f5f5f5")
            ai_content_frame.pack(side=tk.LEFT, fill=tk.X, expand=True)
            ai_content_frame.msg_id = msg_id

            # ReAct思考链（替代原来的思考过程）
            if REACT_CHAIN_AVAILABLE:
                thought_chain = create_react_thought_chain(ai_content_frame)
                ai_content_frame.thought_chain = thought_chain
            else:
                # 向后兼容：使用旧的思考过程
                thinking_section = self._create_thinking_section(ai_content_frame, msg_id, msg.get("thinking", ""))
                ai_content_frame.thinking_section = thinking_section

            # 回复内容 - 使用可选择Text组件
            content = msg.get("content", "")
            lines = content.split('\n')
            height = max(1, len(lines))
            for line in lines:
                height += max(0, (len(line) // 45))
            height = min(height, 30)
            
            content_text = tk.Text(
                ai_content_frame,
                font=("微软雅黑", 12),
                bg="#ffffff",
                fg="#1a1a1a",
                wrap=tk.WORD,
                width=55,
                height=height,
                padx=15,
                pady=10,
                relief=tk.FLAT,
                highlightthickness=0,
                selectbackground="#b3d9ff",
                selectforeground="#000000"
            )
            content_text.insert("1.0", content)
            content_text.config(state=tk.DISABLED)
            content_text.pack(anchor="w", pady=(5, 0), fill=tk.X)
            content_text.bind("<Button-3>", lambda e: self._show_text_context_menu(e, content_text))
            ai_content_frame.content_text = content_text

            # 存储消息引用
            bubble_frame.ai_content_frame = ai_content_frame

        # 图片显示
        if msg.get("images"):
            for img_path in msg["images"]:
                try:
                    img_label = tk.Label(bubble_frame, text=f"[图片: {os.path.basename(img_path)}]",
                                        bg="#f5f5f5", fg="#4e6ef2")
                    img_label.pack()
                except:
                    pass

        return bubble_frame

    def _create_thinking_section(self, parent, msg_id, initial_content=""):
        """创建可折叠的思考过程区域（支持流式更新和状态显示）- 使用Text组件支持选择"""
        # 思考过程容器
        thinking_container = tk.Frame(parent, bg="#f5f5f5")
        thinking_container.pack(fill=tk.X, pady=(0, 5))

        # 头部（可点击折叠/展开）
        header_frame = tk.Frame(thinking_container, bg="#f0f0f0", cursor="hand2")
        header_frame.pack(fill=tk.X)

        # 展开/折叠图标
        toggle_icon = tk.Label(
            header_frame,
            text="▼",
            font=("微软雅黑", 10),
            bg="#f0f0f0",
            fg="#666666"
        )
        toggle_icon.pack(side=tk.LEFT, padx=(10, 5))

        # 标题
        title_label = tk.Label(
            header_frame,
            text="思考过程",
            font=("微软雅黑", 10),
            bg="#f0f0f0",
            fg="#666666"
        )
        title_label.pack(side=tk.LEFT)

        # 状态标签（显示当前步骤）
        status_label = tk.Label(
            header_frame,
            text="",
            font=("微软雅黑", 9),
            bg="#f0f0f0",
            fg="#4e6ef2"
        )
        status_label.pack(side=tk.RIGHT, padx=10)

        # 内容区域（默认展开）
        content_frame = tk.Frame(thinking_container, bg="#f8f8f8")
        content_frame.pack(fill=tk.X, padx=0, pady=0)

        # 思考内容 - 使用Text组件支持选择和复制
        lines = initial_content.split('\n') if initial_content else [""]
        height = max(3, min(len(lines) + 1, 10))
        
        thinking_text = tk.Text(
            content_frame,
            font=("微软雅黑", 10),
            bg="#f8f8f8",
            fg="#888888",
            wrap=tk.WORD,
            width=60,
            height=height,
            padx=15,
            pady=10,
            relief=tk.FLAT,
            highlightthickness=0,
            selectbackground="#d0d0d0",
            selectforeground="#000000"
        )
        thinking_text.insert("1.0", initial_content)
        thinking_text.config(state=tk.DISABLED)
        thinking_text.pack(fill=tk.X)
        thinking_text.bind("<Button-3>", lambda e: self._show_text_context_menu(e, thinking_text))

        # 存储引用以便流式更新
        thinking_container.msg_id = msg_id
        thinking_container.thinking_text = thinking_text
        thinking_container.content_frame = content_frame
        thinking_container.toggle_icon = toggle_icon
        thinking_container.header_frame = header_frame
        thinking_container.title_label = title_label
        thinking_container.status_label = status_label

        # 点击事件处理
        def toggle_thinking():
            if content_frame.winfo_viewable():
                content_frame.pack_forget()
                toggle_icon.config(text="▶")
                header_frame.config(bg="#e8e8e8")
                toggle_icon.config(bg="#e8e8e8")
                title_label.config(bg="#e8e8e8")
                status_label.config(bg="#e8e8e8")
            else:
                content_frame.pack(fill=tk.X, padx=0, pady=0)
                toggle_icon.config(text="▼")
                header_frame.config(bg="#f0f0f0")
                toggle_icon.config(bg="#f0f0f0")
                title_label.config(bg="#f0f0f0")
                status_label.config(bg="#f0f0f0")

        header_frame.bind("<Button-1>", lambda e: toggle_thinking())
        toggle_icon.bind("<Button-1>", lambda e: toggle_thinking())
        title_label.bind("<Button-1>", lambda e: toggle_thinking())

        return thinking_container

    def _refresh_messages(self):
        """刷新消息显示"""
        for widget in self.messages_frame.winfo_children():
            widget.destroy()

        for i, msg in enumerate(self.messages):
            self._create_message_bubble(msg, msg_index=i)

    def _stream_update_thinking(self, msg_index, text, status=None):
        """流式更新思考过程 - 使用Text组件，支持状态显示"""
        if msg_index < len(self.messages_frame.winfo_children()):
            bubble_frame = self.messages_frame.winfo_children()[msg_index]
            if hasattr(bubble_frame, 'ai_content_frame'):
                ai_frame = bubble_frame.ai_content_frame
                if hasattr(ai_frame, 'thinking_section'):
                    thinking_text = ai_frame.thinking_section.thinking_text
                    thinking_text.config(state=tk.NORMAL)
                    thinking_text.delete("1.0", tk.END)
                    thinking_text.insert("1.0", text)
                    thinking_text.config(state=tk.DISABLED)
                    # 调整高度
                    lines = text.split('\n')
                    height = max(3, min(len(lines) + 1, 10))
                    thinking_text.config(height=height)
                    
                    # 更新状态标签
                    if status:
                        status_label = ai_frame.thinking_section.status_label
                        status_label.config(text=status)
                    
                    self.update_idletasks()

    def _stream_update_content(self, msg_index, text):
        """流式更新回复内容 - 使用Text组件"""
        if msg_index < len(self.messages_frame.winfo_children()):
            bubble_frame = self.messages_frame.winfo_children()[msg_index]
            if hasattr(bubble_frame, 'ai_content_frame'):
                content_text = bubble_frame.ai_content_frame
                content_text.config(state=tk.NORMAL)
                content_text.delete("1.0", tk.END)
                content_text.insert("1.0", text)
                content_text.config(state=tk.DISABLED)
                # 调整高度
                lines = text.split('\n')
                height = max(3, len(lines))
                for line in lines:
                    height += max(0, (len(line) // 45))
                height = min(height, 30)
                content_text.config(height=height)
                self.update_idletasks()
                # 自动滚动到底部
                self._scroll_to_bottom()

    def _scroll_to_bottom(self):
        """滚动消息区域到底部"""
        try:
            self.after(100, lambda: self.msg_canvas.yview_moveto(1.0))
        except Exception as e:
            print(f"滚动失败: {e}")

    def _refresh_session_list(self):
        """刷新会话列表"""
        for widget in self.session_list_frame.winfo_children():
            widget.destroy()

        for session in self.sessions:
            session_id = session["id"]
            title = session.get("title", "新对话")
            is_active = session_id == self.current_session_id

            # 会话项
            item_frame = tk.Frame(
                self.session_list_frame,
                bg="#e8f0fe" if is_active else "#ffffff",
                height=40,
                cursor="hand2"
            )
            item_frame.pack(fill=tk.X, pady=2)
            item_frame.pack_propagate(False)
            
            # 按钮容器（先pack确保按钮始终可见，不被标题挤压）
            btn_frame = tk.Frame(
                item_frame,
                bg="#e8f0fe" if is_active else "#ffffff",
                width=60  # 固定宽度确保按钮不被挤压
            )
            btn_frame.pack(side=tk.RIGHT, padx=5)
            btn_frame.pack_propagate(False)  # 防止子组件改变大小
            
            # 标题（放在按钮后面pack，使用剩余空间）
            title_label = tk.Label(
                item_frame,
                text=title[:15] + "..." if len(title) > 15 else title,
                font=("微软雅黑", 11),
                bg="#e8f0fe" if is_active else "#ffffff",
                fg="#1a1a1a" if is_active else "#666666",
                anchor="w"
            )
            title_label.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(15, 5))
            
            # 重命名按钮（增强可见性）
            rename_btn = tk.Label(
                btn_frame,
                text="✏️",
                font=("微软雅黑", 10),
                bg="#e8f0fe" if is_active else "#ffffff",
                fg="#ff9800",  # 橙色更显眼
                cursor="hand2",
                padx=5,
                pady=2
            )
            rename_btn.pack(side=tk.LEFT, padx=2)
            rename_btn.bind("<Button-1>", lambda e, sid=session_id: self._rename_session(sid))
            rename_btn.bind("<Enter>", lambda e: rename_btn.config(bg="#ffe0b2"))
            rename_btn.bind("<Leave>", lambda e: rename_btn.config(bg="#e8f0fe" if is_active else "#ffffff"))
            
            # 删除按钮（增强可见性）
            delete_btn = tk.Label(
                btn_frame,
                text="🗑️",
                font=("微软雅黑", 10),
                bg="#e8f0fe" if is_active else "#ffffff",
                fg="#f44336",  # 红色更显眼
                cursor="hand2",
                padx=5,
                pady=2
            )
            delete_btn.pack(side=tk.LEFT, padx=2)
            delete_btn.bind("<Button-1>", lambda e, sid=session_id: self._delete_session(sid))
            delete_btn.bind("<Enter>", lambda e: delete_btn.config(bg="#ffcdd2"))
            delete_btn.bind("<Leave>", lambda e: delete_btn.config(bg="#e8f0fe" if is_active else "#ffffff"))

            # 点击切换会话
            item_frame.bind("<Button-1>", lambda e, sid=session_id: self._load_session(sid))
            title_label.bind("<Button-1>", lambda e, sid=session_id: self._load_session(sid))

    def _rename_session(self, session_id):
        """重命名会话"""
        # 找到当前会话
        session = None
        for s in self.sessions:
            if s["id"] == session_id:
                session = s
                break
        
        if not session:
            return
        
        # 弹出重命名对话框
        dialog = tk.Toplevel(self)
        dialog.title("重命名对话")
        dialog.geometry("300x120")
        dialog.transient(self)
        dialog.grab_set()
        
        tk.Label(
            dialog,
            text="输入新名称:",
            font=("微软雅黑", 11)
        ).pack(pady=(10, 5))
        
        entry = tk.Entry(dialog, font=("微软雅黑", 11), width=30)
        entry.pack(pady=5)
        entry.insert(0, session.get("title", "新对话"))
        entry.select_range(0, tk.END)
        entry.focus()
        
        def do_rename():
            new_title = entry.get().strip()
            if new_title:
                session["title"] = new_title
                self._save_sessions()
                self._refresh_session_list()
            dialog.destroy()
        
        btn_frame = tk.Frame(dialog)
        btn_frame.pack(pady=10)
        
        tk.Button(
            btn_frame,
            text="确定",
            command=do_rename,
            bg="#4e6ef2",
            fg="white",
            font=("微软雅黑", 10),
            width=8
        ).pack(side=tk.LEFT, padx=5)
        
        tk.Button(
            btn_frame,
            text="取消",
            command=dialog.destroy,
            font=("微软雅黑", 10),
            width=8
        ).pack(side=tk.LEFT, padx=5)
        
        # 回车确认
        entry.bind("<Return>", lambda e: do_rename())
    
    def _delete_session(self, session_id):
        """删除会话"""
        # 找到当前会话
        session = None
        session_index = -1
        for i, s in enumerate(self.sessions):
            if s["id"] == session_id:
                session = s
                session_index = i
                break
        
        if not session:
            return
        
        # 确认删除
        if not messagebox.askyesno("确认删除", f"确定要删除对话：{session.get('title', '新对话')}？\n\n此操作不可恢复！"):
            return
        
        # 删除会话
        self.sessions.pop(session_index)
        
        # 如果删除的是当前会话，切换到第一个会话或创建新会话
        if session_id == self.current_session_id:
            if self.sessions:
                self._load_session(self.sessions[0]["id"])
            else:
                self._create_new_session()
        else:
            self._save_sessions()
            self._refresh_session_list()

    def _create_new_session(self):
        """创建新会话"""
        session_id = f"session_{int(time.time() * 1000)}"
        session = {
            "id": session_id,
            "title": "新对话",
            "created_at": datetime.now().isoformat(),
            "messages": []
        }
        self.sessions.insert(0, session)
        self._save_sessions()
        self._refresh_session_list()
        self._load_session(session_id)

    def _load_session(self, session_id):
        """加载会话"""
        self.current_session_id = session_id
        for session in self.sessions:
            if session["id"] == session_id:
                self.messages = session.get("messages", [])
                break
        self._refresh_session_list()
        self._refresh_messages()

    def _update_session_title(self, first_message):
        """根据第一条消息自动命名会话"""
        if not first_message:
            return

        title = first_message[:20] if len(first_message) <= 20 else first_message[:17] + "..."

        for session in self.sessions:
            if session["id"] == self.current_session_id:
                session["title"] = title
                break

        self._save_sessions()
        self._refresh_session_list()

    def _send_message(self):
        """发送消息"""
        content = self.input_box.get("1.0", tk.END).strip()
        if not content and not self.pending_images:
            return

        # 添加用户消息
        user_msg = {
            "role": "user",
            "content": content,
            "images": self.pending_images.copy(),
            "timestamp": datetime.now().isoformat()
        }
        self.messages.append(user_msg)

        # 更新会话标题（如果是第一条消息）
        if len(self.messages) == 1:
            self._update_session_title(content)

        # 清空输入
        self.input_box.delete("1.0", tk.END)
        self.pending_images = []

        # 刷新显示
        self._refresh_messages()
        self._save_current_session()

        # TODO: 调用AI API获取回复
        # 这里添加AI回复逻辑
        self._simulate_ai_response()

    def _log_to_file(self, message):
        """记录日志到文件"""
        try:
            log_file = os.path.join(BASE_DIR, "chat_api.log")
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            with open(log_file, 'a', encoding='utf-8') as f:
                f.write(f"[{timestamp}] {message}\n")
        except Exception as e:
            print(f"日志记录失败: {e}")

    def _show_retrieved_chunks(self, msg_index, search_results):
        """显示召回的文档片段（可折叠）"""
        try:
            if not search_results:
                return
            
            # 获取消息气泡
            if msg_index < len(self.messages_frame.winfo_children()):
                bubble_frame = self.messages_frame.winfo_children()[msg_index]
                if hasattr(bubble_frame, 'ai_content_frame'):
                    ai_frame = bubble_frame.ai_content_frame
                    
                    # 创建召回片段显示组件
                    from retrieved_chunks_view import create_retrieved_chunks_view
                    chunks_view = create_retrieved_chunks_view(ai_frame, search_results)
                    
                    # 在思考过程后面插入
                    if hasattr(ai_frame, 'thinking_section'):
                        chunks_view.frame.pack(after=ai_frame.thinking_section, fill=tk.X, pady=(5, 0))
                    else:
                        chunks_view.frame.pack(fill=tk.X, pady=(5, 0))
                    
                    self._log_to_file(f"显示召回片段: {len(search_results)}个")
        except Exception as e:
            print(f"显示召回片段失败: {e}")
            import traceback
            traceback.print_exc()

    def _call_ai_api_stream(self, user_message, is_thinking=False, rag_context=None):
        """调用火山引擎API获取流式回复（使用配置参数）
        
        Args:
            user_message: 用户消息
            is_thinking: 是否是思考过程
            rag_context: RAG检索上下文（如果已检索过，直接传入避免重复检索）
        """
        # 使用配置中的模型和API Key
        model = self.ai_config.get("model", AI_CHAT_MODEL)
        api_key = self.ai_config.get("api_key", AI_CHAT_API_KEY)
        
        self._log_to_file(f"开始调用API: 模型={model}, 用户消息={user_message[:50]}..., is_thinking={is_thinking}")
        
        try:
            from volcenginesdkarkruntime import Ark

            client = Ark(base_url=AI_CHAT_API_URL, api_key=api_key)
            self._log_to_file("Ark客户端创建成功")

            # 查询RAG知识库获取相关上下文（只在需要且未传入时检索）
            if rag_context is None and RAG_AVAILABLE and not is_thinking and self._rag_kb:
                try:
                    search_results = self._rag_kb.search(user_message, top_k=3)
                    if search_results:
                        rag_context = "\n\n【知识库参考信息】\n"
                        for i, result in enumerate(search_results, 1):
                            rag_context += f"{i}. {result['content'][:300]}...\n"
                        self._log_to_file(f"RAG检索到 {len(search_results)} 条相关知识")
                except Exception as e:
                    self._log_to_file(f"RAG检索失败: {e}")

            # 根据is_thinking从配置中获取系统提示词
            if is_thinking:
                system_prompt = self.ai_config.get("thinking_system_prompt", 
                    "你是一个善于分析的AI助手。请简要分析用户问题的关键点，列出3-5个思考步骤。用简洁的要点形式回答。")
            else:
                system_prompt = self.ai_config.get("response_system_prompt",
                    "你是一个专业的AI助手，擅长回答各种问题。请根据用户的问题提供准确、有用的回答。")
            
            if rag_context:
                system_prompt += f"\n\n在回答时，请参考以下知识库信息：{rag_context}\n请基于以上信息回答用户问题。"

            # 构建input（新API格式）
            input_messages = [
                {"role": "system", "content": system_prompt}
            ]
            
            # 添加历史对话（最近10条）
            for msg in self.messages[-10:]:
                if msg["role"] == "user":
                    input_messages.append({"role": "user", "content": msg["content"]})
                elif msg["role"] == "assistant" and msg.get("content"):
                    input_messages.append({"role": "assistant", "content": msg["content"]})

            self._log_to_file(f"调用API: 模型={model}, 消息数={len(input_messages)}")
            
            # 保存上下文到JSON文件
            self._save_context_to_json(input_messages)

            # 调用API（流式输出）- 使用chat.completions.create
            self._log_to_file("准备调用chat.completions.create...")
            
            backup_model = self.ai_config.get("model_backup", AI_CHAT_MODEL_BACKUP)
            models_try = [model]
            if backup_model and backup_model != model:
                models_try.append(backup_model)
            last_stream_err = None
            for mi, use_model in enumerate(models_try):
                try:
                    label = "主" if mi == 0 else "备"
                    self._log_to_file(f"流式API: {label}接入点 {use_model}")
                    response = client.chat.completions.create(
                        model=use_model,
                        messages=input_messages,
                        stream=True,
                        temperature=self.ai_config.get("temperature", 0.7),
                        max_tokens=self.ai_config.get("max_tokens", 4096),
                        top_p=self.ai_config.get("top_p", 0.9)
                    )
                    self._log_to_file("API调用成功，开始接收流式响应")

                    chunk_count = 0
                    for chunk in response:
                        chunk_count += 1
                        if chunk.choices and len(chunk.choices) > 0:
                            delta = chunk.choices[0].delta
                            if hasattr(delta, 'content') and delta.content:
                                content = delta.content
                                self._log_to_file(f"收到内容 [{chunk_count}]: {content[:50]}...")
                                yield {"type": "content", "content": content}

                    self._log_to_file(f"流式响应结束，共收到 {chunk_count} 个chunk")
                    return
                except Exception as api_error:
                    last_stream_err = api_error
                    self._log_to_file(f"API调用异常 ({use_model}): {api_error}")
                    import traceback
                    self._log_to_file(traceback.format_exc())
            yield {"type": "error", "content": f"API调用失败: {str(last_stream_err)}"}

        except Exception as e:
            error_msg = f"API调用错误: {e}"
            self._log_to_file(error_msg)
            import traceback
            self._log_to_file(traceback.format_exc())
            yield {"type": "error", "content": f"API调用失败: {str(e)}"}

    def _call_ai_api_once(self, messages, temperature: float = 0.1, max_tokens: int = 900):
        """非流式调用 Ark chat.completions，用于工具规划等小请求。"""
        model = self.ai_config.get("model", AI_CHAT_MODEL)
        api_key = self.ai_config.get("api_key", AI_CHAT_API_KEY)

        from volcenginesdkarkruntime import Ark
        client = Ark(base_url=AI_CHAT_API_URL, api_key=api_key)

        backup_model = self.ai_config.get("model_backup", AI_CHAT_MODEL_BACKUP)
        models_try = [model]
        if backup_model and backup_model != model:
            models_try.append(backup_model)

        last_err = None
        for use_model in models_try:
            try:
                resp = client.chat.completions.create(
                    model=use_model,
                    messages=messages,
                    stream=False,
                    temperature=temperature,
                    max_tokens=max_tokens,
                    top_p=self.ai_config.get("top_p", 0.9),
                )
                return (resp.choices[0].message.content or "").strip()
            except Exception as e:
                last_err = e
        raise RuntimeError(f"non-stream call failed: {last_err}")

    def _save_context_to_json(self, messages):
        """保存上下文到JSON文件"""
        try:
            context_file = os.path.join(BASE_DIR, "chat_context.json")
            context_data = {
                "timestamp": datetime.now().isoformat(),
                "model": AI_CHAT_MODEL,
                "message_count": len(messages),
                "messages": messages
            }
            with open(context_file, 'w', encoding='utf-8') as f:
                json.dump(context_data, f, ensure_ascii=False, indent=2)
            self._log_to_file(f"上下文已保存到 {context_file}")
        except Exception as e:
            self._log_to_file(f"保存上下文失败: {e}")

    def _simulate_ai_response(self):
        """调用真实AI API（流式展示效果）- 使用多线程避免卡顿 - ReAct风格"""
        # 获取用户最后一条消息
        user_message = ""
        for msg in reversed(self.messages):
            if msg["role"] == "user":
                user_message = msg["content"]
                break

        # 创建空消息
        ai_msg = {
            "role": "assistant",
            "content": "",
            "thinking": "",
            "timestamp": datetime.now().isoformat()
        }
        self.messages.append(ai_msg)
        self._refresh_messages()

        msg_index = len(self.messages) - 1

        # 在新线程中执行流式输出，避免卡住主界面
        def stream_in_thread():
            try:
                # 初始化ReAct思考链步骤引用
                intent_step = None
                rewrite_step = None
                retrieve_step = None
                reason_step = None
                
                # ========== Step 1: 意图识别 ==========
                intent_result = None
                
                if REACT_CHAIN_AVAILABLE and RAG_TOOLS_AVAILABLE and self._rag_tool:
                    # 添加意图识别步骤到思考链
                    def add_intent_step():
                        nonlocal intent_step
                        bubble_frame = self.messages_frame.winfo_children()[msg_index]
                        if hasattr(bubble_frame, 'ai_content_frame'):
                            ai_frame = bubble_frame.ai_content_frame
                            if hasattr(ai_frame, 'thought_chain'):
                                intent_step = ai_frame.thought_chain.add_step(
                                    "intent", "意图识别", "正在分析用户意图...", 
                                    status="running"
                                )
                    self.after(0, add_intent_step)
                    
                    # 执行意图识别
                    intent_result = self._rag_tool.intent_recognizer.recognize(user_message, use_llm=True)
                    self._log_to_file(f"意图识别结果: {intent_result.intent.value}, 需要RAG: {intent_result.needs_rag}")
                    
                    # 更新意图识别步骤
                    def update_intent_step():
                        if intent_step:
                            intent_step.update_content(
                                f"意图: {intent_result.intent.value}\n"
                                f"置信度: {intent_result.confidence:.2f}\n"
                                f"需要RAG: {intent_result.needs_rag}\n"
                                f"原因: {intent_result.reason}"
                            )
                            intent_step.update_status("completed")
                    self.after(0, update_intent_step)
                    
                    # 如果不需要RAG，跳过后续步骤
                    if not intent_result.needs_rag:
                        self._log_to_file("意图识别：无需RAG，直接回答")
                
                # ========== Step 2: Query改写 ==========
                rewritten_query = user_message
                query_keywords = []
                suggested_tags = None
                
                if REACT_CHAIN_AVAILABLE and RAG_TOOLS_AVAILABLE and self._rag_tool and \
                   (intent_result is None or intent_result.needs_rag):
                    # 添加Query改写步骤
                    def add_rewrite_step():
                        nonlocal rewrite_step
                        bubble_frame = self.messages_frame.winfo_children()[msg_index]
                        if hasattr(bubble_frame, 'ai_content_frame'):
                            ai_frame = bubble_frame.ai_content_frame
                            if hasattr(ai_frame, 'thought_chain'):
                                rewrite_step = ai_frame.thought_chain.add_step(
                                    "rewrite", "Query改写", "正在改写查询...", 
                                    status="running"
                                )
                    self.after(0, add_rewrite_step)
                    
                    # 执行Query改写
                    try:
                        rewriter = QueryRewriter(llm_client=self._rag_tool.intent_recognizer.llm_client)
                        rewrite_result = rewriter.rewrite(user_message, intent_result)
                        
                        rewritten_query = rewrite_result.rewritten_query
                        query_keywords = rewrite_result.keywords
                        suggested_tags = rewrite_result.suggested_tags
                        
                        self._log_to_file(f"Query改写: {user_message[:50]}... -> {rewritten_query[:50]}...")
                        
                        # 更新Query改写步骤
                        def update_rewrite_step():
                            if rewrite_step:
                                content = f"原始查询: {user_message}\n"
                                content += f"改写后: {rewritten_query}\n"
                                if query_keywords:
                                    content += f"关键词: {', '.join(query_keywords)}\n"
                                if suggested_tags:
                                    content += f"建议标签: {suggested_tags.domain}/{suggested_tags.module}/{suggested_tags.doc_type}"
                                
                                rewrite_step.update_content(content)
                                rewrite_step.update_status("completed")
                        self.after(0, update_rewrite_step)
                        
                        # 如果需要用户澄清
                        if rewrite_result.needs_clarification and rewrite_result.clarification_question:
                            def add_clarification_step():
                                if rewrite_step:
                                    rewrite_step.update_content(
                                        rewrite_step.thought.content + f"\n\n需要澄清: {rewrite_result.clarification_question}"
                                    )
                            self.after(0, add_clarification_step)
                            
                    except Exception as e:
                        self._log_to_file(f"Query改写失败: {e}")
                        def update_rewrite_error():
                            if rewrite_step:
                                rewrite_step.update_content(f"改写失败: {str(e)}\n使用原始查询")
                                rewrite_step.update_status("failed")
                        self.after(0, update_rewrite_error)
                
                # ========== Step 3: 知识检索 ==========
                search_results = []
                rag_context = ""
                
                if REACT_CHAIN_AVAILABLE and RAG_TOOLS_AVAILABLE and self._rag_tool and \
                   (intent_result is None or intent_result.needs_rag):
                    # 添加知识检索步骤
                    def add_retrieve_step():
                        nonlocal retrieve_step
                        bubble_frame = self.messages_frame.winfo_children()[msg_index]
                        if hasattr(bubble_frame, 'ai_content_frame'):
                            ai_frame = bubble_frame.ai_content_frame
                            if hasattr(ai_frame, 'thought_chain'):
                                retrieve_step = ai_frame.thought_chain.add_step(
                                    "retrieve", "知识检索", "正在检索知识库...", 
                                    status="running"
                                )
                    self.after(0, add_retrieve_step)
                    
                    try:
                        # 使用改写后的查询进行检索
                        import concurrent.futures
                        
                        def do_rag_search():
                            # 使用改写后的查询
                            return self._rag_tool.search(
                                rewritten_query, 
                                top_k=3,
                                skip_intent=True
                            )
                        
                        # 在单独的线程中执行RAG检索，带超时
                        with concurrent.futures.ThreadPoolExecutor() as executor:
                            future = executor.submit(do_rag_search)
                            try:
                                _, retrieved_chunks = future.result(timeout=60)
                            except concurrent.futures.TimeoutError:
                                self._log_to_file("RAG检索超时")
                                retrieved_chunks = []
                        
                        # 转换为旧格式兼容
                        search_results = []
                        for chunk in retrieved_chunks:
                            search_results.append({
                                'content': chunk.content,
                                'source_file': chunk.doc_id,
                                'metadata': chunk.metadata.to_dict(),
                                'score': chunk.similarity
                            })
                        
                        # 更新知识检索步骤
                        def update_retrieve_step():
                            if retrieve_step:
                                if search_results:
                                    content = f"使用查询: {rewritten_query[:100]}...\n"
                                    content += f"检索到 {len(search_results)} 条相关知识:\n"
                                    for i, result in enumerate(search_results, 1):
                                        content += f"{i}. [{result['score']:.3f}] {result['content'][:80]}...\n"
                                    retrieve_step.update_content(content)
                                    retrieve_step.update_status("completed")
                                else:
                                    retrieve_step.update_content("未找到相关知识")
                                    retrieve_step.update_status("completed")
                        self.after(0, update_retrieve_step)
                        
                        # 构建RAG上下文
                        if search_results:
                            rag_context = "\n\n【知识库参考信息】\n"
                            for i, result in enumerate(search_results, 1):
                                rag_context += f"{i}. {result['content'][:300]}...\n"
                            self._log_to_file(f"RAG检索到 {len(search_results)} 条相关知识")
                            
                            # 显示召回片段
                            self.after(0, lambda: self._show_retrieved_chunks(msg_index, search_results))
                            
                    except Exception as e:
                        self._log_to_file(f"RAG检索失败: {e}")
                        def update_retrieve_error():
                            if retrieve_step:
                                retrieve_step.update_content(f"检索失败: {str(e)}")
                                retrieve_step.update_status("failed")
                        self.after(0, update_retrieve_error)
                
                # ========== Step 4: 推理分析 ==========
                if REACT_CHAIN_AVAILABLE:
                    def add_reason_step():
                        nonlocal reason_step
                        bubble_frame = self.messages_frame.winfo_children()[msg_index]
                        if hasattr(bubble_frame, 'ai_content_frame'):
                            ai_frame = bubble_frame.ai_content_frame
                            if hasattr(ai_frame, 'thought_chain'):
                                reason_step = ai_frame.thought_chain.add_step(
                                    "reason", "推理分析", "正在分析问题...", 
                                    status="running"
                                )
                    self.after(0, add_reason_step)
                
                # 生成思考过程
                thinking_content = ""
                thinking_prompt = f"请分析这个问题，简要说明你的思考步骤（3-5点）：{user_message}"
                
                for chunk in self._call_ai_api_stream(thinking_prompt, is_thinking=True):
                    if chunk["type"] == "content":
                        thinking_content += chunk["content"]
                        ai_msg["thinking"] = thinking_content
                        
                        # 实时更新推理步骤
                        if REACT_CHAIN_AVAILABLE and reason_step:
                            def update_reason(content):
                                if reason_step:
                                    reason_step.update_content(content)
                            self.after(0, lambda c=thinking_content: update_reason(c))
                            
                    elif chunk["type"] == "error":
                        thinking_content = "思考过程生成失败"
                        ai_msg["thinking"] = thinking_content
                        break
                
                # 完成推理步骤
                if REACT_CHAIN_AVAILABLE and reason_step:
                    def complete_reason_step():
                        if reason_step:
                            reason_step.update_content(thinking_content)
                            reason_step.update_status("completed")
                    self.after(0, complete_reason_step)
                
                self._log_to_file(f"思考过程完成: {thinking_content[:100]}...")

                # ========== Step 4.5: 工具调用（Cursor风格） ==========
                tool_ctx = ""
                tool_step = None
                if CURSOR_TOOLS_AVAILABLE and self._cursor_tools:
                    if REACT_CHAIN_AVAILABLE:
                        def add_tool_step():
                            nonlocal tool_step
                            bubble_frame = self.messages_frame.winfo_children()[msg_index]
                            if hasattr(bubble_frame, 'ai_content_frame'):
                                ai_frame = bubble_frame.ai_content_frame
                                if hasattr(ai_frame, 'thought_chain'):
                                    tool_step = ai_frame.thought_chain.add_step(
                                        "tools", "工具调用", "正在判断是否需要调用工具...", status="running"
                                    )
                        self.after(0, add_tool_step)

                    try:
                        def _llm_call(msgs):
                            return self._call_ai_api_once(msgs, temperature=0.1, max_tokens=900)

                        calls = plan_tool_calls_with_llm(
                            llm_call=_llm_call,
                            tools=self._cursor_tools,
                            user_message=user_message,
                            max_calls=3,
                        )
                        results = execute_tool_calls(self._cursor_tools, calls)
                        tool_ctx = format_tool_results_for_context(results)

                        if REACT_CHAIN_AVAILABLE and tool_step:
                            def update_tool_step():
                                if tool_step:
                                    if results:
                                        brief = "\n".join([f"- {r['name']}: {'OK' if r['ok'] else 'FAIL'}" for r in results])
                                        tool_step.update_content(f"已执行工具：\n{brief}")
                                    else:
                                        tool_step.update_content("本轮无需调用工具。")
                                    tool_step.update_status("completed")
                            self.after(0, update_tool_step)
                    except Exception as e:
                        if REACT_CHAIN_AVAILABLE and tool_step:
                            def update_tool_step_fail():
                                if tool_step:
                                    tool_step.update_content(f"工具调用失败：{type(e).__name__}: {e}")
                                    tool_step.update_status("failed")
                            self.after(0, update_tool_step_fail)
                
                # ========== Step 5: 生成回答 ==========
                self._log_to_file("开始生成回答...")

                rag_context_final = rag_context + (tool_ctx or "")
                # 标准 LangChain 架构：模型+工具循环+memory 统一交给 AgentExecutor
                if self._standard_runtime and self._standard_runtime.ready:
                    augmented_input = user_message
                    if rag_context_final:
                        augmented_input += f"\n\n请参考上下文：\n{rag_context_final}"
                    result = self._standard_runtime.invoke(augmented_input)
                    ai_msg["content"] = result.get("output", "") if result.get("ok") else result.get("output", "标准运行时失败")
                    self.after(0, lambda t=ai_msg["content"]: self._stream_update_content(msg_index, t))
                else:
                    # 兼容旧链路
                    for chunk in self._call_ai_api_stream(user_message, is_thinking=False, rag_context=rag_context_final):
                        if chunk["type"] == "content":
                            ai_msg["content"] += chunk["content"]
                            self.after(0, lambda t=ai_msg["content"]: self._stream_update_content(msg_index, t))
                        elif chunk["type"] == "error":
                            ai_msg["content"] = chunk["content"]
                            self.after(0, lambda t=ai_msg["content"]: self._stream_update_content(msg_index, t))

                # 保存会话
                self.after(0, self._save_current_session)

            except Exception as e:
                print(f"流式输出错误: {e}")
                import traceback
                traceback.print_exc()
                ai_msg["content"] = f"发生错误: {str(e)}"
                self.after(0, lambda t=ai_msg["content"]: self._stream_update_content(msg_index, t))

        # 启动后台线程
        threading.Thread(target=stream_in_thread, daemon=True).start()

    def _on_enter_pressed(self, event):
        """回车发送"""
        self._send_message()
        return "break"

    def _on_shift_enter(self, event):
        """Shift+Enter换行"""
        pass

    def _upload_image(self):
        """上传图片"""
        file_path = filedialog.askopenfilename(
            title="选择图片",
            filetypes=[("图片文件", "*.png *.jpg *.jpeg *.gif *.bmp")]
        )
        if file_path:
            self.pending_images.append(file_path)
            messagebox.showinfo("提示", f"已选择图片: {os.path.basename(file_path)}")

    def _save_current_session(self):
        """保存当前会话"""
        for session in self.sessions:
            if session["id"] == self.current_session_id:
                session["messages"] = self.messages
                break
        self._save_sessions()

    def _save_sessions(self):
        """保存所有会话到文件"""
        try:
            sessions_file = os.path.join(BASE_DIR, "chat_sessions.json")
            with open(sessions_file, 'w', encoding='utf-8') as f:
                json.dump(self.sessions, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"保存会话失败: {e}")

    def _load_sessions(self):
        """从文件加载会话"""
        try:
            sessions_file = os.path.join(BASE_DIR, "chat_sessions.json")
            if os.path.exists(sessions_file):
                with open(sessions_file, 'r', encoding='utf-8') as f:
                    self.sessions = json.load(f)
        except Exception as e:
            print(f"加载会话失败: {e}")
            self.sessions = []

    def _open_rag_manager(self):
        """打开 RAG 知识库管理窗口（独立窗口）"""
        try:
            # 导入 RAG 管理器
            from rag_manager_gui import RAGManagerGUI
            
            # 检查RAG是否已初始化
            if not self._rag_kb:
                messagebox.showerror("错误", "RAG知识库未初始化，请重启程序")
                return
            
            # 创建独立窗口，传入已初始化的RAG
            rag_manager = RAGManagerGUI(self, rag_kb=self._rag_kb)
            
            print("RAG 知识库管理窗口已打开")
        except Exception as e:
            messagebox.showerror("错误", f"打开知识库管理窗口失败：{e}")
            import traceback
            traceback.print_exc()
    
    def _add_file_to_kb(self):
        """添加文件到知识库（支持 TXT 和 MD）"""
        if not RAG_AVAILABLE:
            messagebox.showerror("错误", "RAG 知识库模块未安装")
            return
        
        file_path = filedialog.askopenfilename(
            title="选择文件",
            filetypes=[
                ("文本文件", "*.txt"),
                ("Markdown 文件", "*.md"),
                ("所有支持的文件", "*.txt *.md"),
                ("所有文件", "*.*")
            ]
        )
        
        if file_path:
            try:
                if not self._rag_kb:
                    messagebox.showerror("错误", "RAG知识库未初始化")
                    return
                
                success = self._rag_kb.add_document(file_path)
                if success:
                    messagebox.showinfo("成功", f"文件已添加到知识库:\n{os.path.basename(file_path)}")
                else:
                    messagebox.showerror("错误", "添加文件失败")
            except Exception as e:
                messagebox.showerror("错误", f"添加文件失败: {e}")

    def _add_folder_to_kb(self):
        """添加文件夹到知识库"""
        if not RAG_AVAILABLE:
            messagebox.showerror("错误", "RAG知识库模块未安装")
            return
        
        folder_path = filedialog.askdirectory(title="选择文件夹")
        
        if folder_path:
            try:
                if not self._rag_kb:
                    messagebox.showerror("错误", "RAG知识库未初始化")
                    return
                
                # 遍历文件夹中的所有文件
                added_count = 0
                total_files = 0
                for root, dirs, files in os.walk(folder_path):
                    for file in files:
                        if file.endswith(('.txt', '.md')):
                            total_files += 1
                            file_path = os.path.join(root, file)
                            try:
                                if self._rag_kb.add_document(file_path):
                                    added_count += 1
                            except Exception as e:
                                print(f"添加文件失败 {file_path}: {e}")
                
                messagebox.showinfo("成功", f"已添加 {added_count}/{total_files} 个文件到知识库")
            except Exception as e:
                messagebox.showerror("错误", f"添加文件夹失败: {e}")

    def _rebuild_kb_index(self):
        """重建知识库索引"""
        if not RAG_AVAILABLE:
            messagebox.showerror("错误", "RAG知识库模块未安装")
            return
        
        try:
            if not self._rag_kb:
                messagebox.showerror("错误", "RAG知识库未初始化")
                return
            
            # 重新加载索引
            self._rag_kb.load_index()
            
            # 获取统计信息
            doc_count = len(self._rag_kb.chunks)
            messagebox.showinfo("成功", f"索引重建完成！\n当前知识库共有 {doc_count} 个文档片段")
        except Exception as e:
            messagebox.showerror("错误", f"重建索引失败：{e}")
    
    def _open_rag_manager(self):
        """打开 RAG 知识库管理窗口（独立窗口）"""
        try:
            # 导入 RAG 管理器
            from rag_manager_gui import RAGManagerGUI
            
            # 检查RAG是否已初始化
            if not self._rag_kb:
                messagebox.showerror("错误", "RAG知识库未初始化，请重启程序")
                return
            
            # 创建独立窗口，传入已初始化的RAG
            rag_manager = RAGManagerGUI(self, rag_kb=self._rag_kb)
            
            print("RAG 知识库管理窗口已打开")
        except Exception as e:
            messagebox.showerror("错误", f"打开知识库管理窗口失败：{e}")
            import traceback
            traceback.print_exc()

    def _open_rag_manager(self):
        """打开 RAG 知识库管理窗口（独立窗口）"""
        try:
            # 导入 RAG 管理器
            from rag_manager_gui import RAGManagerGUI
            
            # 检查RAG是否已初始化
            if not self._rag_kb:
                messagebox.showerror("错误", "RAG知识库未初始化，请重启程序")
                return
            
            # 创建独立窗口，传入已初始化的RAG
            rag_manager = RAGManagerGUI(self, rag_kb=self._rag_kb)
            
            print("RAG 知识库管理窗口已打开")
        except Exception as e:
            messagebox.showerror("错误", f"打开知识库管理窗口失败：{e}")
            import traceback
            traceback.print_exc()
    
    def _open_settings(self):
        """打开AI配置界面（COZE风格）"""
        settings_window = tk.Toplevel(self)
        settings_window.title("AI配置")
        settings_window.geometry("700x600")
        settings_window.configure(bg="#f5f5f5")
        
        # 居中显示
        settings_window.transient(self)
        settings_window.grab_set()
        
        # 主容器
        main_frame = tk.Frame(settings_window, bg="#f5f5f5")
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # 标题
        title_label = tk.Label(
            main_frame,
            text="🤖 AI助手配置",
            font=("微软雅黑", 18, "bold"),
            bg="#f5f5f5",
            fg="#1a1a1a"
        )
        title_label.pack(anchor="w", pady=(0, 20))
        
        # 创建Notebook（标签页）
        notebook = ttk.Notebook(main_frame)
        notebook.pack(fill=tk.BOTH, expand=True)
        
        # === 标签页1: 系统提示词 ===
        prompt_frame = tk.Frame(notebook, bg="#ffffff", padx=20, pady=20)
        notebook.add(prompt_frame, text="系统提示词")
        
        # 思考过程提示词
        tk.Label(
            prompt_frame,
            text="思考过程提示词:",
            font=("微软雅黑", 11, "bold"),
            bg="#ffffff",
            fg="#333333"
        ).pack(anchor="w", pady=(0, 5))
        
        thinking_text = scrolledtext.ScrolledText(
            prompt_frame,
            height=6,
            font=("Consolas", 10),
            wrap=tk.WORD,
            bg="#fafafa",
            fg="#333333"
        )
        thinking_text.pack(fill=tk.X, pady=(0, 15))
        thinking_text.insert(tk.END, self.ai_config.get("thinking_system_prompt", ""))
        
        # 回复提示词
        tk.Label(
            prompt_frame,
            text="回复提示词:",
            font=("微软雅黑", 11, "bold"),
            bg="#ffffff",
            fg="#333333"
        ).pack(anchor="w", pady=(0, 5))
        
        response_text = scrolledtext.ScrolledText(
            prompt_frame,
            height=6,
            font=("Consolas", 10),
            wrap=tk.WORD,
            bg="#fafafa",
            fg="#333333"
        )
        response_text.pack(fill=tk.X, pady=(0, 15))
        response_text.insert(tk.END, self.ai_config.get("response_system_prompt", ""))
        
        # === 标签页2: 模型参数 ===
        param_frame = tk.Frame(notebook, bg="#ffffff", padx=20, pady=20)
        notebook.add(param_frame, text="模型参数")
        
        # Temperature
        tk.Label(
            param_frame,
            text="Temperature (创造性):",
            font=("微软雅黑", 11),
            bg="#ffffff",
            fg="#333333"
        ).pack(anchor="w", pady=(0, 5))
        
        temp_scale = tk.Scale(
            param_frame,
            from_=0.0,
            to=2.0,
            resolution=0.1,
            orient=tk.HORIZONTAL,
            bg="#ffffff",
            highlightthickness=0
        )
        temp_scale.set(self.ai_config.get("temperature", 0.7))
        temp_scale.pack(fill=tk.X, pady=(0, 15))
        
        # Max Tokens
        tk.Label(
            param_frame,
            text="Max Tokens (最大长度):",
            font=("微软雅黑", 11),
            bg="#ffffff",
            fg="#333333"
        ).pack(anchor="w", pady=(0, 5))
        
        max_tokens_entry = tk.Entry(
            param_frame,
            font=("Consolas", 11),
            bg="#fafafa",
            fg="#333333"
        )
        max_tokens_entry.pack(fill=tk.X, pady=(0, 15))
        max_tokens_entry.insert(0, str(self.ai_config.get("max_tokens", 4096)))
        
        # Top P
        tk.Label(
            param_frame,
            text="Top P (多样性):",
            font=("微软雅黑", 11),
            bg="#ffffff",
            fg="#333333"
        ).pack(anchor="w", pady=(0, 5))
        
        top_p_scale = tk.Scale(
            param_frame,
            from_=0.0,
            to=1.0,
            resolution=0.1,
            orient=tk.HORIZONTAL,
            bg="#ffffff",
            highlightthickness=0
        )
        top_p_scale.set(self.ai_config.get("top_p", 0.9))
        top_p_scale.pack(fill=tk.X, pady=(0, 15))
        
        # === 标签页3: API配置 ===
        api_frame = tk.Frame(notebook, bg="#ffffff", padx=20, pady=20)
        notebook.add(api_frame, text="API配置")
        
        # API Key
        tk.Label(
            api_frame,
            text="API Key:",
            font=("微软雅黑", 11),
            bg="#ffffff",
            fg="#333333"
        ).pack(anchor="w", pady=(0, 5))
        
        api_key_entry = tk.Entry(
            api_frame,
            font=("Consolas", 11),
            bg="#fafafa",
            fg="#333333",
            show="*"
        )
        api_key_entry.pack(fill=tk.X, pady=(0, 15))
        api_key_entry.insert(0, self.ai_config.get("api_key", ""))
        
        # Model
        tk.Label(
            api_frame,
            text="Model (接入点ID):",
            font=("微软雅黑", 11),
            bg="#ffffff",
            fg="#333333"
        ).pack(anchor="w", pady=(0, 5))
        
        model_entry = tk.Entry(
            api_frame,
            font=("Consolas", 11),
            bg="#fafafa",
            fg="#333333"
        )
        model_entry.pack(fill=tk.X, pady=(0, 15))
        model_entry.insert(0, self.ai_config.get("model", ""))
        
        # 底部按钮
        btn_frame = tk.Frame(main_frame, bg="#f5f5f5")
        btn_frame.pack(fill=tk.X, pady=(20, 0))
        
        def save_settings():
            """保存设置"""
            self.ai_config["thinking_system_prompt"] = thinking_text.get(1.0, tk.END).strip()
            self.ai_config["response_system_prompt"] = response_text.get(1.0, tk.END).strip()
            self.ai_config["temperature"] = temp_scale.get()
            self.ai_config["max_tokens"] = int(max_tokens_entry.get())
            self.ai_config["top_p"] = top_p_scale.get()
            self.ai_config["api_key"] = api_key_entry.get().strip()
            self.ai_config["model"] = model_entry.get().strip()
            
            self._save_ai_config()
            messagebox.showinfo("成功", "配置已保存！")
            settings_window.destroy()
        
        def reset_defaults():
            """恢复默认设置"""
            thinking_text.delete(1.0, tk.END)
            thinking_text.insert(tk.END, self.DEFAULT_AI_CONFIG["thinking_system_prompt"])
            response_text.delete(1.0, tk.END)
            response_text.insert(tk.END, self.DEFAULT_AI_CONFIG["response_system_prompt"])
            temp_scale.set(self.DEFAULT_AI_CONFIG["temperature"])
            max_tokens_entry.delete(0, tk.END)
            max_tokens_entry.insert(0, str(self.DEFAULT_AI_CONFIG["max_tokens"]))
            top_p_scale.set(self.DEFAULT_AI_CONFIG["top_p"])
            api_key_entry.delete(0, tk.END)
            api_key_entry.insert(0, self.DEFAULT_AI_CONFIG["api_key"])
            model_entry.delete(0, tk.END)
            model_entry.insert(0, self.DEFAULT_AI_CONFIG["model"])
        
        tk.Button(
            btn_frame,
            text="保存",
            command=save_settings,
            bg="#4e6ef2",
            fg="#ffffff",
            font=("微软雅黑", 11),
            width=10,
            cursor="hand2"
        ).pack(side=tk.RIGHT, padx=(10, 0))
        
        tk.Button(
            btn_frame,
            text="恢复默认",
            command=reset_defaults,
            bg="#ffffff",
            fg="#666666",
            font=("微软雅黑", 11),
            width=10,
            cursor="hand2"
        ).pack(side=tk.RIGHT, padx=(10, 0))
        
        tk.Button(
            btn_frame,
            text="取消",
            command=settings_window.destroy,
            bg="#ffffff",
            fg="#666666",
            font=("微软雅黑", 11),
            width=10,
            cursor="hand2"
        ).pack(side=tk.RIGHT)


# ==================== 主应用类 ====================
# 加载配置
CONFIG = load_config()

PLATFORMS = {
    "小红书": {
        "api_endpoint": "https://www.hellotik.app/zh/rednote",
        "headers": {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Accept': 'application/json, text/plain, */*',
            'Content-Type': 'application/json',
            'Referer': 'https://hellotik.app/',
            'Origin': 'https://hellotik.app'
        },
        "payload": {
            "requestURL": "{url}",
            "isMobile": "false",
            "isoCode": "HK",
            "adType": "adsense",
            "uwx_id": "uwx_350696y5juIO",
            "successCount": "0",
            "totalSuccessCount": "2",
            "firstSuccessDate": "2026-01-10",
            "time": "{timestamp}",
            "key": "xaq8pkc7"
        },
        "url_key_candidates": ["video_url", "download_url", "url"]
    }
}

class App:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title(APP_TITLE)
        self.root.geometry("1200x800")
        self.root.configure(bg=UI_BG)
        # 允许窗口大小调整
        self.root.resizable(True, True)

        self.link_var = tk.StringVar()
        
        # 队列系统初始化
        self.task_queue = []
        self._task_queue_lock = threading.Lock()
        self._scheduler_start_lock = threading.Lock()
        self.processing_queue = False
        self.current_task_index = 0
        self.queue_max_size = 50  # 默认队列最大大小
        
        # 任务取消标志 - 用于停止正在执行的任务
        self.task_cancel_flags = {}  # {link: cancel_event}
        
        # 历史记录初始化
        self.history = load_history()
        
        # 线程池初始化
        self.cpu_count = multiprocessing.cpu_count()
        # 从配置文件中加载线程数量，如果不存在则使用默认值
        self.max_workers = CONFIG.get("max_workers", min(self.cpu_count, 8))
        # 确保线程数量在合理范围内
        self.max_workers = max(1, min(self.max_workers, self.cpu_count))
        self.executor = concurrent.futures.ThreadPoolExecutor(max_workers=self.max_workers)
        self.active_futures = {}  # 存储活跃的任务未来对象 {link: future}
        self._pipeline_log_lock = threading.Lock()  # pipeline.log 线程安全
        # 运维 Agent：日志/事件上报去重（避免 ERROR 刷屏重复出报告）
        self._ops_incident_lock = threading.Lock()
        self._ops_incident_last_ts = {}  # fingerprint -> monotonic time
        self._ops_incident_cooldown = float(os.environ.get("OPS_INCIDENT_COOLDOWN_SEC", "120"))

        # 缓存机制初始化
        self.model_cache = None  # Whisper 模型缓存（主实例）
        self.model_cache_lock = threading.Lock()  # 模型缓存锁
        self.video_cache = {}  # 视频缓存，键为链接，值为本地文件路径
        self.video_cache_lock = threading.Lock()  # 视频缓存锁
        self.file_operation_lock = threading.Lock()  # 文件操作锁，防止文件竞争
        self.file_counter_lock = threading.Lock()  # 文件计数器锁，确保文件名唯一
        
        # Whisper：多实例池（每槽位独立 whisper.load_model，可真正并行 transcribe；单实例共享模型不安全）
        self._whisper_pool_queue = None  # queue.Queue，元素为 (slot_id, model)
        self._whisper_pool_init_lock = threading.Lock()
        self._whisper_pool_model_name = os.environ.get("WHISPER_MODEL", "tiny").strip() or "tiny"

        # 【优化】程序启动时初始化RAG知识库（只初始化一次）
        # 优先使用新的RAG工具系统（包含意图识别和元数据管理）
        self._rag_kb = None
        self._rag_tool = None
        
        if RAG_TOOLS_AVAILABLE and KB_MANAGER_AVAILABLE:
            try:
                print("[RAG] 程序启动，正在初始化新的RAG工具系统（意图识别+元数据）...")
                kb = get_knowledge_base()
                from rag_tools import RAGTool, IntentRecognizer
                intent_recognizer = IntentRecognizer(llm_client=None)  # 暂时不使用LLM
                self._rag_tool = RAGTool(kb, intent_recognizer)
                self._rag_kb = kb  # 向后兼容
                stats = kb.get_stats()
                print(f"[RAG] RAG工具系统初始化成功: {stats['total_chunks']} chunks")
            except Exception as e:
                print(f"[RAG] RAG工具系统初始化失败：{e}")
                import traceback
                traceback.print_exc()
                self._rag_tool = None
                self._rag_kb = None
        
        # 向后兼容：如果新的不可用，使用旧的kb_manager
        if self._rag_kb is None and KB_MANAGER_AVAILABLE:
            try:
                print("[RAG] 程序启动，正在初始化新的知识库管理器...")
                self._rag_kb = get_knowledge_base()
                stats = self._rag_kb.get_stats()
                print(f"[RAG] 知识库初始化成功: {stats['total_chunks']} chunks, model_loaded={stats['model_loaded']}")
            except Exception as e:
                print(f"[RAG] 新的知识库管理器初始化失败：{e}")
                self._rag_kb = None
        
        # 向后兼容：如果新的不可用，使用旧的RAG
        if self._rag_kb is None and RAG_AVAILABLE:
            try:
                print("[RAG] 程序启动，正在初始化RAG知识库（旧版）...")
                self._rag_kb = RAGKnowledgeBase()
                print("[RAG] 知识库初始化成功")
            except Exception as e:
                print(f"[RAG] 知识库初始化失败：{e}")
                self._rag_kb = None

        # 先构建UI，确保所有UI组件都已创建
        self._build_ui()
        
        # 运维Agent初始化（必须在UI构建之后，因为需要使用append_log）
        if OPS_AGENT_AVAILABLE:
            try:
                # 使用备用API配置（Doubao-Seed-2.0-mini）
                self.ops_agent = create_ops_agent(
                    api_key=CONFIG.get("volcengine_api_key") or AI_CHAT_API_KEY,
                    api_model=CONFIG.get("ai_chat_model") or AI_CHAT_MODEL,
                )
                self.append_log("运维Agent初始化完成")
            except Exception as e:
                self.append_log(f"运维Agent初始化失败: {e}")
                self.ops_agent = None
        else:
            self.ops_agent = None
        
        # 自动恢复未完成任务
        self.recover_unfinished_tasks()
        
        # 显示系统信息
        self.append_log(f"系统CPU核心数：{self.cpu_count}")
        self.append_log(f"线程池最大工作线程数：{self.max_workers}")
        
        # 绑定窗口关闭事件
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)

    def _build_ui(self):
        # 设置样式
        style = ttk.Style()
        style.theme_use("clam")
        
        # 自定义样式
        style.configure(
            "TButton",
            padding=(10, 5),
            font=UI_FONT,
            foreground=UI_ACCENT,
            background=UI_CARD,
            borderwidth=0,
        )
        style.map(
            "TButton",
            foreground=[("disabled", UI_TEXT_LIGHT), ("pressed", UI_ACCENT), ("active", UI_ACCENT)],
            background=[("pressed", UI_ACCENT_SOFT), ("active", UI_ACCENT_SOFT)],
        )
        style.configure("TLabel", font=UI_FONT, foreground=UI_TEXT)
        style.configure("TEntry", padding=(8, 5), font=UI_FONT)
        style.configure("TLabelframe", font=UI_FONT_BOLD, foreground=UI_TEXT)
        
        # 创建圆角输入框样式
        style.configure(
            "Rounded.TEntry",
            padding=(10, 6),
            font=UI_FONT,
            borderwidth=0,
            relief="flat"
        )
        
        # 主容器
        main_container = tk.Frame(self.root, bg=UI_BG)
        main_container.pack(fill=tk.BOTH, expand=True, padx=16, pady=16)
        
        # 顶部标题区域
        title_frame = tk.Frame(main_container, bg=UI_BG)
        title_frame.pack(fill=tk.X, pady=(0, 12))
        
        title_label = tk.Label(
            title_frame,
            text="视频转文字处理工具",
            font=UI_FONT_TITLE,
            foreground=UI_TEXT,
            bg=UI_BG,
        )
        title_label.pack(anchor=tk.W)
        
        subtitle_label = tk.Label(
            title_frame,
            text="智能视频分析与文本转换",
            font=UI_FONT,
            foreground=UI_TEXT_MUTED,
            bg=UI_BG,
        )
        subtitle_label.pack(anchor=tk.W, pady=(4, 0))
        
        # 导航：与背景同色的扁平分段，选中项用浅色底 + 强调色文字
        nav_frame = tk.Frame(main_container, bg=UI_BG, bd=0, highlightthickness=0)
        nav_frame.pack(fill=tk.X, pady=(0, 12))
        
        nav_container = tk.Frame(nav_frame, bg=UI_BG)
        nav_container.pack(fill=tk.X, padx=0, pady=4)
        
        nav_btn_style = {
            "font": UI_FONT_NAV,
            "padx": 18,
            "pady": 6,
            "bd": 0,
            "relief": tk.FLAT,
            "cursor": "hand2",
            "activebackground": UI_ACCENT_SOFT,
            "activeforeground": UI_ACCENT,
        }
        
        # 视频处理页面按钮（默认选中）
        self.nav_video_btn = tk.Button(
            nav_container,
            text="视频处理",
            command=self._show_video_page,
            bg=UI_CARD,
            fg=UI_ACCENT,
            **nav_btn_style,
        )
        self.nav_video_btn.pack(side=tk.LEFT, padx=(0, 6))
        
        # AI问答页面按钮
        self.nav_chat_btn = tk.Button(
            nav_container,
            text="AI 问答",
            command=self._show_chat_page,
            bg=UI_BG,
            fg=UI_TEXT_MUTED,
            **nav_btn_style,
        )
        self.nav_chat_btn.pack(side=tk.LEFT, padx=(0, 6))
        
        # 多模态文档处理页面按钮
        self.nav_multimodal_btn = tk.Button(
            nav_container,
            text="文档处理",
            command=self._show_multimodal_page,
            bg=UI_BG,
            fg=UI_TEXT_MUTED,
            **nav_btn_style,
        )
        self.nav_multimodal_btn.pack(side=tk.LEFT, padx=(0, 6))
        
        # 页面内容容器
        self.content_frame = tk.Frame(main_container, bg=UI_BG)
        self.content_frame.pack(fill=tk.BOTH, expand=True)
        
        # 创建视频处理页面
        self.video_page = tk.Frame(self.content_frame, bg=UI_BG)
        self.video_page.pack(fill=tk.BOTH, expand=True)
        
        # 创建AI问答页面（初始隐藏）
        self.chat_page = tk.Frame(self.content_frame, bg=UI_BG)
        
        # 创建多模态文档处理页面（初始隐藏）
        self.multimodal_page = None
        
        # 构建视频处理页面内容
        self._build_video_page(self.video_page)
        
        # 构建AI问答页面内容
        self._build_chat_page(self.chat_page)
        
    def _build_video_page(self, parent):
        """构建视频处理页面"""
        # 核心功能区域
        core_frame = tk.Frame(parent, bg=UI_BG)
        core_frame.pack(fill=tk.X, pady=(0, 12))
        
        # 视频链接输入区域
        link_frame = tk.Frame(core_frame, bg=UI_CARD, bd=0, relief=tk.FLAT)
        link_frame.pack(fill=tk.X, padx=0, pady=(0, 8))
        link_frame.configure(
            bg=UI_CARD,
            highlightbackground=UI_BORDER,
            highlightthickness=1,
            borderwidth=0,
            highlightcolor=UI_BORDER,
        )
        
        link_label = tk.Label(
            link_frame,
            text="视频链接",
            font=UI_FONT_BOLD,
            foreground=UI_TEXT,
            bg=UI_CARD,
        )
        link_label.pack(side=tk.LEFT, padx=(14, 10), pady=14)
        
        self.link_entry = tk.Entry(
            link_frame,
            textvariable=self.link_var,
            font=UI_FONT,
            bd=0,
            bg=UI_CARD,
            fg=UI_TEXT,
            relief=tk.FLAT,
            highlightthickness=0,
        )
        self.link_entry.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True, pady=15)
        
        # 飞书同步（链接分析流程：生成 MD 后上传；勾选状态写入 config.json）
        feishu_quick_frame = tk.Frame(core_frame, bg=UI_CARD, bd=0, relief=tk.FLAT)
        feishu_quick_frame.pack(fill=tk.X, padx=0, pady=(0, 8))
        feishu_quick_frame.configure(
            bg=UI_CARD,
            highlightbackground=UI_BORDER,
            highlightthickness=1,
            borderwidth=0,
            highlightcolor=UI_BORDER,
        )
        inner_f = tk.Frame(feishu_quick_frame, bg=UI_CARD)
        inner_f.pack(fill=tk.X, padx=12, pady=10)

        self.feishu_sync_var = tk.BooleanVar(value=bool(CONFIG.get("feishu_sync_enabled", False)))

        def _persist_feishu_sync_toggle():
            global CONFIG
            CONFIG = {**CONFIG, "feishu_sync_enabled": bool(self.feishu_sync_var.get())}
            save_config(CONFIG)

        feishu_chk = tk.Checkbutton(
            inner_f,
            text="同步到飞书（生成 MD 后上传至知识库路径）",
            variable=self.feishu_sync_var,
            command=_persist_feishu_sync_toggle,
            font=UI_FONT,
            bg=UI_CARD,
            fg=UI_TEXT,
            activebackground=UI_CARD,
            highlightthickness=0,
            selectcolor=UI_CARD,
        )
        feishu_chk.pack(anchor=tk.W)

        self.task_feishu_folder_var = tk.StringVar(
            value=(CONFIG.get("feishu_default_folder_path") or "").strip()
        )
        row2 = tk.Frame(inner_f, bg=UI_CARD)
        row2.pack(fill=tk.X, pady=(6, 0))
        tk.Label(
            row2,
            text="本任务飞书路径（可选，覆盖默认）：",
            font=("Microsoft YaHei UI", 9),
            fg=UI_TEXT_MUTED,
            bg=UI_CARD,
        ).pack(side=tk.LEFT)
        self.task_feishu_folder_entry = tk.Entry(
            row2,
            textvariable=self.task_feishu_folder_var,
            font=("Microsoft YaHei UI", 9),
            bd=1,
            relief=tk.FLAT,
            highlightthickness=1,
            highlightbackground=UI_BORDER,
            bg=UI_LOG_BG,
            fg=UI_TEXT,
        )
        self.task_feishu_folder_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(8, 0))
        tk.Label(
            inner_f,
            text="留空则使用「AI配置」中的默认飞书文件夹路径；统一凭证与默认路径在「AI配置」中设置。",
            font=("Microsoft YaHei UI", 8),
            fg=UI_TEXT_LIGHT,
            bg=UI_CARD,
        ).pack(anchor=tk.W, pady=(4, 0))

        # 按钮区域
        button_frame = tk.Frame(core_frame, bg=UI_BG)
        button_frame.pack(fill=tk.X, padx=0, pady=8)
        
        btn_container = tk.Frame(button_frame, bg=UI_BG)
        btn_container.pack(anchor=tk.E)
        
        # 开始处理按钮 - 使用tk.Button并设置蓝色样式
        self.start_btn = tk.Button(
            btn_container,
            text="开始处理",
            command=self.start,
            bg=UI_ACCENT,
            fg="#ffffff",
            font=UI_FONT_BOLD,
            relief=tk.FLAT,
            cursor="hand2",
            padx=22,
            pady=6,
            activebackground="#1d4ed8",
            activeforeground="#ffffff",
            bd=0,
        )
        self.start_btn.pack(side=tk.LEFT, padx=10)
        
        self.ai_config_btn = ttk.Button(btn_container, text="AI配置", command=self.open_ai_config_window)
        self.ai_config_btn.pack(side=tk.LEFT, padx=10)
        
        # AI API配置按钮（新的API配置界面）
        self.ai_api_config_btn = ttk.Button(btn_container, text="API设置", command=self.open_ai_api_config_window)
        self.ai_api_config_btn.pack(side=tk.LEFT, padx=10)
        
        # 批量导入按钮
        self.batch_import_btn = ttk.Button(btn_container, text="批量导入", command=self.batch_import)
        self.batch_import_btn.pack(side=tk.LEFT, padx=10)
        
        # 历史查询按钮
        self.history_btn = ttk.Button(btn_container, text="历史查询", command=self.show_history)
        self.history_btn.pack(side=tk.LEFT, padx=10)
        
        # 线程配置按钮
        self.thread_config_btn = ttk.Button(btn_container, text="线程配置", command=self.open_thread_config_window)
        self.thread_config_btn.pack(side=tk.LEFT, padx=10)
        
        # 任务状态区域
        self.status_frame = tk.Frame(parent, bg=UI_CARD, bd=0, relief=tk.FLAT)
        self.status_frame.pack(fill=tk.X, pady=(0, 12))
        self.status_frame.configure(
            bg=UI_CARD,
            highlightbackground=UI_BORDER,
            highlightthickness=1,
            borderwidth=0,
            highlightcolor=UI_BORDER,
        )
        
        self.status_info = tk.Label(
            self.status_frame,
            text="任务状态：就绪",
            font=UI_FONT,
            foreground=UI_TEXT,
            bg=UI_CARD,
        )
        self.status_info.pack(side=tk.LEFT, padx=(14, 10), pady=10)
        
        # 队列状态显示
        self.queue_status = tk.Label(
            self.status_frame,
            text="队列：0 个任务",
            font=UI_FONT,
            foreground=UI_ACCENT,
            bg=UI_CARD,
        )
        self.queue_status.pack(side=tk.RIGHT, padx=(10, 15), pady=10)
        
        # User Prompt 输入区域
        user_prompt_frame = tk.Frame(parent, bg=UI_CARD, bd=0, relief=tk.FLAT)
        user_prompt_frame.pack(fill=tk.X, pady=(0, 12))
        user_prompt_frame.configure(
            bg=UI_CARD,
            highlightbackground=UI_BORDER,
            highlightthickness=1,
            borderwidth=0,
            highlightcolor=UI_BORDER,
        )
        
        prompt_label_frame = tk.Frame(user_prompt_frame, bg=UI_CARD)
        prompt_label_frame.pack(fill=tk.X, padx=15, pady=(14, 0))
        
        prompt_label = tk.Label(
            prompt_label_frame,
            text="User Prompt（可选）",
            font=UI_FONT_BOLD,
            foreground=UI_TEXT,
            bg=UI_CARD,
        )
        prompt_label.pack(side=tk.LEFT)
        
        # 添加展开按钮
        def open_prompt_window():
            """打开大文本框窗口"""
            prompt_window = tk.Toplevel(self.root)
            prompt_window.title("User Prompt 编辑")
            prompt_window.geometry("800x500")
            prompt_window.configure(bg="#f0f4f8")
            prompt_window.transient(self.root)
            prompt_window.grab_set()
            
            # 居中显示
            prompt_window.update_idletasks()
            width = prompt_window.winfo_width()
            height = prompt_window.winfo_height()
            x = (prompt_window.winfo_screenwidth() // 2) - (width // 2)
            y = (prompt_window.winfo_screenheight() // 2) - (height // 2)
            prompt_window.geometry(f"{width}x{height}+{x}+{y}")
            
            # 文本框标签
            text_frame = tk.Frame(prompt_window, bg="#ffffff")
            text_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=(20, 10))
            text_frame.configure(highlightbackground="#0066cc", highlightthickness=1)
            
            # 大文本框
            text_widget = tk.Text(
                text_frame, 
                font=("微软雅黑", 10),
                bg="#f9f9f9",
                wrap=tk.WORD
            )
            text_widget.pack(fill=tk.BOTH, expand=True, padx=15, pady=15)
            
            # 插入当前值
            current_value = self.user_prompt_var.get()
            text_widget.insert(tk.END, current_value)
            
            # 字符计数标签
            window_char_count_var = tk.StringVar(value=f"{len(current_value)}/500")
            count_frame = tk.Frame(prompt_window, bg="#f0f4f8")
            count_frame.pack(fill=tk.X, padx=20, pady=(0, 10))
            
            count_label = tk.Label(
                count_frame, 
                textvariable=window_char_count_var, 
                font=("微软雅黑", 9),
                foreground="#999",
                bg="#f0f4f8"
            )
            count_label.pack(anchor=tk.E)
            
            # 实时更新字符计数
            def update_char_count(event):
                text = text_widget.get("1.0", tk.END).strip()
                if len(text) > 500:
                    # 限制字符数
                    text_widget.delete("1.0", tk.END)
                    text_widget.insert(tk.END, text[:500])
                    text = text[:500]
                window_char_count_var.set(f"{len(text)}/500")
            
            text_widget.bind("<KeyRelease>", update_char_count)
            
            # 按钮区域
            button_frame = tk.Frame(prompt_window, bg="#f0f4f8")
            button_frame.pack(fill=tk.X, padx=20, pady=10)
            
            def save_prompt():
                """保存User Prompt"""
                value = text_widget.get("1.0", tk.END).strip()
                if len(value) > 500:
                    value = value[:500]
                self.user_prompt_var.set(value)
                prompt_window.destroy()
            
            save_btn = ttk.Button(button_frame, text="保存", command=save_prompt)
            save_btn.pack(side=tk.RIGHT, padx=10)
            
            cancel_btn = ttk.Button(button_frame, text="取消", command=prompt_window.destroy)
            cancel_btn.pack(side=tk.RIGHT, padx=10)
        
        # 展开按钮
        expand_btn = ttk.Button(
            prompt_label_frame, 
            text="展开编辑", 
            command=open_prompt_window
        )
        expand_btn.pack(side=tk.RIGHT, padx=10)
        
        prompt_desc = tk.Label(
            prompt_label_frame,
            text="每次处理视频时的额外提示信息，最多500字符",
            font=("Microsoft YaHei UI", 9),
            foreground=UI_TEXT_MUTED,
            bg=UI_CARD,
        )
        prompt_desc.pack(side=tk.LEFT, padx=10)
        
        # 字符计数标签
        char_count_var = tk.StringVar(value="0/500")
        char_count_label = tk.Label(
            prompt_label_frame,
            textvariable=char_count_var,
            font=("Microsoft YaHei UI", 9),
            foreground=UI_TEXT_LIGHT,
            bg=UI_CARD,
        )
        char_count_label.pack(side=tk.RIGHT)
        
        self.user_prompt_var = tk.StringVar(value=CONFIG.get("user_prompt", DEFAULT_CONFIG["user_prompt"]))
        self.user_prompt_entry = tk.Entry(
            user_prompt_frame,
            textvariable=self.user_prompt_var,
            font=UI_FONT,
            bd=0,
            bg=UI_LOG_BG,
            fg=UI_TEXT,
            relief=tk.FLAT,
            highlightthickness=0,
        )
        self.user_prompt_entry.pack(fill=tk.X, padx=15, pady=(10, 15))
        
        # 点击输入框也打开大文本框
        def on_entry_click(event):
            open_prompt_window()
        
        self.user_prompt_entry.bind("<Button-1>", on_entry_click)
        
        # 字符数限制和计数
        def limit_characters(*args):
            value = self.user_prompt_var.get()
            if len(value) > 500:
                self.user_prompt_var.set(value[:500])
            char_count_var.set(f"{len(self.user_prompt_var.get())}/500")
        
        self.user_prompt_var.trace_add("write", limit_characters)
        # 初始计数
        limit_characters()
        
        # 日志区域
        log_frame = tk.Frame(parent, bg=UI_CARD, bd=0, relief=tk.FLAT)
        log_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 12))
        log_frame.configure(
            bg=UI_CARD,
            highlightbackground=UI_BORDER,
            highlightthickness=1,
            borderwidth=0,
            highlightcolor=UI_BORDER,
        )
        
        log_title_frame = tk.Frame(log_frame, bg=UI_CARD)
        log_title_frame.pack(fill=tk.X, padx=15, pady=(14, 8))
        
        log_title = tk.Label(
            log_title_frame,
            text="处理日志",
            font=UI_FONT_BOLD,
            foreground=UI_TEXT,
            bg=UI_CARD,
        )
        log_title.pack(anchor=tk.W)
        
        self.log = scrolledtext.ScrolledText(
            log_frame,
            height=15,
            font=("Consolas", 10),
            bd=0,
            bg=UI_LOG_BG,
            fg=UI_TEXT,
            relief=tk.FLAT,
            wrap=tk.WORD,
            highlightthickness=0,
        )
        self.log.pack(fill=tk.BOTH, expand=True, padx=15, pady=(0, 15))
        
        # 底部状态区域
        status_frame = tk.Frame(parent, bg=UI_BG)
        status_frame.pack(fill=tk.X, pady=(0, 8))
        
        self.status_var = tk.StringVar(value="就绪")
        status_label = tk.Label(
            status_frame,
            textvariable=self.status_var,
            font=UI_FONT,
            foreground=UI_TEXT_MUTED,
            bg=UI_BG,
        )
        status_label.pack(anchor=tk.W)
        
        # 状态指示器
        status_indicator = tk.Label(
            status_frame, 
            text="●", 
            font=("微软雅黑", 12),
            foreground="#00cc66",
            bg="#f0f4f8"
        )
        status_indicator.pack(side=tk.LEFT, padx=(0, 10))

    # 日志与状态（非主线程通过 after 投递，避免 Tk 死锁导致队列卡死在「处理中」）
    def append_log(self, msg, *args):
        ts = datetime.now().strftime("%H:%M:%S")
        thread_id = threading.current_thread().name
        level = "INFO"
        if args and isinstance(args[-1], str):
            u = args[-1].upper()
            if u in ("ERROR", "EXCEPTION", "WARNING", "INFO", "DEBUG"):
                level = u
        line = f"[{ts}] [线程:{thread_id}] {msg}\n"

        def _do():
            self.log.insert(tk.END, line)
            self.log.see(tk.END)
            self.root.update_idletasks()

        if threading.current_thread() is threading.main_thread():
            _do()
        else:
            try:
                self.root.after(0, _do)
            except Exception:
                _do()

        if self._ops_should_forward_log_to_agent(msg, level):
            threading.Thread(
                target=self._ops_dispatch_log_incident,
                args=(msg, level),
                daemon=True,
            ).start()

    def _ops_should_forward_log_to_agent(self, msg: str, level: str) -> bool:
        """仅 ERROR / EXCEPTION 上报运维 Agent；WARNING 不自动上报。"""
        if not getattr(self, "ops_agent", None):
            return False
        if "[运维Agent]" in msg or "运维事件已记录" in msg:
            return False
        if os.environ.get("OPS_LOG_INCIDENT_DISABLE", "").strip() in ("1", "true", "yes"):
            return False
        return level in ("ERROR", "EXCEPTION")

    def _ops_incident_fingerprint(self, category: str, text: str) -> str:
        import hashlib

        h = hashlib.md5(f"{category}|{text[:500]}".encode("utf-8", errors="ignore")).hexdigest()
        return h

    def _ops_incident_allow(self, fp: str) -> bool:
        import time as _t

        now = _t.monotonic()
        with self._ops_incident_lock:
            last = self._ops_incident_last_ts.get(fp, 0.0)
            if now - last < self._ops_incident_cooldown:
                return False
            self._ops_incident_last_ts[fp] = now
        return True

    def _ops_collect_ui_logs(self, max_lines: int = 200) -> list:
        out: list = []
        ev = threading.Event()

        def _grab():
            try:
                w = getattr(self, "log", None)
                if w is not None:
                    raw = w.get("1.0", tk.END)
                    lines = raw.split("\n")
                    out.append(lines[-max_lines:] if len(lines) > max_lines else lines)
                else:
                    out.append([])
            finally:
                ev.set()

        try:
            if threading.current_thread() is threading.main_thread():
                _grab()
            else:
                self.root.after(0, _grab)
                ev.wait(timeout=3.0)
        except Exception:
            out.append([])
        return out[0] if out else []

    def _ops_dispatch_log_incident(self, msg: str, level: str) -> None:
        fp = self._ops_incident_fingerprint("log_" + level, msg)
        if not self._ops_incident_allow(fp):
            return
        if not self.ops_agent:
            return
        try:
            logs = self._ops_collect_ui_logs()
            err = {"type": f"Log{level}", "message": msg[:4000], "traceback": ""}
            tid = f"log_{fp[:10]}"
            md = self.ops_agent.monitor_task_completion(
                link="_gui_log_",
                task_id=tid,
                status="failed",
                logs=logs,
                error_info=err,
            )
            if md:
                self.append_log(f"[运维Agent] 已根据日志{level}生成维护建议: {md}", "INFO")
        except Exception as e:
            self.append_log(f"[运维Agent] 日志事件上报失败: {e}", "INFO")

    def _schedule_ops_volcengine_degraded(self, primary_err: str, primary_ep: str, backup_ep: str) -> None:
        msg = (
            f"火山主接入点失败但备用成功。主: {primary_ep} 错误: {primary_err[:800]}；"
            f"已用备: {backup_ep}。建议：检查控制台 Safe Experience/限额，或把 config.json 中 "
            f"ai_chat_model 与 ai_chat_model_backup 对调。"
        )
        fp = self._ops_incident_fingerprint("volcengine_degraded", primary_err + primary_ep)
        if not self._ops_incident_allow(fp):
            return
        if not self.ops_agent:
            return
        try:
            logs = self._ops_collect_ui_logs()
            err = {
                "type": "VolcenginePrimaryFailedBackupOk",
                "message": msg,
                "traceback": "",
            }
            md = self.ops_agent.monitor_task_completion(
                link="_volcengine_",
                task_id=f"ve_{fp[:10]}",
                status="failed",
                logs=logs,
                error_info=err,
            )
            if md:
                self.append_log(f"[运维Agent] 主备切换事件已记录: {md}", "INFO")
        except Exception as e:
            self.append_log(f"[运维Agent] 主备事件上报失败: {e}", "INFO")

    def set_status(self, msg: str):
        self.status_var.set(msg)
        self.root.update_idletasks()
    
    def update_progress(self, percentage: float, status: str = ""):
        """更新任务状态显示"""
        if status:
            self.set_status(status)
            self.status_info.config(text=f"任务状态：{status}")
        self.root.update_idletasks()
    
    def update_queue_status(self):
        """更新队列状态显示"""
        def _do():
            pending = self._task_queue_len()
            running = len(self.active_futures)
            processing_status = "处理中" if self.processing_queue else "就绪"
            self.queue_status.config(
                text=f"队列：{pending} 待处理 + {running} 执行中 | 状态：{processing_status}"
            )
            self.root.update_idletasks()

        if threading.current_thread() is threading.main_thread():
            _do()
        else:
            try:
                self.root.after(0, _do)
            except Exception:
                _do()

    def _task_queue_len(self) -> int:
        with self._task_queue_lock:
            return len(self.task_queue)

    def _task_queue_append_unique(self, link: str) -> bool:
        with self._task_queue_lock:
            if link in self.task_queue:
                return False
            self.task_queue.append(link)
            return True

    def _task_queue_pop_front(self):
        with self._task_queue_lock:
            if not self.task_queue:
                return None
            return self.task_queue.pop(0)

    def _task_queue_remove_if_present(self, link: str) -> bool:
        with self._task_queue_lock:
            if link in self.task_queue:
                self.task_queue.remove(link)
                return True
            return False

    def _total_queued_work(self) -> int:
        return self._task_queue_len() + len(self.active_futures)

    def _pipeline_log(self, message: str):
        """写入 src/agent/pipeline.log，便于排查卡住、超时（多线程安全）"""
        try:
            path = os.path.join(BASE_DIR, "pipeline.log")
            ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            th = threading.current_thread().name
            line = f"[{ts}] [{th}] {message}\n"
            with self._pipeline_log_lock:
                with open(path, "a", encoding="utf-8") as f:
                    f.write(line)
        except Exception:
            pass

    # 恢复未完成任务
    def recover_unfinished_tasks(self):
        """恢复上次未完成的任务"""
        unfinished_tasks = [task for task in self.history.get("tasks", []) 
                          if task.get("status") not in ["completed", "failed"]]
        
        if unfinished_tasks:
            self.append_log(f"发现 {len(unfinished_tasks)} 个未完成任务，正在恢复...")
            for task in unfinished_tasks:
                link = task.get("link")
                self._task_queue_append_unique(link)
            
            self.append_log(f"已恢复 {len(unfinished_tasks)} 个任务到队列")
            self.append_log(f"当前待处理：{self._task_queue_len()}")
            
            # 自动开始处理队列
            if not self.processing_queue and self._task_queue_len() > 0:
                self.start_queue_processing()
        else:
            self.append_log("无未完成任务需要恢复")
    
    # 检查链接是否已导入
    def is_link_already_imported(self, link):
        """检查链接是否已经导入过"""
        # 使用链接的hash作为任务ID，提高比较效率
        link_hash = hashlib.md5(link.encode()).hexdigest()
        
        # 检查历史记录中是否存在该链接
        for task in self.history.get("tasks", []):
            if task.get("id") == link_hash:
                return True
        
        return False
    
    # 从链接中提取标题
    def extract_title_from_link(self, link):
        """从链接中提取标题
        
        Args:
            link: 视频链接
            
        Returns:
            提取的标题，如果无法提取则返回None
        """
        try:
            # 对于小红书链接，尝试获取页面标题
            if "xiaohongshu.com" in link:
                import requests
                
                # 尝试导入BeautifulSoup
                try:
                    from bs4 import BeautifulSoup
                    
                    headers = {
                        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/144.0.0.0 Safari/537.36',
                        'Referer': 'https://www.xiaohongshu.com/'
                    }
                    
                    response = requests.get(link, headers=headers, timeout=2)
                    if response.status_code == 200:
                        soup = BeautifulSoup(response.text, 'html.parser')
                        # 尝试从title标签提取
                        title_tag = soup.find('title')
                        if title_tag and title_tag.text:
                            title = title_tag.text.strip()
                            # 清理标题
                            title = title.replace('\n', '').replace('\r', '').replace('  ', ' ')
                            self.append_log(f"从链接中提取标题：{title}")
                            return title
                except ImportError:
                    # 如果缺少bs4模块，跳过网页解析，直接从链接中提取
                    self.append_log("缺少bs4模块，跳过网页解析，直接从链接中提取标题")
                except Exception as e:
                    # 超时或其他异常，跳过网页解析
                    self.append_log(f"小红书链接解析异常：{e}，跳过网页解析")
            
            # 对于B站链接，尝试获取页面标题
            elif "bilibili.com" in link:
                import requests
                
                # 尝试导入BeautifulSoup
                try:
                    from bs4 import BeautifulSoup
                    
                    headers = {
                        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/144.0.0.0 Safari/537.36',
                        'Referer': 'https://www.bilibili.com/'
                    }
                    
                    response = requests.get(link, headers=headers, timeout=2)
                    if response.status_code == 200:
                        soup = BeautifulSoup(response.text, 'html.parser')
                        # 尝试从title标签提取
                        title_tag = soup.find('title')
                        if title_tag and title_tag.text:
                            title = title_tag.text.strip()
                            # 清理标题
                            title = title.replace('\n', '').replace('\r', '').replace('  ', ' ')
                            # 移除B站标题后缀 "_哔哩哔哩_bilibili"
                            if title.endswith('_哔哩哔哩_bilibili'):
                                title = title[:-len('_哔哩哔哩_bilibili')]
                            self.append_log(f"从B站链接中提取标题：{title}")
                            return title
                except ImportError:
                    # 如果缺少bs4模块，跳过网页解析，直接从链接中提取
                    self.append_log("缺少bs4模块，跳过网页解析，直接从链接中提取标题")
                except Exception as e:
                    # 超时或其他异常，跳过网页解析
                    self.append_log(f"B站链接解析异常：{e}，跳过网页解析")
            
            # 如果无法从页面获取，尝试从链接中提取
            import re
            # 对于B站链接，尝试提取BV号
            bv_match = re.search(r'BV[0-9A-Za-z]{10}', link)
            if bv_match:
                return bv_match.group(0)
            
            # 尝试匹配链接中的数字或有意义的部分
            match = re.search(r'[a-zA-Z0-9_-]{8,}', link)
            if match:
                return match.group(0)
            
            # 返回默认标题
            return "未知标题"
        except Exception as e:
            self.append_log(f"从链接提取标题异常：{e}")
            # 异常时返回默认标题
            return "未知标题"
    
    def _feishu_sync_enabled(self):
        """是否将生成结果同步到飞书（config.json 持久化 feishu_sync_enabled）。"""
        return bool(CONFIG.get("feishu_sync_enabled", False))

    def _feishu_upload_effective_for_link(self, link):
        """任务级可覆盖：feishu_sync_override True/False 优先，否则用全局勾选。"""
        for task in self.history.get("tasks", []):
            if task.get("link") == link:
                o = task.get("feishu_sync_override")
                if o is not None:
                    return bool(o)
                break
        return self._feishu_sync_enabled()

    def _make_feishu_kb(self):
        """飞书应用凭证：config.json 的 feishu_app_id / feishu_app_secret，或环境变量 FEISHU_APP_ID / FEISHU_APP_SECRET。"""
        try:
            from feishu_integration import FeishuKnowledgeBase
        except ImportError:
            return None
        aid = (
            (CONFIG.get("feishu_app_id") or "").strip()
            or (os.environ.get("FEISHU_APP_ID") or "").strip()
        )
        sec = (
            (CONFIG.get("feishu_app_secret") or "").strip()
            or (os.environ.get("FEISHU_APP_SECRET") or "").strip()
        )
        if not aid or not sec:
            return None
        return FeishuKnowledgeBase(aid, sec)

    def _resolve_feishu_folder_for_task(self):
        """单任务飞书路径：主界面「本任务」输入优先，否则默认路径。"""
        if hasattr(self, "task_feishu_folder_var"):
            t = self.task_feishu_folder_var.get().strip()
            if t:
                return t
        d = (CONFIG.get("feishu_default_folder_path") or "").strip()
        return d or None

    def _run_feishu_upload_if_enabled(self, link, md_file, user_prompt, feishu_folder_path=None):
        """生成 MD 后的飞书上传：受全局/任务级同步开关与 AI 配置中的凭证控制。"""
        if not self._feishu_upload_effective_for_link(link):
            self.append_log("未开启「同步到飞书」，跳过上传步骤", "INFO")
            self.update_task_status(link, "feishu_upload", "completed", "未开启飞书同步")
            return
        self.update_task_status(link, "feishu_upload", "in_progress")
        self.update_progress(90, "上传到飞书...")
        try:
            with open(md_file, "r", encoding="utf-8") as f:
                md_content = f.read()
            feishu = self._make_feishu_kb()
            if feishu is None:
                self.append_log(
                    "未配置飞书凭证：请在「AI配置」中填写飞书 App ID / App Secret 并保存，"
                    "或设置环境变量 FEISHU_APP_ID、FEISHU_APP_SECRET；"
                    f"当前 config.json 路径：{CONFIG_FILE}",
                    "WARNING",
                )
                self.update_task_status(link, "feishu_upload", "completed", "未配置飞书凭证")
                return
            prompt_folder = feishu.parse_feishu_folder_from_prompt(user_prompt)
            default_cfg = (CONFIG.get("feishu_default_folder_path") or "").strip() or None
            final_folder = feishu_folder_path or prompt_folder or default_cfg
            folder_token_cfg = (CONFIG.get("feishu_folder_token") or "").strip() or None
            doc_title = os.path.basename(md_file).replace(".md", "")
            doc_token = feishu.upload_document(
                doc_title,
                md_content,
                feishu_folder_path=final_folder,
                folder_token=folder_token_cfg,
            )
            if doc_token:
                self.append_log(f"文档已上传到飞书：{doc_token}")
                self.update_task_status(link, "feishu_upload", "completed", doc_token)
            else:
                self.append_log("上传到飞书失败")
                self.update_task_status(link, "feishu_upload", "failed")
        except Exception as e:
            self.append_log(f"飞书上传异常：{e}")
            self.update_task_status(link, "feishu_upload", "failed")

    # 添加任务到历史记录
    def add_task_to_history(self, link, user_prompt="", feishu_folder_path=None, feishu_sync_override=None):
        """添加任务到历史记录
        
        Args:
            link: 视频链接
            user_prompt: 用户提示词
            feishu_folder_path: 飞书文件夹路径（单任务/单批覆盖）
            feishu_sync_override: 是否同步飞书 None=跟随全局勾选 True/False=本任务覆盖
            
        Returns:
            任务对象
        """
        # 检查是否已存在
        link_hash = hashlib.md5(link.encode()).hexdigest()
        for task in self.history.get("tasks", []):
            if task.get("id") == link_hash:
                # 更新用户提示词和飞书文件夹路径
                if user_prompt:
                    task["user_prompt"] = user_prompt
                if feishu_folder_path:
                    task["feishu_folder_path"] = feishu_folder_path
                if feishu_sync_override is not None:
                    task["feishu_sync_override"] = feishu_sync_override
                task["updated_at"] = datetime.now().isoformat()
                save_history(self.history)
                return task
        
        # 从链接中提取标题
        link_title = self.extract_title_from_link(link)
        
        # 创建新任务记录
        new_task = {
            "id": link_hash,
            "link": link,
            "title": link_title,
            "status": "pending",
            "user_prompt": user_prompt,
            "feishu_folder_path": feishu_folder_path,
            "feishu_sync_override": feishu_sync_override,
            "stages": {
                "download": {"status": "pending", "result": None},
                "transcribe": {"status": "pending", "result": None},
                "ai_analysis": {"status": "pending", "result": None},
                "generate_md": {"status": "pending", "result": None},
                "feishu_upload": {"status": "pending", "result": None}
            },
            "created_at": datetime.now().isoformat(),
            "updated_at": datetime.now().isoformat()
        }
        
        self.history.setdefault("tasks", []).append(new_task)
        save_history(self.history)
        return new_task
    
    # 更新任务状态
    def update_task_status(self, link, stage, status, result=None):
        """更新任务状态"""
        for task in self.history.get("tasks", []):
            if task.get("link") == link:
                # 更新阶段状态
                if stage in task.get("stages", {}):
                    task["stages"][stage]["status"] = status
                    if result:
                        task["stages"][stage]["result"] = result
                
                # 更新整体状态
                if (stage == "generate_md" or stage == "feishu_upload") and status == "completed":
                    task["status"] = "completed"
                elif status == "failed":
                    task["status"] = "failed"
                
                task["updated_at"] = datetime.now().isoformat()
                save_history(self.history)
                return True
        return False
    
    # 页面切换方法
    def _show_video_page(self):
        """显示视频处理页面"""
        if hasattr(self, 'current_page') and self.current_page == "video":
            return
        
        # 隐藏AI问答页面
        if hasattr(self, 'chat_page') and self.chat_page:
            self.chat_page.pack_forget()
        
        # 隐藏多模态文档处理页面
        if hasattr(self, 'multimodal_page') and self.multimodal_page:
            self.multimodal_page.pack_forget()
        
        # 显示视频处理页面
        if hasattr(self, 'video_page') and self.video_page:
            self.video_page.pack(fill=tk.BOTH, expand=True)
        
        # 更新按钮样式
        if hasattr(self, 'nav_video_btn') and hasattr(self, 'nav_chat_btn'):
            self.nav_video_btn.configure(bg=UI_CARD, fg=UI_ACCENT)
            self.nav_chat_btn.configure(bg=UI_BG, fg=UI_TEXT_MUTED)
        if hasattr(self, 'nav_multimodal_btn'):
            self.nav_multimodal_btn.configure(bg=UI_BG, fg=UI_TEXT_MUTED)
        
        self.current_page = "video"
        self.root.title(APP_TITLE)
    
    def _show_chat_page(self):
        """显示AI问答页面"""
        if hasattr(self, 'current_page') and self.current_page == "chat":
            return
        
        # 隐藏视频处理页面
        if hasattr(self, 'video_page') and self.video_page:
            self.video_page.pack_forget()
        
        # 隐藏多模态文档处理页面
        if hasattr(self, 'multimodal_page') and self.multimodal_page:
            self.multimodal_page.pack_forget()
        
        # 创建AI问答页面（如果还没有创建）
        if not hasattr(self, 'chat_page') or self.chat_page is None:
            if AI_CHAT_PAGE_AVAILABLE:
                # 使用新的AIChatPage（带任务管理）
                self.chat_page = AIChatPage(self.content_frame)
            elif CHAT_GUI_AVAILABLE:
                # 回退到旧的ChatGUI
                self.chat_page = tk.Frame(self.content_frame, bg=UI_BG)
                chat_gui = ChatGUI(self.chat_page)
            else:
                messagebox.showerror("错误", "AI问答系统模块未安装")
                return
        
        # 显示AI问答页面
        self.chat_page.pack(fill=tk.BOTH, expand=True)
        
        # 更新按钮样式
        if hasattr(self, 'nav_video_btn') and hasattr(self, 'nav_chat_btn'):
            self.nav_video_btn.configure(bg=UI_BG, fg=UI_TEXT_MUTED)
            self.nav_chat_btn.configure(bg=UI_CARD, fg=UI_ACCENT)
        if hasattr(self, 'nav_multimodal_btn'):
            self.nav_multimodal_btn.configure(bg=UI_BG, fg=UI_TEXT_MUTED)
        
        self.current_page = "chat"
        self.root.title(f"{APP_TITLE} - AI问答")
    
    def _show_multimodal_page(self):
        """显示多模态文档处理页面"""
        if hasattr(self, 'current_page') and self.current_page == "multimodal":
            return
        
        if not MULTIMODAL_AVAILABLE:
            messagebox.showerror("错误", "多模态文档处理模块未安装")
            return
        
        # 隐藏视频处理页面
        if hasattr(self, 'video_page') and self.video_page:
            self.video_page.pack_forget()
        
        # 隐藏AI问答页面
        if hasattr(self, 'chat_page') and self.chat_page:
            self.chat_page.pack_forget()
        
        # 创建多模态文档处理页面（如果还没有创建）
        if not hasattr(self, 'multimodal_page') or self.multimodal_page is None:
            self.multimodal_page = MultimodalProcessingPage(self.content_frame)
        
        # 显示多模态文档处理页面
        self.multimodal_page.pack(fill=tk.BOTH, expand=True)
        
        # 更新按钮样式
        if hasattr(self, 'nav_video_btn') and hasattr(self, 'nav_chat_btn'):
            self.nav_video_btn.configure(bg=UI_BG, fg=UI_TEXT_MUTED)
            self.nav_chat_btn.configure(bg=UI_BG, fg=UI_TEXT_MUTED)
        if hasattr(self, 'nav_multimodal_btn'):
            self.nav_multimodal_btn.configure(bg=UI_CARD, fg=UI_ACCENT)
        
        self.current_page = "multimodal"
        self.root.title(f"{APP_TITLE} - 文档处理")
    
    # 打开AI问答窗口
    def open_ai_chat_window(self):
        """打开AI问答系统窗口"""
        if not CHAT_GUI_AVAILABLE:
            messagebox.showerror("错误", "AI问答系统模块未安装，请检查chat_gui.py是否存在")
            return
        
        try:
            # 创建新窗口
            chat_window = tk.Toplevel(self.root)
            chat_window.title("🤖 AI技术专家问答系统")
            chat_window.geometry("1200x800")
            chat_window.minsize(1000, 600)
            
            # 创建ChatGUI实例
            chat_gui = ChatGUI(chat_window)
            
            # 窗口关闭时清理
            def on_close():
                chat_window.destroy()
            
            chat_window.protocol("WM_DELETE_WINDOW", on_close)
            
            self.append_log("打开AI问答系统窗口")
            
        except Exception as e:
            messagebox.showerror("错误", f"打开AI问答系统失败：{e}")
            import traceback
            traceback.print_exc()
    
    def _build_chat_page(self, parent):
        """构建AI问答页面 - 使用豆包AI风格界面"""
        # 使用豆包AI风格聊天界面，传入RAG知识库和RAG工具
        chat_page = DoubaoChatPage(parent, rag_kb=self._rag_kb, rag_tool=self._rag_tool)
        chat_page.pack(fill=tk.BOTH, expand=True)
        self.ai_chat_page = chat_page
    
    # 显示历史记录
    def show_history(self):
        """显示历史记录窗口 - 增强版：支持实时状态、停止按钮、队列排序"""
        history_window = tk.Toplevel(self.root)
        history_window.title("历史记录查询")
        history_window.geometry("1200x700")
        history_window.configure(bg="#f0f4f8")
        
        # 设置样式
        style = ttk.Style()
        style.theme_use("clam")
        
        # 主容器
        main_frame = tk.Frame(history_window, bg="#f0f4f8")
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # 标题
        title_frame = tk.Frame(main_frame, bg="#f0f4f8")
        title_frame.pack(fill=tk.X, pady=(0, 15))
        
        title_label = tk.Label(
            title_frame, 
            text="历史记录查询", 
            font=("微软雅黑", 16, "bold"),
            foreground="#0066cc",
            bg="#f0f4f8"
        )
        title_label.pack(side=tk.LEFT)
        
        # 刷新按钮
        refresh_btn = tk.Button(
            title_frame,
            text="🔄 刷新",
            font=("微软雅黑", 10),
            bg="#2196F3",
            fg="white",
            cursor="hand2",
            relief=tk.FLAT,
            padx=15,
            pady=5
        )
        refresh_btn.pack(side=tk.RIGHT, padx=5)
        
        # 历史记录列表
        tree_frame = tk.Frame(main_frame, bg="#ffffff")
        tree_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 15))
        tree_frame.configure(highlightbackground="#0066cc", highlightthickness=1)
        
        # 使用grid布局管理器来确保树状图完全填充空间
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)
        
        # 创建树状图 - 新增"当前阶段"和"队列位置"列
        columns = ("id", "title", "link", "status", "current_stage", "queue_pos", "created_at", "updated_at")
        tree = ttk.Treeview(tree_frame, columns=columns, show="headings")
        
        # 设置列标题
        tree.heading("id", text="任务ID")
        tree.heading("title", text="标题")
        tree.heading("link", text="视频链接")
        tree.heading("status", text="状态")
        tree.heading("current_stage", text="当前阶段")
        tree.heading("queue_pos", text="队列位置")
        tree.heading("created_at", text="创建时间")
        tree.heading("updated_at", text="更新时间")
        
        # 设置列宽
        tree.column("id", width=80)
        tree.column("title", width=180)
        tree.column("link", width=250)
        tree.column("status", width=80)
        tree.column("current_stage", width=120)
        tree.column("queue_pos", width=80)
        tree.column("created_at", width=130)
        tree.column("updated_at", width=130)
        
        # 添加垂直滚动条
        v_scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
        v_scrollbar.grid(row=0, column=1, sticky="ns")
        tree.configure(yscroll=v_scrollbar.set)
        
        # 添加水平滚动条
        h_scrollbar = ttk.Scrollbar(tree_frame, orient="horizontal", command=tree.xview)
        h_scrollbar.grid(row=1, column=0, sticky="ew")
        tree.configure(xscroll=h_scrollbar.set)
        
        # 放置树状图
        tree.grid(row=0, column=0, sticky="nsew")
        
        # 添加鼠标滚轮支持
        def on_mouse_wheel(event):
            try:
                if event.state & 0x1:
                    tree.xview_scroll(-1 * (event.delta // 120), "units")
                else:
                    tree.yview_scroll(-1 * (event.delta // 120), "units")
            except Exception:
                pass
        
        tree.bind_all("<MouseWheel>", on_mouse_wheel)
        
        # 绑定窗口关闭事件
        def on_window_close():
            try:
                tree.unbind_all("<MouseWheel>")
            except Exception:
                pass
            history_window.destroy()
        
        history_window.protocol("WM_DELETE_WINDOW", on_window_close)
        
        # 阶段名称映射
        stage_names = {
            "download": "📥 下载中",
            "transcribe": "🎤 转写中",
            "ai_analysis": "🤖 AI分析中",
            "generate_md": "📝 生成文档中",
            "download_failed": "❌ 下载失败",
            "transcribe_failed": "❌ 转写失败",
            "ai_analysis_failed": "❌ AI分析失败",
            "generate_md_failed": "❌ 生成失败",
            "completed": "✅ 已完成",
            "cancelled": "🛑 已取消",
            "pending": "⏳ 等待中",
            "unknown": "❓ 未知"
        }
        
        # 填充数据的函数
        def refresh_tree():
            """刷新树状图数据"""
            # 清空现有数据
            for item in tree.get_children():
                tree.delete(item)
            
            # 获取任务数据
            tasks = self.history.get("tasks", [])
            tasks.sort(key=lambda x: x.get("created_at", ""), reverse=True)
            
            for task in tasks:
                task_id = task.get("id", "")
                title = task.get("title", "")
                if not title:
                    title = "生成中..."
                link = task.get("link", "")
                status = task.get("status", "")
                created_at = task.get("created_at", "")
                updated_at = task.get("updated_at", "")
                
                # 获取当前阶段
                current_stage = self.get_task_current_stage(link)
                stage_display = stage_names.get(current_stage, current_stage)
                
                # 获取队列位置
                queue_pos = ""
                if link in self.active_futures:
                    queue_pos = "执行中"
                elif link in self.task_queue:
                    with self._task_queue_lock:
                        pos = self.task_queue.index(link) + 1
                    queue_pos = f"第{pos}位（待处理）"
                
                # 根据状态设置标签
                tag = status
                if status == "completed":
                    tag = "completed"
                elif status == "failed":
                    tag = "failed"
                elif status == "in_progress" or current_stage in ["download", "transcribe", "ai_analysis", "generate_md"]:
                    tag = "in_progress"
                elif status == "cancelled":
                    tag = "cancelled"
                else:
                    tag = "pending"
                
                # 插入数据
                tree.insert("", "end", values=(
                    task_id,
                    title,
                    link,
                    status,
                    stage_display,
                    queue_pos,
                    created_at,
                    updated_at
                ), tags=(tag,))
            
            # 设置标签颜色
            tree.tag_configure("completed", background="#e6ffe6")
            tree.tag_configure("failed", background="#ffe6e6")
            tree.tag_configure("in_progress", background="#fff0e6")
            tree.tag_configure("pending", background="#f0f0f0")
            tree.tag_configure("cancelled", background="#ffe0e0")
        
        # 初始填充
        refresh_tree()
        refresh_btn.config(command=refresh_tree)
        
        # 自动刷新（每3秒）
        def auto_refresh():
            if history_window.winfo_exists():
                refresh_tree()
                history_window.after(3000, auto_refresh)
        
        auto_refresh()
        
        # 清理完成任务的缓存文件
        self.cleanup_completed_tasks()
        
        # 详情查看按钮
        def view_details():
            selected = tree.selection()
            if not selected:
                messagebox.showinfo("提示", "请选择一个任务查看详情")
                return
            
            item = tree.item(selected[0])
            task_id = item["values"][0]
            
            # 查找任务
            task = None
            for t in self.history.get("tasks", []):
                if t.get("id") == task_id:
                    task = t
                    break
            
            if not task:
                return
            
            # 显示详情窗口
            detail_window = tk.Toplevel(history_window)
            detail_window.title("任务详情")
            detail_window.geometry("800x500")
            detail_window.configure(bg="#f0f4f8")
            
            # 详情容器
            detail_frame = tk.Frame(detail_window, bg="#f0f4f8")
            detail_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            
            # 链接信息
            link_frame = tk.Frame(detail_frame, bg="#ffffff")
            link_frame.pack(fill=tk.X, pady=(0, 10))
            link_frame.configure(highlightbackground="#0066cc", highlightthickness=1)
            
            ttk.Label(link_frame, text="视频链接：", font=("微软雅黑", 10), background="#ffffff").pack(anchor=tk.W, padx=10, pady=(10, 5))
            link_text = scrolledtext.ScrolledText(link_frame, height=2, font=("Consolas", 10), bg="#f9f9f9")
            link_text.pack(fill=tk.X, padx=10, pady=(0, 10))
            link_text.insert(tk.END, task.get("link", ""))
            link_text.configure(state=tk.DISABLED)
            
            # 阶段信息
            stages_frame = tk.Frame(detail_frame, bg="#ffffff")
            stages_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
            stages_frame.configure(highlightbackground="#0066cc", highlightthickness=1)
            
            ttk.Label(stages_frame, text="处理阶段：", font=("微软雅黑", 10, "bold"), background="#ffffff").pack(anchor=tk.W, padx=10, pady=(10, 5))
            
            # 创建带滚动条的阶段内容框架
            content_frame = tk.Frame(stages_frame, bg="#ffffff")
            content_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
            
            # 添加垂直滚动条
            canvas = tk.Canvas(content_frame, bg="#ffffff")
            scrollbar = ttk.Scrollbar(content_frame, orient="vertical", command=canvas.yview)
            canvas.configure(yscrollcommand=scrollbar.set)
            
            # 内部内容框架
            inner_frame = tk.Frame(canvas, bg="#ffffff")
            canvas.create_window((0, 0), window=inner_frame, anchor="nw")
            
            # 填充阶段信息
            stages = task.get("stages", {})
            for stage_name, stage_info in stages.items():
                stage_frame = tk.Frame(inner_frame, bg="#f9f9f9")
                stage_frame.pack(fill=tk.X, padx=5, pady=5)
                
                stage_label = ttk.Label(stage_frame, text=f"{stage_name}：", font=("微软雅黑", 9), background="#f9f9f9")
                stage_label.pack(side=tk.LEFT, padx=5)
                
                status_label = ttk.Label(stage_frame, text=stage_info.get("status", ""), font=("微软雅黑", 9), background="#f9f9f9")
                status_label.pack(side=tk.LEFT, padx=10)
                
                # 根据状态设置颜色
                if stage_info.get("status") == "completed":
                    status_label.configure(foreground="#00cc66")
                elif stage_info.get("status") == "failed":
                    status_label.configure(foreground="#ff3333")
                    
                    # 添加重试按钮
                    def retry_stage(s=stage_name):
                        link = task.get("link")
                        # 更新阶段状态为pending
                        self.update_task_status(link, s, "pending")
                        # 单独处理失败的阶段
                        threading.Thread(target=self._retry_stage, args=(link, s, task), daemon=True).start()
                        self.append_log(f"开始重试处理阶段 {s}：{link}")
                        detail_window.destroy()
                    
                    retry_btn = ttk.Button(stage_frame, text="重试", command=retry_stage)
                    retry_btn.pack(side=tk.RIGHT, padx=5)
                elif stage_info.get("status") == "in_progress":
                    status_label.configure(foreground="#ff9900")
            
            # 更新画布大小
            def update_scrollregion():
                canvas.update_idletasks()
                canvas.config(scrollregion=canvas.bbox("all"))
            
            # 布局滚动条和画布
            canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            
            # 初始更新滚动区域
            update_scrollregion()
            
            # 按钮区域
            button_frame = tk.Frame(detail_window, bg="#f0f4f8")
            button_frame.pack(fill=tk.X, pady=10)
            
            # 继续处理按钮
            if task.get("status") not in ["completed", "failed"]:
                def continue_task():
                    link = task.get("link")
                    self._task_queue_append_unique(link)
                    self.append_log(f"继续处理任务：{link}")
                    if not self.processing_queue:
                        self.start_queue_processing()
                    detail_window.destroy()
                
                ttk.Button(button_frame, text="继续处理", command=continue_task).pack(side=tk.LEFT, padx=10)
            
            # 关闭按钮
            ttk.Button(button_frame, text="关闭", command=detail_window.destroy).pack(side=tk.RIGHT, padx=10)
        
        # 按钮框架
        button_frame = tk.Frame(main_frame, bg="#f0f4f8")
        button_frame.pack(fill=tk.X, pady=10)
        
        # 查看详情按钮
        def view_details():
            selected = tree.selection()
            if not selected:
                messagebox.showinfo("提示", "请选择一个任务查看详情")
                return
            
            item = tree.item(selected[0])
            task_id = item["values"][0]
            
            # 查找任务
            task = None
            for t in self.history.get("tasks", []):
                if t.get("id") == task_id:
                    task = t
                    break
            
            if not task:
                return
            
            # 显示详情窗口
            detail_window = tk.Toplevel(history_window)
            detail_window.title("任务详情")
            detail_window.geometry("800x500")
            detail_window.configure(bg="#f0f4f8")
            
            # 详情容器
            detail_frame = tk.Frame(detail_window, bg="#f0f4f8")
            detail_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            
            # 链接信息
            link_frame = tk.Frame(detail_frame, bg="#ffffff")
            link_frame.pack(fill=tk.X, pady=(0, 10))
            link_frame.configure(highlightbackground="#0066cc", highlightthickness=1)
            
            ttk.Label(link_frame, text="视频链接：", font=("微软雅黑", 10), background="#ffffff").pack(anchor=tk.W, padx=10, pady=(10, 5))
            link_text = scrolledtext.ScrolledText(link_frame, height=2, font=("Consolas", 10), bg="#f9f9f9")
            link_text.pack(fill=tk.X, padx=10, pady=(0, 10))
            link_text.insert(tk.END, task.get("link", ""))
            link_text.configure(state=tk.DISABLED)
            
            # 阶段信息
            stages_frame = tk.Frame(detail_frame, bg="#ffffff")
            stages_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
            stages_frame.configure(highlightbackground="#0066cc", highlightthickness=1)
            
            ttk.Label(stages_frame, text="处理阶段：", font=("微软雅黑", 10, "bold"), background="#ffffff").pack(anchor=tk.W, padx=10, pady=(10, 5))
            
            # 创建阶段内容框架
            content_frame = tk.Frame(stages_frame, bg="#ffffff")
            content_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
            
            stages = task.get("stages", {})
            for stage_name, stage_info in stages.items():
                stage_frame = tk.Frame(content_frame, bg="#f9f9f9")
                stage_frame.pack(fill=tk.X, padx=5, pady=5)
                
                stage_label = ttk.Label(stage_frame, text=f"{stage_name}：", font=("微软雅黑", 9), background="#f9f9f9")
                stage_label.pack(side=tk.LEFT, padx=5)
                
                status_label = ttk.Label(stage_frame, text=stage_info.get("status", ""), font=("微软雅黑", 9), background="#f9f9f9")
                status_label.pack(side=tk.LEFT, padx=10)
                
                # 根据状态设置颜色
                if stage_info.get("status") == "completed":
                    status_label.configure(foreground="#00cc66")
                elif stage_info.get("status") == "failed":
                    status_label.configure(foreground="#ff3333")
                    
                    # 添加重试按钮
                    def retry_stage(s=stage_name):
                        link = task.get("link")
                        # 更新阶段状态为pending
                        self.update_task_status(link, s, "pending")
                        # 单独处理失败的阶段
                        threading.Thread(target=self._retry_stage, args=(link, s, task), daemon=True).start()
                        self.append_log(f"开始重试处理阶段 {s}：{link}")
                        detail_window.destroy()
                    
                    retry_btn = ttk.Button(stage_frame, text="重试", command=retry_stage)
                    retry_btn.pack(side=tk.RIGHT, padx=5)
                elif stage_info.get("status") == "in_progress":
                    status_label.configure(foreground="#ff9900")
            
            # 按钮区域
            button_frame = tk.Frame(detail_window, bg="#f0f4f8")
            button_frame.pack(fill=tk.X, pady=10)
            
            # 继续处理按钮
            if task.get("status") not in ["completed", "failed"]:
                def continue_task():
                    link = task.get("link")
                    self._task_queue_append_unique(link)
                    self.update_queue_status()
                    self.append_log(f"继续处理任务：{link}")
                    if not self.processing_queue:
                        self.start_queue_processing()
                    detail_window.destroy()
                
                ttk.Button(button_frame, text="继续处理", command=continue_task).pack(side=tk.LEFT, padx=10)
            
            # 关闭按钮
            ttk.Button(button_frame, text="关闭", command=detail_window.destroy).pack(side=tk.RIGHT, padx=10)
        
        # 继续处理按钮
        def continue_task():
            selected = tree.selection()
            if not selected:
                messagebox.showinfo("提示", "请选择一个任务继续处理")
                return
            
            for item_id in selected:
                item = tree.item(item_id)
                task_id = item["values"][0]
                
                # 查找任务
                task = None
                for t in self.history.get("tasks", []):
                    if t.get("id") == task_id:
                        task = t
                        break
                
                if task:
                    link = task.get("link")
                    self._task_queue_append_unique(link)
                    self.update_queue_status()
                    self.append_log(f"继续处理任务：{link}")
            
            # 自动开始处理队列
            if not self.processing_queue and self._task_queue_len() > 0:
                self.start_queue_processing()
        
        # 刷新按钮
        def refresh_tree():
            # 清空树状图
            for item in tree.get_children():
                tree.delete(item)
            
            # 重新填充数据 - 按创建时间降序排序
            tasks = self.history.get("tasks", [])
            # 按created_at字段降序排序
            tasks.sort(key=lambda x: x.get("created_at", ""), reverse=True)
            
            for task in tasks:
                task_id = task.get("id", "")
                title = task.get("title", "")
                # 如果标题为None或空字符串，显示"未知标题"
                if not title:
                    title = "未知标题"
                link = task.get("link", "")
                status = task.get("status", "")
                created_at = task.get("created_at", "")
                updated_at = task.get("updated_at", "")
                
                # 根据状态设置标签
                tag = status
                if status == "completed":
                    tag = "completed"
                elif status == "failed":
                    tag = "failed"
                elif status == "in_progress":
                    tag = "in_progress"
                else:
                    tag = "pending"
                
                # 插入数据并应用标签
                tree.insert("", "end", values=(
                    task_id,
                    title,
                    link,
                    status,
                    created_at,
                    updated_at
                ), tags=(tag,))
        
        # 按钮框架
        action_frame = tk.Frame(main_frame, bg="#f0f4f8")
        action_frame.pack(fill=tk.X, pady=10)
        
        # 批量操作按钮
        def batch_start():
            selected = tree.selection()
            if not selected:
                messagebox.showinfo("提示", "请选择要开始的任务")
                return
            
            for item_id in selected:
                item = tree.item(item_id)
                task_id = item["values"][0]
                
                # 查找任务
                task = None
                for t in self.history.get("tasks", []):
                    if t.get("id") == task_id:
                        task = t
                        break
                
                if task:
                    link = task.get("link")
                    self._task_queue_append_unique(link)
                    self.update_queue_status()
                    self.append_log(f"开始处理任务：{link}")
            
            # 自动开始处理队列
            if not self.processing_queue and self._task_queue_len() > 0:
                self.start_queue_processing()
        
        def set_prompt_for_task():
            """为选定的任务设置用户提示词"""
            selected = tree.selection()
            if not selected:
                messagebox.showinfo("提示", "请选择要设置提示词的任务")
                return
            
            # 获取当前的用户提示词
            current_prompt = self.user_prompt_var.get().strip()
            
            # 创建提示词设置窗口
            prompt_window = tk.Toplevel(history_window)
            prompt_window.title("设置用户提示词")
            prompt_window.geometry("600x300")
            prompt_window.configure(bg="#f0f4f8")
            
            # 提示词输入区域
            prompt_frame = tk.Frame(prompt_window, bg="#ffffff")
            prompt_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
            prompt_frame.configure(highlightbackground="#0066cc", highlightthickness=1)
            
            ttk.Label(prompt_frame, text="用户提示词：", font=("微软雅黑", 10), background="#ffffff").pack(anchor=tk.W, padx=15, pady=(15, 5))
            
            prompt_text = tk.Text(prompt_frame, height=8, font=("微软雅黑", 10), bg="#f9f9f9")
            prompt_text.pack(fill=tk.BOTH, expand=True, padx=15, pady=(0, 15))
            prompt_text.insert(tk.END, current_prompt)
            
            # 字符计数
            char_count_var = tk.StringVar(value=f"{len(current_prompt)}/500")
            char_count_label = tk.Label(prompt_frame, textvariable=char_count_var, font=("微软雅黑", 9), foreground="#999", background="#ffffff")
            char_count_label.pack(anchor=tk.E, padx=15, pady=(0, 10))
            
            def update_char_count(event):
                text = prompt_text.get("1.0", tk.END).strip()
                if len(text) > 500:
                    prompt_text.delete("1.0", tk.END)
                    prompt_text.insert(tk.END, text[:500])
                    text = text[:500]
                char_count_var.set(f"{len(text)}/500")
            
            prompt_text.bind("<KeyRelease>", update_char_count)
            
            # 应用按钮
            def apply_prompt():
                new_prompt = prompt_text.get("1.0", tk.END).strip()
                
                for item_id in selected:
                    item = tree.item(item_id)
                    task_id = item["values"][0]
                    
                    # 查找任务
                    task = None
                    for t in self.history.get("tasks", []):
                        if t.get("id") == task_id:
                            task = t
                            break
                    
                    if task:
                        task["user_prompt"] = new_prompt
                        task["updated_at"] = datetime.now().isoformat()
                        self.append_log(f"为任务设置提示词：{task.get('link')}")
                
                # 保存历史记录
                save_history(self.history)
                messagebox.showinfo("成功", f"已为 {len(selected)} 个任务设置提示词")
                prompt_window.destroy()
                refresh_tree()
            
            # 按钮区域
            button_frame = tk.Frame(prompt_window, bg="#f0f4f8")
            button_frame.pack(fill=tk.X, padx=20, pady=(0, 20))
            
            ttk.Button(button_frame, text="应用", command=apply_prompt).pack(side=tk.RIGHT, padx=10)
            ttk.Button(button_frame, text="取消", command=prompt_window.destroy).pack(side=tk.RIGHT, padx=10)
        
        def batch_stop():
            selected = tree.selection()
            if not selected:
                messagebox.showinfo("提示", "请选择要停止的任务")
                return
            
            stopped_count = 0
            for item_id in selected:
                item = tree.item(item_id)
                task_id = item["values"][0]
                
                # 查找任务
                task = None
                for t in self.history.get("tasks", []):
                    if t.get("id") == task_id:
                        task = t
                        break
                
                if task:
                    link = task.get("link")
                    # 使用新的stop_task方法停止任务
                    if self.stop_task(link):
                        stopped_count += 1
                        # 从队列中移除
                        if self._task_queue_remove_if_present(link):
                            self.update_queue_status()
            
            if stopped_count > 0:
                messagebox.showinfo("成功", f"已停止 {stopped_count} 个任务")
                refresh_tree()
        
        def move_task_up():
            """将选中的任务在队列中上移"""
            selected = tree.selection()
            if not selected:
                messagebox.showinfo("提示", "请选择要上移的任务")
                return
            
            if len(selected) > 1:
                messagebox.showinfo("提示", "请只选择一个任务进行移动")
                return
            
            item = tree.item(selected[0])
            task_id = item["values"][0]
            
            # 查找任务
            task = None
            for t in self.history.get("tasks", []):
                if t.get("id") == task_id:
                    task = t
                    break
            
            if task:
                link = task.get("link")
                if self.move_task_in_queue(link, "up"):
                    messagebox.showinfo("成功", "任务已上移")
                    refresh_tree()
        
        def move_task_down():
            """将选中的任务在队列中下移"""
            selected = tree.selection()
            if not selected:
                messagebox.showinfo("提示", "请选择要下移的任务")
                return
            
            if len(selected) > 1:
                messagebox.showinfo("提示", "请只选择一个任务进行移动")
                return
            
            item = tree.item(selected[0])
            task_id = item["values"][0]
            
            # 查找任务
            task = None
            for t in self.history.get("tasks", []):
                if t.get("id") == task_id:
                    task = t
                    break
            
            if task:
                link = task.get("link")
                if self.move_task_in_queue(link, "down"):
                    messagebox.showinfo("成功", "任务已下移")
                    refresh_tree()
        
        def clear_completed():
            """清理已完成的任务"""
            if messagebox.askyesno("确认", "确定要清理所有已完成的任务吗？"):
                completed_tasks = [task for task in self.history.get("tasks", []) 
                                  if task.get("status") == "completed"]
                
                for task in completed_tasks:
                    self.history["tasks"].remove(task)
                
                save_history(self.history)
                self.append_log(f"已清理 {len(completed_tasks)} 个已完成任务")
                refresh_tree()
        
        ttk.Button(action_frame, text="查看详情", command=view_details).pack(side=tk.LEFT, padx=10)
        ttk.Button(action_frame, text="继续处理", command=continue_task).pack(side=tk.LEFT, padx=10)
        ttk.Button(action_frame, text="批量开始", command=batch_start).pack(side=tk.LEFT, padx=10)
        ttk.Button(action_frame, text="批量停止", command=batch_stop).pack(side=tk.LEFT, padx=10)
        ttk.Button(action_frame, text="⬆️ 上移", command=move_task_up).pack(side=tk.LEFT, padx=10)
        ttk.Button(action_frame, text="⬇️ 下移", command=move_task_down).pack(side=tk.LEFT, padx=10)
        ttk.Button(action_frame, text="设置提示词", command=set_prompt_for_task).pack(side=tk.LEFT, padx=10)
        ttk.Button(action_frame, text="清理已完成", command=clear_completed).pack(side=tk.LEFT, padx=10)
        ttk.Button(action_frame, text="刷新", command=refresh_tree).pack(side=tk.LEFT, padx=10)
        ttk.Button(action_frame, text="关闭", command=history_window.destroy).pack(side=tk.RIGHT, padx=10)
        
        # 任务统计信息
        stats_frame = tk.Frame(main_frame, bg="#f0f4f8")
        stats_frame.pack(fill=tk.X, pady=(0, 10))
        
        total_tasks = len(self.history.get("tasks", []))
        completed_tasks = len([t for t in self.history.get("tasks", []) if t.get("status") == "completed"])
        failed_tasks = len([t for t in self.history.get("tasks", []) if t.get("status") == "failed"])
        pending_tasks = len([t for t in self.history.get("tasks", []) if t.get("status") == "pending"])
        in_progress_tasks = len([t for t in self.history.get("tasks", []) if t.get("status") == "in_progress"])
        
        stats_label = tk.Label(
            stats_frame, 
            text=f"任务统计：总计 {total_tasks} | 已完成 {completed_tasks} | 处理中 {in_progress_tasks} | 待处理 {pending_tasks} | 失败 {failed_tasks}", 
            font=("微软雅黑", 9),
            foreground="#666",
            bg="#f0f4f8"
        )
        stats_label.pack(anchor=tk.W, padx=10)
    
    # 批量导入方法
    def batch_import(self):
        """批量导入视频链接，支持先设定user prompt再提交"""
        # 选择Excel文件
        file_path = filedialog.askopenfilename(
            title="选择Excel文件",
            filetypes=[("Excel文件", "*.xlsx;*.xls"), ("所有文件", "*.*")]
        )
        
        if not file_path:
            return
        
        try:
            # 解析Excel文件
            self.append_log(f"开始解析Excel文件：{file_path}")
            links = self.parse_excel_file(file_path)
            
            if not links:
                messagebox.showwarning("提示", "Excel文件中未找到有效的视频链接")
                return
            
            # 创建批量导入设置窗口
            import_window = tk.Toplevel(self.root)
            import_window.title("批量导入设置")
            import_window.geometry("800x600")
            import_window.configure(bg="#f0f4f8")
            # 允许窗口大小调整
            import_window.resizable(True, True)
            
            # 主容器
            main_frame = tk.Frame(import_window, bg="#f0f4f8")
            main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
            
            # 导入信息
            info_frame = tk.Frame(main_frame, bg="#ffffff")
            info_frame.pack(fill=tk.X, pady=(0, 20))
            info_frame.configure(highlightbackground="#0066cc", highlightthickness=1)
            
            ttk.Label(info_frame, text=f"发现 {len(links)} 个视频链接", font=("微软雅黑", 10), background="#ffffff").pack(anchor=tk.W, padx=15, pady=15)
            
            # 用户提示词设置
            prompt_frame = tk.Frame(main_frame, bg="#ffffff")
            prompt_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 20))
            prompt_frame.configure(highlightbackground="#0066cc", highlightthickness=1)
            
            ttk.Label(prompt_frame, text="用户提示词（批量设置）：", font=("微软雅黑", 10), background="#ffffff").pack(anchor=tk.W, padx=15, pady=(15, 5))
            
            # 提示词输入区域
            prompt_text = tk.Text(prompt_frame, height=8, font=("微软雅黑", 10), bg="#f9f9f9")
            prompt_text.pack(fill=tk.BOTH, expand=True, padx=15, pady=(0, 15))
            
            # 获取当前的用户提示词
            current_prompt = self.user_prompt_var.get().strip()
            if current_prompt:
                prompt_text.insert(tk.END, current_prompt)
            
            # 字符计数
            char_count_var = tk.StringVar(value=f"{len(current_prompt)}/500")
            char_count_label = tk.Label(prompt_frame, textvariable=char_count_var, font=("微软雅黑", 9), foreground="#999", background="#ffffff")
            char_count_label.pack(anchor=tk.E, padx=15, pady=(0, 10))
            
            def update_char_count(event):
                text = prompt_text.get("1.0", tk.END).strip()
                if len(text) > 500:
                    prompt_text.delete("1.0", tk.END)
                    prompt_text.insert(tk.END, text[:500])
                    text = text[:500]
                char_count_var.set(f"{len(text)}/500")
            
            prompt_text.bind("<KeyRelease>", update_char_count)
            
            # 飞书：本批次是否同步 + 单批路径（覆盖 AI 配置中的默认路径）
            feishu_frame = tk.Frame(main_frame, bg="#ffffff")
            feishu_frame.pack(fill=tk.X, pady=(0, 20))
            feishu_frame.configure(highlightbackground="#0066cc", highlightthickness=1)
            
            batch_feishu_sync_var = tk.BooleanVar(value=bool(CONFIG.get("feishu_sync_enabled", False)))
            ttk.Checkbutton(
                feishu_frame,
                text="本批次同步到飞书（关闭则仅入队分析，不触发飞书上传；与主界面勾选独立可改）",
                variable=batch_feishu_sync_var,
            ).pack(anchor=tk.W, padx=15, pady=(12, 5))
            
            tk.Label(feishu_frame, text="本批次飞书文件夹路径（可选，留空用 AI 配置中的默认路径）：", font=("微软雅黑", 10), bg="#ffffff", fg="#333").pack(anchor=tk.W, padx=15, pady=(5, 5))
            
            feishu_folder_var = tk.StringVar(value=(CONFIG.get("feishu_default_folder_path") or ""))
            feishu_folder_entry = ttk.Entry(feishu_frame, textvariable=feishu_folder_var, font=("微软雅黑", 10))
            feishu_folder_entry.pack(fill=tk.X, padx=15, pady=(0, 10))
            
            tk.Label(feishu_frame, text="示例：就业技术文档集/八股", font=("微软雅黑", 9), fg="#999", bg="#ffffff").pack(anchor=tk.W, padx=15, pady=(0, 10))
            
            # 按钮区域
            button_frame = tk.Frame(main_frame, bg="#f0f4f8")
            button_frame.pack(fill=tk.X, pady=10)
            
            def start_import():
                """开始导入"""
                user_prompt = prompt_text.get("1.0", tk.END).strip()
                if batch_feishu_sync_var.get():
                    raw_f = feishu_folder_var.get().strip()
                    feishu_folder_path = raw_f or (CONFIG.get("feishu_default_folder_path") or "").strip() or None
                else:
                    feishu_folder_path = None
                
                # 批量添加到历史记录
                self.append_log(f"开始创建 {len(links)} 个任务的历史记录")
                new_links_count = 0
                existing_links_count = 0
                
                for link in links:
                    # 检查链接是否已经导入
                    if not self.is_link_already_imported(link):
                        self.add_task_to_history(
                            link,
                            user_prompt,
                            feishu_folder_path,
                            feishu_sync_override=batch_feishu_sync_var.get(),
                        )
                        self._task_queue_append_unique(link)
                        new_links_count += 1
                    else:
                        existing_links_count += 1
                        self.append_log(f"链接已存在，跳过导入：{link}")
                
                self.append_log(f"批量导入完成：新添加 {new_links_count} 个链接，跳过 {existing_links_count} 个已存在的链接")
                
                # 更新队列状态显示
                self.update_queue_status()
                
                self.append_log(f"成功导入 {new_links_count} 个视频链接到队列")
                self.append_log(f"当前待处理：{self._task_queue_len()}，执行中：{len(self.active_futures)}")
                
                # 自动开始处理队列（已在处理时由调度器自动取新任务）
                if new_links_count > 0 and not self.processing_queue:
                    self.start_queue_processing()
                
                import_window.destroy()
                messagebox.showinfo("成功", f"批量导入完成：新添加 {new_links_count} 个链接")
            
            ttk.Button(button_frame, text="确认", command=start_import).pack(side=tk.RIGHT, padx=10)
            ttk.Button(button_frame, text="取消", command=import_window.destroy).pack(side=tk.RIGHT, padx=10)
            
        except Exception as e:
            self.append_log(f"批量导入失败：{e}")
            messagebox.showerror("错误", f"批量导入失败：{e}")
    
    # 解析Excel文件
    def parse_excel_file(self, file_path):
        try:
            # 尝试使用pandas解析
            try:
                import pandas as pd
                df = pd.read_excel(file_path, header=None)
                links = df.iloc[:, 0].dropna().tolist()
                
            except ImportError:
                # 如果pandas不可用，使用openpyxl
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(file_path)
                    ws = wb.active
                    links = []
                    for row in ws.iter_rows(min_row=1, values_only=True):
                        if row[0]:
                            links.append(str(row[0]))
                except ImportError:
                    # 如果都不可用，提示安装依赖
                    self.append_log("请安装pandas或openpyxl来解析Excel文件")
                    raise Exception("缺少Excel解析库，请安装pandas或openpyxl")
            
            # 过滤和验证链接
            valid_links = []
            for link in links:
                link = str(link).strip()
                if link and ('http' in link.lower() or 'www.' in link.lower()):
                    valid_links.append(link)
            
            return valid_links
        except Exception as e:
            self.append_log(f"解析Excel文件失败：{e}")
            raise
    

    # 开始队列处理（线程池滑动窗口：最多 max_workers 路并行，FIFO，新入队任务不必等整批结束）
    def start_queue_processing(self):
        with self._scheduler_start_lock:
            if self.processing_queue:
                self.append_log("队列调度已在运行，新任务将在有空闲线程时按顺序自动执行")
                self.update_queue_status()
                return
            if self._task_queue_len() == 0:
                self.append_log("队列为空，无需处理")
                self.update_queue_status()
                return
            self.processing_queue = True

        self.update_queue_status()
        self.append_log(
            f"启动队列调度：最大并行 {self.max_workers}，当前待处理 {self._task_queue_len()}（先进先出，接近 Java 线程池行为）"
        )
        self._pipeline_log(f"scheduler_start workers={self.max_workers} pending={self._task_queue_len()}")
        threading.Thread(target=self._queue_scheduler_loop, daemon=True).start()

    def _queue_scheduler_loop(self):
        """维持不超过 max_workers 个并发；有空槽即从队首取任务提交。"""
        import time as time_mod

        poll = float(os.environ.get("PIPELINE_SCHEDULER_POLL_SEC", "0.5"))
        timeout_sec = int(os.environ.get("PIPELINE_BATCH_TIMEOUT_SEC", "7200"))
        wall0 = time_mod.time()
        active = {}
        completed_count = 0

        def _submit_link(link: str) -> None:
            task_prompt = ""
            feishu_folder_path = None
            for task in self.history.get("tasks", []):
                if task.get("link") == link:
                    task_prompt = task.get("user_prompt", "")
                    feishu_folder_path = task.get("feishu_folder_path")
                    break
            if not task_prompt:
                task_prompt = self.user_prompt_var.get().strip()
            self.append_log(f"任务提示词：{task_prompt[:50]}{'...' if len(task_prompt) > 50 else ''}")
            if feishu_folder_path:
                self.append_log(f"飞书文件夹路径：{feishu_folder_path}")
            cancel_event = threading.Event()
            self.task_cancel_flags[link] = cancel_event
            fut = self.executor.submit(
                self._run_pipeline_with_cancel, link, task_prompt, feishu_folder_path, cancel_event
            )
            self.active_futures[link] = fut
            active[link] = fut
            self.append_log(
                f"提交到线程池（并行 {len(active)}/{self.max_workers}）：{link[:120]}{'...' if len(link) > 120 else ''}"
            )
            self._pipeline_log(f"submit active={len(active)}")
            self.update_queue_status()

        try:
            while True:
                while len(active) < self.max_workers:
                    link = self._task_queue_pop_front()
                    if link is None:
                        break
                    _submit_link(link)

                if not active:
                    if self._task_queue_len() == 0:
                        self.append_log("所有队列任务处理完成")
                        self._pipeline_log("scheduler_done all_complete")
                        break
                    time_mod.sleep(poll)
                    continue

                if time_mod.time() - wall0 > timeout_sec:
                    self.append_log(f"队列调度总超时（{timeout_sec}s），正在取消未完成任务", "ERROR")
                    self._pipeline_log(f"scheduler_timeout active={len(active)}")
                    for f in list(active.values()):
                        f.cancel()
                    for link in list(active.keys()):
                        active.pop(link, None)
                        self.task_cancel_flags.pop(link, None)
                        self.active_futures.pop(link, None)
                    break

                done, _ = concurrent.futures.wait(
                    set(active.values()),
                    timeout=poll,
                    return_when=concurrent.futures.FIRST_COMPLETED,
                )
                for fut in done:
                    link = None
                    for lk, ff in list(active.items()):
                        if ff is fut:
                            link = lk
                            break
                    if link is None:
                        continue
                    try:
                        fut.result()
                    except concurrent.futures.CancelledError:
                        self.append_log("任务已取消", "WARNING")
                        self._pipeline_log("future_cancelled")
                    except Exception as e:
                        self.append_log(f"任务执行异常：{type(e).__name__}: {e}", "ERROR")
                        self._pipeline_log(f"future_exc {type(e).__name__}: {e!r}")
                    finally:
                        completed_count += 1
                        active.pop(link, None)
                        self.task_cancel_flags.pop(link, None)
                        self.active_futures.pop(link, None)
                        self.append_log(
                            f"任务完成进度：{completed_count}（当前并行 {len(active)}/{self.max_workers}）"
                        )
                        self.update_queue_status()
        except Exception as e:
            self.append_log(f"队列调度异常：{type(e).__name__}: {e}", "ERROR")
            import traceback
            self.append_log(traceback.format_exc(), "ERROR")
            self._pipeline_log(f"scheduler_exc {type(e).__name__}: {e!r}")
        finally:
            self.processing_queue = False
            self.update_queue_status()

    def stop_task(self, link: str) -> bool:
        """停止正在执行的任务
        
        Args:
            link: 任务链接
            
        Returns:
            bool: 是否成功停止
        """
        try:
            # 设置取消标志
            if link in self.task_cancel_flags:
                self.task_cancel_flags[link].set()
                self.append_log(f"已发送停止信号：{link}")
            
            # 取消future
            if link in self.active_futures:
                future = self.active_futures[link]
                if future and not future.done():
                    cancelled = future.cancel()
                    if cancelled:
                        self.append_log(f"已取消任务：{link}")
                    else:
                        self.append_log(f"任务无法取消（可能已在运行）：{link}")
            
            # 更新任务状态
            self.update_task_status(link, "download", "cancelled")
            self.update_task_status(link, "transcribe", "cancelled")
            self.update_task_status(link, "ai_analysis", "cancelled")
            self.update_task_status(link, "generate_md", "cancelled")
            
            # 从历史记录中更新状态
            for task in self.history.get("tasks", []):
                if task.get("link") == link:
                    task["status"] = "cancelled"
                    break
            save_history(self.history)
            
            return True
        except Exception as e:
            self.append_log(f"停止任务失败：{e}")
            return False
    
    def move_task_in_queue(self, link: str, direction: str) -> bool:
        """调整任务在队列中的位置
        
        Args:
            link: 任务链接
            direction: 移动方向，"up"或"down"
            
        Returns:
            bool: 是否成功移动
        """
        try:
            with self._task_queue_lock:
                if link not in self.task_queue:
                    self.append_log(f"任务不在待处理队列中（可能正在执行）：{link}")
                    return False

                idx = self.task_queue.index(link)

                if direction == "up":
                    if idx == 0:
                        self.append_log("任务已在队列最前面")
                        return False
                    self.task_queue[idx], self.task_queue[idx - 1] = (
                        self.task_queue[idx - 1],
                        self.task_queue[idx],
                    )
                    self.append_log(f"任务已上移：{link}")

                elif direction == "down":
                    if idx == len(self.task_queue) - 1:
                        self.append_log("任务已在队列最后面")
                        return False
                    self.task_queue[idx], self.task_queue[idx + 1] = (
                        self.task_queue[idx + 1],
                        self.task_queue[idx],
                    )
                    self.append_log(f"任务已下移：{link}")

                else:
                    self.append_log(f"未知的移动方向：{direction}")
                    return False

            self.update_queue_status()
            return True

        except Exception as e:
            self.append_log(f"移动任务失败：{e}")
            return False
    
    def get_task_current_stage(self, link: str) -> str:
        """获取任务当前执行的阶段
        
        Args:
            link: 任务链接
            
        Returns:
            str: 当前阶段名称
        """
        for task in self.history.get("tasks", []):
            if task.get("link") == link:
                stages = task.get("stages", {})
                # 按顺序检查各阶段状态
                stage_order = ["download", "transcribe", "ai_analysis", "generate_md"]
                for stage in stage_order:
                    if stage in stages:
                        status = stages[stage].get("status", "")
                        if status == "in_progress":
                            return stage
                        elif status == "failed":
                            return f"{stage}_failed"
                
                # 检查是否已完成
                all_completed = all(
                    stages.get(s, {}).get("status") == "completed" 
                    for s in stage_order if s in stages
                )
                if all_completed:
                    return "completed"
                
                # 检查是否已取消
                if task.get("status") == "cancelled":
                    return "cancelled"
                
                return "pending"
        
        return "unknown"
    
    # 单独重试失败阶段
    def _retry_stage(self, link, stage, task):
        """单独重试失败的阶段，使用缓存的结果"""
        try:
            self.append_log(f"开始重试阶段 {stage}：{link}")
            
            # 获取缓存的结果
            stages = task.get("stages", {})
            cached_results = {}
            
            # 收集已完成阶段的结果
            for s, info in stages.items():
                if info.get("status") == "completed" and info.get("result"):
                    cached_results[s] = info.get("result")
            
            # 根据阶段名称执行相应的操作
            if stage == "download":
                # 重新下载视频
                self.update_task_status(link, stage, "in_progress")
                video_file = self.download_video(link)
                if video_file:
                    self.update_task_status(link, stage, "completed", video_file)
                    # 自动继续下一个阶段
                    self._continue_from_stage(link, "transcribe", {"download": video_file})
                else:
                    self.update_task_status(link, stage, "failed")
                    self.append_log(f"阶段 {stage} 重试失败")
            
            elif stage == "transcribe":
                # 重新转写
                video_file = cached_results.get("download")
                if not video_file:
                    self.append_log(f"缺少前一阶段结果，无法重试 {stage}")
                    return
                
                self.update_task_status(link, stage, "in_progress")
                result_data = self.speech_to_text(video_file, "")
                if result_data:
                    self.update_task_status(link, stage, "completed", result_data.get("segments", []))
                    # 更新AI分析结果
                    self.update_task_status(link, "ai_analysis", "completed", result_data.get("ai_summary", ""))
                    # 自动继续下一个阶段
                    self._continue_from_stage(link, "generate_md", {
                        "download": video_file,
                        "transcribe": result_data.get("segments", []),
                        "ai_analysis": result_data.get("ai_summary", "")
                    })
                else:
                    self.update_task_status(link, stage, "failed")
                    self.append_log(f"阶段 {stage} 重试失败")
            
            elif stage == "ai_analysis":
                # 重新AI分析
                segments = cached_results.get("transcribe")
                if not segments:
                    self.append_log(f"缺少前一阶段结果，无法重试 {stage}")
                    return
                
                self.update_task_status(link, stage, "in_progress")
                # 构建临时结果数据
                result_data = {"segments": segments}
                # 提取转写文本
                transcript_lines = []
                for seg in segments:
                    text = seg if isinstance(seg, str) else seg.get("text", "")
                    transcript_lines.append(text)
                transcript = " ".join(transcript_lines)
                # 重新分析
                summary = self.summarize_with_volcengine(transcript, "")
                if summary:
                    self.update_task_status(link, stage, "completed", summary)
                    # 自动继续下一个阶段
                    self._continue_from_stage(link, "generate_md", {
                        "download": cached_results.get("download"),
                        "transcribe": segments,
                        "ai_analysis": summary
                    })
                else:
                    self.update_task_status(link, stage, "failed")
                    self.append_log(f"阶段 {stage} 重试失败")
            
            elif stage == "generate_md":
                # 重新生成MD
                segments = cached_results.get("transcribe")
                summary = cached_results.get("ai_analysis")
                if not segments or not summary:
                    self.append_log(f"缺少前一阶段结果，无法重试 {stage}")
                    return
                
                self.update_task_status(link, stage, "in_progress")
                # 构建临时结果数据
                result_data = {
                    "segments": segments,
                    "ai_summary": summary
                }
                md_file = self.generate_md(result_data, link, "视频")
                if md_file:
                    self.update_task_status(link, stage, "completed", md_file)
                    self.append_log(f"阶段 {stage} 重试成功：{md_file}")
                else:
                    self.update_task_status(link, stage, "failed")
                    self.append_log(f"阶段 {stage} 重试失败")
            
            self.append_log(f"阶段 {stage} 重试完成")
            
        except Exception as e:
            self.append_log(f"重试阶段 {stage} 异常：{e}")
            self.update_task_status(link, stage, "failed")
    
    # 从指定阶段继续处理
    def _continue_from_stage(self, link, next_stage, cached_results):
        """从指定阶段继续处理"""
        try:
            self.append_log(f"从阶段 {next_stage} 继续处理：{link}")
            
            if next_stage == "transcribe":
                video_file = cached_results.get("download")
                if video_file:
                    self.update_task_status(link, next_stage, "in_progress")
                    result_data = self.speech_to_text(video_file, "")
                    if result_data:
                        self.update_task_status(link, next_stage, "completed", result_data.get("segments", []))
                        self.update_task_status(link, "ai_analysis", "completed", result_data.get("ai_summary", ""))
                        self._continue_from_stage(link, "generate_md", {
                            "download": video_file,
                            "transcribe": result_data.get("segments", []),
                            "ai_analysis": result_data.get("ai_summary", "")
                        })
                    else:
                        self.update_task_status(link, next_stage, "failed")
            
            elif next_stage == "generate_md":
                segments = cached_results.get("transcribe")
                summary = cached_results.get("ai_analysis")
                if segments and summary:
                    self.update_task_status(link, next_stage, "in_progress")
                    result_data = {
                        "segments": segments,
                        "ai_summary": summary
                    }
                    md_file = self.generate_md(result_data, link, "视频")
                    if md_file:
                        self.update_task_status(link, next_stage, "completed", md_file)
                        self.append_log(f"任务处理完成：{link}")
                        # 清理缓存文件
                        self.cleanup_task_cache(link, cached_results)
                    else:
                        self.update_task_status(link, next_stage, "failed")
        
        except Exception as e:
            self.append_log(f"继续处理异常：{e}")
    
    # 清理任务缓存文件
    def cleanup_task_cache(self, link, cached_results):
        """清理任务的缓存文件"""
        try:
            # 清理视频文件
            video_file = cached_results.get("download")
            if video_file and os.path.exists(video_file):
                with self.file_operation_lock:
                    try:
                        os.remove(video_file)
                        self.append_log(f"已清理视频缓存文件：{video_file}")
                    except PermissionError:
                        self.append_log(f"文件被占用，跳过清理：{video_file}")
                    except Exception as e:
                        self.append_log(f"清理视频文件失败：{e}")
            
            # 清理可能的临时文件
            import glob
            url_hash = hashlib.md5(link.encode()).hexdigest()[:8]
            temp_files = glob.glob(os.path.join(VIDEO_DIR, f"*{url_hash}*"))
            for temp_file in temp_files:
                if os.path.exists(temp_file):
                    with self.file_operation_lock:
                        try:
                            os.remove(temp_file)
                            self.append_log(f"已清理临时缓存文件：{temp_file}")
                        except PermissionError:
                            self.append_log(f"临时文件被占用，跳过清理：{temp_file}")
                        except Exception as e:
                            self.append_log(f"清理临时文件失败：{e}")
        except Exception as e:
            self.append_log(f"清理缓存文件异常：{e}")
    
    # 清理所有完成任务的缓存文件
    def cleanup_completed_tasks(self):
        """清理所有完成任务的缓存文件"""
        try:
            completed_tasks = [task for task in self.history.get("tasks", []) 
                              if task.get("status") == "completed"]
            
            for task in completed_tasks:
                stages = task.get("stages", {})
                cached_results = {}
                
                # 检查是否已经生成了MD文件
                generate_md_stage = stages.get("generate_md", {})
                if generate_md_stage.get("status") != "completed":
                    # 如果没有生成MD文件，不清理缓存
                    continue
                
                for s, info in stages.items():
                    if info.get("status") == "completed" and info.get("result"):
                        cached_results[s] = info.get("result")
                
                # 清理缓存文件
                if cached_results:
                    self.cleanup_task_cache(task.get("link"), cached_results)
        except Exception as e:
            self.append_log(f"清理完成任务缓存异常：{e}")
    
    def _detect_platform(self, link: str) -> str:
        """检测链接类型"""
        if 'xiaohongshu.com' in link.lower():
            # 所有小红书 /explore/ 路径的链接都视为视频（让 link_analyzer 进一步判断）
            if '/explore/' in link:
                return "小红书视频"
            else:
                return "小红书图文"
        elif 'mp.weixin.qq.com' in link.lower():
            # 微信公众号文章
            return "微信公众号"
        elif 'douyin.com' in link.lower() or 'tiktok.com' in link.lower():
            return "抖音"
        elif 'bilibili.com' in link.lower():
            return "B 站"
        elif 'youtube.com' in link.lower():
            return "YouTube"
        else:
            return "视频"
    
    def _run_xiaohongshu_analysis(self, link: str, user_prompt: str = "", feishu_folder_path: str = None):
        """运行小红书图文分析流程"""
        try:
            # 导入link_analyzer模块
            try:
                from link_analyzer import LinkAnalyzer
            except ImportError as e:
                self.append_log(f"导入link_analyzer模块失败：{e}")
                self.update_task_status(link, "download", "failed")
                self.update_progress(0, "失败")
                return
            
            # 初始化分析器
            analyzer = LinkAnalyzer()
            
            # 阶段1：提取图文内容
            self.update_task_status(link, "download", "in_progress")
            self.update_progress(10, "提取小红书图文内容...")
            self.append_log("开始分析小红书链接...")
            
            result = analyzer.analyze_link(link)
            
            if not result or result.get('error'):
                error_msg = result.get('error', '未知错误') if result else '分析失败'
                self.append_log(f"小红书内容检测失败：{error_msg}")
                self.update_task_status(link, "download", "failed")
                self.update_progress(0, "失败")
                return
            
            # 【关键】检查返回的类型
            content_type = result.get('type', '')
            self.append_log(f"小红书内容类型检测结果：{content_type}")
            
            if content_type == 'video':
                # 是视频！跳转到视频下载流程
                self.append_log("✓ 检测到小红书视频，跳转到视频下载流程...")
                # 直接调用 download_video 方法（使用 yt-dlp 下载）
                video_file = self.download_video(link)
                if video_file:
                    self.append_log(f"✓ 视频下载成功：{video_file}")
                    # 继续后续处理（转写、生成 MD 等）- 复用主流程逻辑
                    self._process_downloaded_video(video_file, link, user_prompt, feishu_folder_path)
                else:
                    self.append_log("✗ 视频下载失败")
                    self.update_task_status(link, "download", "failed")
                return
            elif content_type == 'xiaohongshu':
                # 是图文，继续图文分析
                self.append_log("✓ 检测到小红书图文，继续图文分析...")
            else:
                # 未知类型，默认按图文处理
                self.append_log(f"⚠ 未知内容类型：{content_type}，默认按图文处理")
            
            self.update_task_status(link, "download", "completed", result.get('summary', ''))
            self.append_log(f"成功提取 {len(result.get('image_links', []))} 张图片")
            
            # 阶段2：AI分析
            self.update_task_status(link, "ai_analysis", "in_progress")
            self.update_progress(60, "使用AI进行内容分析...")
            
            summary = result.get('summary', '')
            if not summary:
                self.append_log("提取的内容为空，无法进行AI分析")
                self.update_task_status(link, "ai_analysis", "failed")
                self.update_progress(0, "失败")
                return
            
            # 调用火山引擎AI进行分析
            ai_summary = self.summarize_with_volcengine(summary, "")
            
            if not ai_summary:
                self.append_log("AI分析失败")
                self.update_task_status(link, "ai_analysis", "failed")
                self.update_progress(0, "失败")
                return
            
            self.update_task_status(link, "ai_analysis", "completed", ai_summary)
            
            # 从AI分析结果中提取标题
            title = self.extract_title_from_summary(ai_summary, link)
            if not title or title == "未知标题":
                title = result.get('title', '小红书内容分析')
                # 清理标题
                import re
                title = re.sub(r' - 小红书$', '', title)
                title = re.sub(r'[\\/:*?"<>|]', '', title)
                title = title.replace(' ', '_')
            
            # 更新任务标题
            for task in self.history.get("tasks", []):
                if task.get("link") == link:
                    task["title"] = title
                    task["updated_at"] = datetime.now().isoformat()
                    save_history(self.history)
                    self.append_log(f"更新任务标题：{title}")
                    break
            
            # 阶段3：生成Markdown
            self.update_task_status(link, "generate_md", "in_progress")
            self.update_progress(80, "生成Markdown文档...")
            
            # 构建result_data格式
            result_data = {
                "transcript": summary,
                "ai_summary": ai_summary,
                "platform": "小红书",
                "title": title,
                "image_analysis": result.get('image_analysis', []),
                "expected_image_count": result.get('expected_image_count', 0),
                "actual_image_count": len(result.get('image_links', []))
            }
            
            md_file = self.generate_md(result_data, link, "小红书")
            
            if not md_file:
                self.append_log("生成文档失败")
                self.update_task_status(link, "generate_md", "failed")
                self.update_progress(0, "失败")
                return
            
            self.update_task_status(link, "generate_md", "completed", md_file)
            
            # 阶段4：上传到飞书（勾选 + AI 配置中的凭证）
            self._run_feishu_upload_if_enabled(link, md_file, user_prompt, feishu_folder_path)
            
            self.update_progress(100, "完成")
            self.append_log("小红书图文分析完成！")
            
        except Exception as e:
            self.append_log(f"小红书图文分析异常：{e}")
            import traceback
            traceback.print_exc()
            self.update_task_status(link, "download", "failed")
            self.update_progress(0, "失败")
    
    def _run_wechat_article_analysis(self, link: str, user_prompt: str = "", feishu_folder_path: str = None):
        """运行微信公众号文章分析流程"""
        try:
            # 导入微信公众号文章处理器
            try:
                from wechat_article_processor import WeChatArticleProcessor
            except ImportError as e:
                self.append_log(f"导入wechat_article_processor模块失败：{e}")
                self.update_task_status(link, "download", "failed")
                self.update_progress(0, "失败")
                return
            
            # 初始化处理器
            processor = WeChatArticleProcessor()
            
            # 阶段1：提取文章内容
            self.update_task_status(link, "download", "in_progress")
            self.update_progress(10, "提取微信公众号文章内容...")
            self.append_log("开始分析微信公众号文章链接...")
            
            result = processor.extract_article(link)
            
            if not result or result.get('error'):
                error_msg = result.get('error', '未知错误') if result else '分析失败'
                self.append_log(f"微信公众号文章提取失败：{error_msg}")
                self.update_task_status(link, "download", "failed")
                self.update_progress(0, "失败")
                return
            
            # 检查提取的内容是否有效
            title = result.get('title', '')
            content = result.get('content', '')
            
            if title == "未找到标题" and len(content) < 100:
                self.append_log(f"文章提取失败：无法获取有效内容，可能是链接无效或文章已删除")
                self.update_task_status(link, "download", "failed")
                self.update_progress(0, "失败")
                return
            
            # 生成摘要
            summary = processor.generate_summary(result)
            
            self.update_task_status(link, "download", "completed", summary)
            self.append_log(f"成功提取文章: {title}")
            self.append_log(f"作者: {result.get('author', '未知')}")
            self.append_log(f"图片数量: {result.get('image_count', 0)}")
            self.append_log(f"正文长度: {len(content)} 字符")
            
            # 阶段2：AI分析
            self.update_task_status(link, "ai_analysis", "in_progress")
            self.update_progress(60, "使用AI进行内容分析...")
            
            # 验证摘要内容
            if not summary or len(summary) < 200:
                self.append_log("提取的内容为空或太短，无法进行AI分析")
                self.update_task_status(link, "ai_analysis", "failed")
                self.update_progress(0, "失败")
                return
            
            # 调用火山引擎AI进行分析
            ai_summary = self.summarize_with_volcengine(summary, "")
            
            if not ai_summary:
                self.append_log("AI分析失败")
                self.update_task_status(link, "ai_analysis", "failed")
                self.update_progress(0, "失败")
                return
            
            self.update_task_status(link, "ai_analysis", "completed", ai_summary)
            
            # 从AI分析结果中提取标题
            title = self.extract_title_from_summary(ai_summary, link)
            if not title or title == "未知标题":
                title = result.get('title', '微信公众号文章分析')
                # 清理标题
                import re
                title = re.sub(r'[\\/:*?"<>|]', '', title)
                title = title.replace(' ', '_')
            
            # 更新任务标题
            for task in self.history.get("tasks", []):
                if task.get("link") == link:
                    task["title"] = title
                    task["updated_at"] = datetime.now().isoformat()
                    save_history(self.history)
                    self.append_log(f"更新任务标题：{title}")
                    break
            
            # 阶段3：生成Markdown
            self.update_task_status(link, "generate_md", "in_progress")
            self.update_progress(80, "生成Markdown文档...")
            
            # 构建result_data格式
            result_data = {
                "transcript": summary,
                "ai_summary": ai_summary,
                "platform": "微信公众号",
                "title": title,
                "image_analysis": result.get('image_analysis', []),
                "expected_image_count": result.get('image_count', 0),
                "actual_image_count": len(result.get('image_links', [])),
                "author": result.get('author', ''),
                "publish_time": result.get('publish_time', '')
            }
            
            md_file = self.generate_md(result_data, link, "微信公众号")
            
            if not md_file:
                self.append_log("生成文档失败")
                self.update_task_status(link, "generate_md", "failed")
                self.update_progress(0, "失败")
                return
            
            self.update_task_status(link, "generate_md", "completed", md_file)
            
            # 阶段4：上传到飞书
            self._run_feishu_upload_if_enabled(link, md_file, user_prompt, feishu_folder_path)
            
            self.update_progress(100, "完成")
            self.append_log("微信公众号文章分析完成！")
            
        except Exception as e:
            self.append_log(f"微信公众号文章分析异常：{e}")
            import traceback
            traceback.print_exc()
            self.update_task_status(link, "download", "failed")
            self.update_progress(0, "失败")
    
    # 从摘要中提取标题
    def extract_title_from_summary(self, summary, link):
        """从AI分析结果中提取标题"""
        try:
            # 检查是否有RULES控制标题提取
            rules = CONFIG.get("rules", "")
            
            # 从摘要中提取标题
            if summary:
                # 尝试提取第一行内容作为标题
                lines = summary.split('\n')
                for i, line in enumerate(lines):
                    line = line.strip()
                    if not line:
                        continue
                    
                    # 如果是Markdown标题格式 (# 标题)，提取#后面的内容
                    if line.startswith('#'):
                        # 去除所有#号和后面的空格
                        title = line.lstrip('#').strip()
                        if title and len(title) > 3:
                            # 清理标题，去除特殊字符
                            import re
                            title = re.sub(r'[\\/:*?"<>|]', '', title)
                            # 截取前20个字符作为文件名
                            title = title[:20].strip()
                            # 替换空格为下划线
                            title = title.replace(' ', '_')
                            # 确保标题不为空
                            if title:
                                self.append_log(f"从AI摘要中提取标题：{title}")
                                return title
                    else:
                        # 普通文本行，跳过太短的内容
                        if len(line) > 5:
                            # 清理标题，去除特殊字符
                            import re
                            # 去除标点符号和特殊字符
                            title = re.sub(r'[\\/:*?"<>|]', '', line)
                            # 去除开头的数字和点（如 "1. "）
                            title = re.sub(r'^[\d\.\s]+', '', title)
                            # 截取前20个字符作为文件名
                            title = title[:20].strip()
                            # 替换空格为下划线
                            title = title.replace(' ', '_')
                            # 确保标题不为空
                            if title and title != '_' and len(title) > 3:
                                self.append_log(f"从AI摘要中提取标题：{title}")
                                return title
            
            # 如果摘要中没有合适的标题，从链接中提取
            import re
            # 首先尝试提取B站BV号
            bv_match = re.search(r'BV[0-9A-Za-z]{10}', link)
            if bv_match:
                bv = bv_match.group(0)
                self.append_log(f"从链接中提取B站BV号作为标题：{bv}")
                return bv
            
            # 尝试提取链接中的数字部分
            doc_name_match = re.search(r'\d+', link.split('/')[-1])
            if doc_name_match:
                doc_name = doc_name_match.group(0)
                self.append_log(f"从链接中提取数字作为标题：{doc_name}")
                return doc_name
            
            # 尝试提取链接中的有意义部分
            path_match = re.search(r'[a-zA-Z0-9_-]{8,}', link)
            if path_match:
                path_part = path_match.group(0)
                self.append_log(f"从链接中提取路径部分作为标题：{path_part}")
                return path_part
            
            # 如果都失败了，使用默认值
            default_title = "未知标题"
            self.append_log(f"使用默认标题：{default_title}")
            return default_title
            
        except Exception as e:
            self.append_log(f"提取标题异常：{e}")
            # 从链接中提取作为后备
            import re
            # 首先尝试提取B站BV号
            bv_match = re.search(r'BV[0-9A-Za-z]{10}', link)
            if bv_match:
                return bv_match.group(0)
            # 尝试提取数字
            doc_name_match = re.search(r'\d+', link.split('/')[-1])
            if doc_name_match:
                return doc_name_match.group(0)
            # 尝试提取路径部分
            path_match = re.search(r'[a-zA-Z0-9_-]{8,}', link)
            if path_match:
                return path_match.group(0)
            # 最终默认值
            return "未知标题"
    
    # 窗口关闭事件处理
    def on_closing(self):
        """处理窗口关闭事件"""
        try:
            self.append_log("正在关闭应用程序...")
            # 关闭线程池
            self.append_log("正在关闭线程池...")
            self.executor.shutdown(wait=False)  # 不等待所有任务完成
            self.append_log("线程池已关闭")
        except Exception as e:
            self.append_log(f"关闭应用程序时发生异常：{e}")
        finally:
            # 保存历史记录
            save_history(self.history)
            self.append_log("历史记录已保存")
            # 关闭窗口
            self.root.destroy()
    
    # 入口
    def start(self):
        link = self.link_var.get().strip()
        user_prompt = self.user_prompt_var.get().strip()
        
        # 尝试从文本中提取URL
        extracted_url = self.extract_url_from_text(link)
        if extracted_url:
            # 移除可能的反引号
            extracted_url = extracted_url.strip('`')
            self.append_log(f"从文本中提取到URL: {extracted_url}")
            # 更新输入框为提取的URL
            self.link_var.set(extracted_url)
            link = extracted_url
        
        if not link:
            messagebox.showwarning("提示", "请先输入视频链接")
            return
            
        # 更宽松的链接验证
        if not ('http' in link.lower() or 'www.' in link.lower()):
            messagebox.showwarning("提示", "请输入有效的视频链接\n提示: 链接应该包含 http 或 www")
            return
        
        # 检查链接是否已经导入
        if self.is_link_already_imported(link):
            # 创建自定义对话框，提供取消和继续上传选项
            dialog = tk.Toplevel(self.root)
            dialog.title("链接已导入")
            dialog.geometry("400x200")
            dialog.configure(bg="#f0f4f8")
            dialog.transient(self.root)
            dialog.grab_set()
            
            # 居中显示
            dialog.update_idletasks()
            width = dialog.winfo_width()
            height = dialog.winfo_height()
            x = (dialog.winfo_screenwidth() // 2) - (width // 2)
            y = (dialog.winfo_screenheight() // 2) - (height // 2)
            dialog.geometry(f"{width}x{height}+{x}+{y}")
            
            # 消息标签
            msg_frame = tk.Frame(dialog, bg="#ffffff")
            msg_frame.pack(fill=tk.X, padx=20, pady=(20, 10))
            msg_frame.configure(highlightbackground="#0066cc", highlightthickness=1)
            
            msg_label = tk.Label(
                msg_frame, 
                text="该链接已经导入过，是否继续上传？", 
                font=("微软雅黑", 10),
                foreground="#333",
                bg="#ffffff"
            )
            msg_label.pack(padx=20, pady=20)
            
            # 按钮区域
            btn_frame = tk.Frame(dialog, bg="#f0f4f8")
            btn_frame.pack(fill=tk.X, padx=20, pady=10)
            
            # 存储用户选择的变量
            user_choice = tk.StringVar(value="cancel")
            
            def cancel_upload():
                user_choice.set("cancel")
                dialog.destroy()
            
            def continue_upload():
                user_choice.set("continue")
                dialog.destroy()
            
            # 取消按钮
            cancel_btn = ttk.Button(btn_frame, text="取消上传", command=cancel_upload)
            cancel_btn.pack(side=tk.LEFT, padx=10)
            
            # 继续上传按钮
            continue_btn = ttk.Button(btn_frame, text="继续上传", command=continue_upload)
            continue_btn.pack(side=tk.RIGHT, padx=10)
            
            # 等待用户选择
            self.root.wait_window(dialog)
            
            # 根据用户选择执行操作
            if user_choice.get() == "cancel":
                return
            # 如果用户选择继续上传，不返回，继续执行后续代码

        # 检查任务队列大小限制
        if self._total_queued_work() >= self.queue_max_size:
            messagebox.showwarning(
                "提示",
                f"任务队列已满（待处理+执行中：{self._total_queued_work()}，最大限制：{self.queue_max_size}）\n请稍后再添加任务或调整队列大小限制",
            )
            return

        # 添加到任务队列
        self._task_queue_append_unique(link)
        self.add_task_to_history(link, user_prompt, self._resolve_feishu_folder_for_task())
        self.update_queue_status()
        
        self.start_btn.config(state=tk.DISABLED)
        self.append_log(f"添加任务到队列：链接={link}")
        self.append_log(f"当前待处理：{self._task_queue_len()}，执行中：{len(self.active_futures)}")

        try:
            # 自动开始处理队列（调度已运行时，新链接会在有空闲线程时自动开始）
            if not self.processing_queue:
                self.start_queue_processing()
            else:
                self.append_log("调度运行中：新链接已入队，将在有空闲线程时按顺序自动执行")
        finally:
            # 确保按钮状态恢复
            self.start_btn.config(state=tk.NORMAL)

    def extract_url_from_text(self, text: str) -> str:
        """从文本中提取URL"""
        import re
        
        # 清理文本中的反引号
        text = text.replace('`', '')
        
        # 匹配http/https开头的URL，直到遇到空格或特殊字符
        url_pattern = r'https?://[^\s<>"\'\)\]\}]+'
        urls = re.findall(url_pattern, text)
        
        # 去重
        unique_urls = list(dict.fromkeys(urls))
        
        # 优先处理抖音链接
        for url in unique_urls:
            if 'douyin.com' in url.lower() or 'v.douyin.com' in url.lower():
                # 移除末尾可能的标点符号
                url = url.rstrip('.,;:!?')
                return url
        
        # 处理小红书链接
        for url in unique_urls:
            if 'xiaohongshu.com' in url.lower():
                # 移除末尾可能的标点符号
                url = url.rstrip('.,;:!?')
                return url
        
        # 处理其他链接
        if unique_urls:
            # 移除末尾可能的标点符号
            url = unique_urls[0].rstrip('.,;:!?')
            return url
                
        return None

    # 主流程：下载 -> 转写 -> 生成MD
    def _run_pipeline(self, link: str, user_prompt: str = "", feishu_folder_path: str = None):
        """原始流程方法（向后兼容）"""
        self._run_pipeline_with_cancel(link, user_prompt, feishu_folder_path, None)
    
    def _run_pipeline_with_cancel(self, link: str, user_prompt: str = "", feishu_folder_path: str = None, cancel_event=None):
        """支持取消的主流程"""
        try:
            # 检查是否已取消
            if cancel_event and cancel_event.is_set():
                self.append_log(f"任务已取消：{link}")
                return
            
            # 添加任务到历史记录，更新飞书文件夹路径
            self.add_task_to_history(link, user_prompt, feishu_folder_path)
            
            # 初始化进度条
            self.update_progress(0, "准备开始处理...")
            
            # 检测链接类型
            platform = self._detect_platform(link)
            
            # 如果是小红书图文链接，使用图文分析流程
            if platform == "小红书图文":
                self.append_log(f"检测到小红书图文链接，使用图文分析流程...")
                self._run_xiaohongshu_analysis(link, user_prompt, feishu_folder_path)
                return
            
            # 如果是小红书视频链接，先尝试图文分析流程（内部会检测视频类型并跳转）
            if platform == "小红书视频":
                self.append_log(f"检测到小红书视频链接，使用图文分析流程（内部会检测视频类型）...")
                # 调用图文分析流程，内部会检测是否为视频并跳转
                self._run_xiaohongshu_analysis(link, user_prompt, feishu_folder_path)
                return
            
            # 如果是微信公众号文章链接，使用文章分析流程
            if platform == "微信公众号":
                self.append_log(f"检测到微信公众号文章链接，使用文章分析流程...")
                self._run_wechat_article_analysis(link, user_prompt, feishu_folder_path)
                return
            
            # 否则使用视频处理流程
            platform = "视频"
            
            # 阶段1：下载视频
            self.update_task_status(link, "download", "in_progress")
            self.update_progress(10, "下载视频...")
            video_file = self.download_video(link)
            if not video_file:
                self.append_log("视频下载失败，流程结束。")
                self.update_task_status(link, "download", "failed")
                self.update_progress(0, "失败")
                
                self._call_ops_agent_for_error(
                    link=link,
                    error_message="视频下载失败：未生成本地文件（download_video 返回空）",
                    stage="download",
                    error_type="DownloadFailed",
                    traceback="",
                )
                return
            self.update_task_status(link, "download", "completed", video_file)

            # 阶段2：语音转文字
            self.update_task_status(link, "transcribe", "in_progress")
            self.update_progress(40, "语音转文字(上传/轮询)...")
            result_data = self.speech_to_text(video_file, user_prompt)
            if not result_data:
                self.append_log("语音转文字失败，流程结束。")
                self.update_task_status(link, "transcribe", "failed")
                self.update_progress(0, "失败")
                return
            self.update_task_status(link, "transcribe", "completed", result_data.get("segments", []))

            # 阶段3：AI分析
            self.update_task_status(link, "ai_analysis", "in_progress")
            self.update_progress(60, "使用AI进行文本分析...")
            # AI分析已在speech_to_text中完成
            summary = result_data.get("ai_summary", "")
            self.update_task_status(link, "ai_analysis", "completed", summary)
            
            # 从AI分析结果中提取标题并更新任务标题
            if summary:
                # 从摘要中提取标题
                title = self.extract_title_from_summary(summary, link)
                # 确保标题有效
                if not title or title == "未知标题":
                    # 如果摘要不为空，使用摘要的前20个字符作为标题
                    if summary:
                        title = summary[:20].strip()
                        # 清理标题
                        import re
                        title = re.sub(r'[\\/:*?"<>|]', '', title)
                        title = title.replace(' ', '_')
                        # 确保标题不为空
                        if not title:
                            title = "视频内容摘要"
                        self.append_log(f"使用AI摘要内容生成标题：{title}")
                
                # 更新任务的标题字段
                for task in self.history.get("tasks", []):
                    if task.get("link") == link:
                        # 总是更新标题，确保标题明确
                        old_title = task.get("title", "")
                        task["title"] = title
                        task["updated_at"] = datetime.now().isoformat()
                        save_history(self.history)
                        if old_title != title:
                            self.append_log(f"更新任务标题：{title}")
                        break

            # 阶段4：生成Markdown
            self.update_task_status(link, "generate_md", "in_progress")
            self.update_progress(80, "生成Markdown文档...")
            md_file = self.generate_md(result_data, link, platform)
            if not md_file:
                self.append_log("生成文档失败")
                self.update_task_status(link, "generate_md", "failed")
                self.update_progress(0, "失败")
                return
            self.update_task_status(link, "generate_md", "completed", md_file)
            
            # 清理缓存文件（只有在生成MD文件成功后才清理）
            self.cleanup_task_cache(link, {
                "download": video_file,
                "transcribe": result_data.get("segments", []),
                "ai_analysis": result_data.get("ai_summary", "")
            })

            # 阶段5：上传到飞书
            self._run_feishu_upload_if_enabled(link, md_file, user_prompt, feishu_folder_path)
            
            self.update_progress(100, "完成")
        except Exception as e:
            self.append_log(f"异常：{type(e).__name__}: {e}")
            self._pipeline_log(f"pipeline_run_exc link={str(link)[:80]!r} err={type(e).__name__}: {e!r}")
            import traceback
            error_traceback = traceback.format_exc()
            traceback.print_exc()
            
            # 确保任务状态被更新为失败
            self.update_task_status(link, "download", "failed")
            self.update_progress(0, "失败")
            
            # 调用运维Agent分析错误
            if self.ops_agent:
                try:
                    self.append_log("正在调用运维Agent分析错误...")
                    
                    # 获取任务ID
                    task_id = None
                    for task in self.history.get("tasks", []):
                        if task.get("link") == link:
                            task_id = task.get("id", "unknown")
                            break
                    
                    # 构建错误信息
                    error_info = {
                        "type": type(e).__name__,
                        "message": str(e),
                        "traceback": error_traceback
                    }
                    
                    # 获取日志（从UI日志中）
                    logs = []
                    try:
                        # 尝试从日志文件或UI获取最近日志
                        log_widget = getattr(self, "log", None)
                        if log_widget:
                            logs = log_widget.get("1.0", tk.END).split("\n")
                    except Exception:
                        pass

                    # 调用运维Agent
                    md_path = self.ops_agent.monitor_task_completion(
                        link=link,
                        task_id=task_id or "unknown",
                        status="failed",
                        logs=logs,
                        error_info=error_info
                    )
                    
                    if md_path:
                        self.append_log(f"运维Agent分析完成，维护建议已保存: {md_path}")
                    else:
                        self.append_log("运维Agent分析完成，无需生成维护建议")
                        
                except Exception as ops_e:
                    self.append_log(f"运维Agent调用失败: {ops_e}")
        finally:
            # 确保任务状态被正确更新
            try:
                # 检查任务是否存在于历史记录中
                task_exists = False
                for task in self.history.get("tasks", []):
                    if task.get("link") == link:
                        task_exists = True
                        break
                
                # 如果任务不存在，添加到历史记录
                if not task_exists:
                    self.add_task_to_history(link, user_prompt, feishu_folder_path)
                    # 更新状态为失败
                    self.update_task_status(link, "download", "failed")
            except Exception as e:
                self.append_log(f"历史记录更新异常：{e}")
            
            # 只有在非队列处理时才启用按钮
            if not self.processing_queue:
                self.start_btn.config(state=tk.NORMAL)

    def _call_ops_agent_for_error(self, link: str, error_message: str, stage: str, 
                                   error_type: str = "Unknown", traceback: str = ""):
        """
        调用运维Agent分析错误
        
        Args:
            link: 任务链接
            error_message: 错误消息
            stage: 失败的阶段（download/transcribe/ai_analysis/generate_md）
            error_type: 错误类型
            traceback: 堆栈跟踪
        """
        if not self.ops_agent:
            self.append_log("[运维Agent] 运维Agent未初始化，跳过错误分析")
            return
        
        try:
            self.append_log("[运维Agent] 正在调用运维Agent分析错误...")
            self.append_log(f"[运维Agent] 错误阶段: {stage}")
            self.append_log(f"[运维Agent] 错误类型: {error_type}")
            self.append_log(f"[运维Agent] 错误消息: {error_message[:100]}...")
            
            # 获取任务ID
            task_id = None
            for task in self.history.get("tasks", []):
                if task.get("link") == link:
                    task_id = task.get("id", "unknown")
                    break
            
            if not task_id:
                task_id = f"task_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
                self.append_log(f"[运维Agent] 未找到任务ID，使用临时ID: {task_id}")
            
            # 构建错误信息
            error_info = {
                "type": error_type,
                "message": error_message,
                "traceback": traceback
            }
            
            # 获取日志（从UI日志中）
            logs = []
            try:
                log_widget = getattr(self, "log", None)
                if log_widget:
                    log_content = log_widget.get("1.0", tk.END)
                    logs = log_content.split("\n")
                    self.append_log(f"[运维Agent] 已获取 {len(logs)} 行日志用于分析")
            except Exception as log_e:
                self.append_log(f"[运维Agent] 获取日志失败: {log_e}")
            
            # 调用运维Agent
            self.append_log("[运维Agent] 开始分析错误...")
            md_path = self.ops_agent.monitor_task_completion(
                link=link,
                task_id=task_id,
                status="failed",
                logs=logs,
                error_info=error_info
            )
            
            if md_path:
                self.append_log(f"[运维Agent] ✓ 分析完成，维护建议已保存: {md_path}")
                self.append_log(f"[运维Agent] 请查看维护文件了解修复建议")
            else:
                self.append_log("[运维Agent] ✓ 分析完成，无需生成维护建议")
                
        except Exception as ops_e:
            self.append_log(f"[运维Agent] ✗ 调用失败: {ops_e}")
            import traceback
            self.append_log(f"[运维Agent] 错误详情: {traceback.format_exc()}")

    # 抖音视频专用下载方法
    def download_douyin_video(self, link: str):
        """使用HTML解析方法下载抖音视频（免登录）"""
        import time
        import requests
        import re
        import json
        
        download_start = time.time()
        
        try:
            self.append_log("使用抖音专用解析器下载视频...", "INFO")
            
            # 清理链接
            link = link.strip('`')
            
            # 获取当前目录下的视频文件数量，作为总序号
            existing_videos = [f for f in os.listdir(VIDEO_DIR) if f.endswith('.mp4')]
            total_count = len(existing_videos) + 1
            
            # 获取当前日期（月-日）
            current_date = time.strftime('%m-%d')
            
            # 从链接中提取文档名称
            doc_name_match = re.search(r'\d+', link.split('/')[-1])
            doc_name = doc_name_match.group(0) if doc_name_match else "douyin"
            
            # 构建新的文件名：总记录序号-月-日-文档名称
            new_filename = f"{total_count:03d}-{current_date}-{doc_name}.mp4"
            output_file = os.path.join(VIDEO_DIR, new_filename)
            
            headers = {
                'User-Agent': 'Mozilla/5.0 (iPhone; CPU iPhone OS 17_2 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) EdgiOS/121.0.2277.107 Version/17.0 Mobile/15E148 Safari/604.1'
            }
            
            # 步骤1: 访问分享链接获取视频ID
            self.append_log("解析抖音视频链接...", "INFO")
            response = requests.get(link, headers=headers, allow_redirects=True, timeout=30)
            video_id = response.url.split("?")[0].strip("/").split("/")[-1]
            self.append_log(f"视频ID: {video_id}", "INFO")
            
            # 步骤2: 访问分享页面获取HTML
            share_url = f'https://www.iesdouyin.com/share/video/{video_id}'
            response = requests.get(share_url, headers=headers, timeout=30)
            response.raise_for_status()
            
            # 步骤3: 从HTML中解析视频信息
            pattern = re.compile(
                pattern=r"window\._ROUTER_DATA\s*=\s*(.*?)</script>",
                flags=re.DOTALL
            )
            find_res = pattern.search(response.text)
            
            if not find_res or not find_res.group(1):
                self.append_log("未能从HTML中解析视频信息", "ERROR")
                return None
            
            json_data = json.loads(find_res.group(1).strip())
            
            # 步骤4: 提取视频URL
            video_url = None
            if "loaderData" in json_data:
                loader_data = json_data["loaderData"]
                for key in loader_data:
                    if "videoInfoRes" in str(loader_data[key]):
                        data = loader_data[key]["videoInfoRes"]["item_list"][0]
                        video_url = data["video"]["play_addr"]["url_list"][0].replace("playwm", "play")
                        self.append_log("成功获取无水印视频链接", "INFO")
                        break
            
            if not video_url:
                self.append_log("未能提取视频URL", "ERROR")
                return None
            
            # 步骤5: 下载视频
            self.append_log("下载视频中...", "INFO")
            video_response = requests.get(video_url, headers=headers, stream=True, timeout=120)
            video_response.raise_for_status()
            
            with open(output_file, 'wb') as f:
                for chunk in video_response.iter_content(chunk_size=8192):
                    if chunk:
                        f.write(chunk)
            
            # 验证文件
            if os.path.exists(output_file) and os.path.getsize(output_file) > 0:
                download_end = time.time()
                self.append_log(f"视频下载成功: {output_file}", "INFO")
                self.append_log(f"视频下载耗时: {download_end - download_start:.2f}秒", "INFO")
                
                # 添加到缓存
                with self.video_cache_lock:
                    self.video_cache[link] = output_file
                
                return output_file
            else:
                self.append_log("视频文件不存在或为空", "ERROR")
            
            return None
            
        except Exception as e:
            download_end = time.time()
            self.append_log(f"抖音视频下载异常: {e}", "ERROR")
            self.append_log(f"视频下载耗时: {download_end - download_start:.2f}秒（异常）", "INFO")
            return None

    # 步骤 1：下载视频
    def download_video(self, link: str):
        import time
        import os
        import subprocess
        import re
        import random
        download_start = time.time()
        
        try:
            # 清理链接中的反引号
            link = link.strip('`')
            
            # 对于抖音链接，优先使用专用解析器
            if "douyin.com" in link or "tiktok.com" in link or "v.douyin.com" in link:
                self.append_log("检测到抖音链接，使用专用解析器...", "INFO")
                result = self.download_douyin_video(link)
                if result:
                    return result
                else:
                    self.append_log("专用解析器失败，尝试使用yt-dlp...", "WARNING")
            
            # 检查视频缓存
            with self.video_cache_lock:
                if link in self.video_cache:
                    cached_file = self.video_cache[link]
                    with self.file_operation_lock:
                        if os.path.exists(cached_file) and os.path.getsize(cached_file) > 0:
                            self.append_log(f"使用缓存的视频文件: {cached_file}", "INFO")
                            download_end = time.time()
                            self.append_log(f"视频下载耗时: {download_end - download_start:.2f}秒（使用缓存）", "INFO")
                            return cached_file
                        else:
                            # 缓存文件不存在或为空，删除缓存条目
                            del self.video_cache[link]
                            self.append_log("缓存视频文件不存在，重新下载", "INFO")
            
            # 使用yt-dlp工具下载视频
            import subprocess
            import os
            
            self.append_log("使用yt-dlp下载视频...", "INFO")
            
            # 获取当前日期（月-日）
            current_date = time.strftime('%m-%d')
            
            # 从链接中提取文档名称（使用链接的最后部分）
            import re
            doc_name_match = re.search(r'\d+', link.split('/')[-1])
            doc_name = doc_name_match.group(0) if doc_name_match else "unknown"
            
            # 构建唯一文件名：时间戳-随机数-文档名称
            import random
            timestamp = int(time.time() * 1000)
            random_suffix = random.randint(100, 999)
            new_filename = f"{timestamp}-{random_suffix}-{current_date}-{doc_name}.mp4"
            output_file = os.path.join(VIDEO_DIR, new_filename)
            
            # 从链接中提取 xsec_token
            xsec_token_match = re.search(r'xsec_token=([^&]+)', link)
            xsec_token = xsec_token_match.group(1) if xsec_token_match else ""
            
            # 构建cookie字符串
            cookie_string = f"xsec_token={xsec_token}" if xsec_token else ""
            
            # 根据链接类型设置不同的referer
            referer = "https://www.xiaohongshu.com/"
            if "bilibili.com" in link or "bilibili" in link:
                referer = "https://www.bilibili.com/"
            elif "youtube.com" in link or "youtu.be" in link:
                referer = "https://www.youtube.com/"
            elif "douyin.com" in link or "tiktok.com" in link:
                referer = "https://www.douyin.com/"
                # 对于抖音链接，添加额外的参数以提高下载成功率
                self.append_log("检测到抖音链接，使用专用配置...", "INFO")
            
            # 构建yt-dlp命令，直接下载到目标文件夹
            cmd = [
                "yt-dlp",
                "--user-agent", "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/144.0.0.0 Safari/537.36",
                "--referer", referer,
                "--no-check-certificate",
                "--quiet",  # 静默模式，减少输出
                "--no-warnings",  # 禁用警告
                # 下载优化参数 - B站音视频分离格式，使用360p降低下载时间
                "--format", "30016+30216/worst",  # 优先360p视频+音频，否则最低质量
                "--merge-output-format", "mp4",  # 合并为mp4格式
                "--concurrent-fragments", "5",  # 5个并发片段下载
                "--buffer-size", "16K",  # 缓冲区大小
                "--retries", "10",  # 重试次数
                "--fragment-retries", "10",  # 片段重试次数
                "--socket-timeout", "30",  #  socket超时
                "-o", output_file,
            ]
            
            # 对于B站链接，尝试使用浏览器cookies提高下载速度
            if "bilibili.com" in link:
                self.append_log("检测到B站链接，尝试使用浏览器cookies...", "INFO")
                try:
                    # 尝试从Firefox获取cookies
                    cmd.extend(["--cookies-from-browser", "firefox"])
                    self.append_log("已添加Firefox cookies参数", "INFO")
                except Exception as e:
                    self.append_log(f"从Firefox获取cookies失败: {e}", "WARNING")
                    self.append_log("提示：在Firefox浏览器中登录B站可以提高下载速度", "INFO")
            
            # 对于抖音链接，添加额外的参数以提高下载成功率
            if "douyin.com" in link or "tiktok.com" in link:
                # 尝试多种方式获取cookies
                cookie_added = False
                
                # 方式1: 检查是否存在抖音cookie文件
                cookie_file = os.path.join(BASE_DIR, "douyin_cookies.txt")
                if os.path.exists(cookie_file):
                    self.append_log(f"使用抖音cookie文件：{cookie_file}", "INFO")
                    cmd.extend(["--cookies", cookie_file])
                    cookie_added = True
                
                # 方式2: 尝试从Firefox获取cookies（优先使用Firefox，因为它不会像Chrome那样被锁定）
                if not cookie_added:
                    try:
                        self.append_log("尝试从Firefox浏览器获取cookies...", "INFO")
                        # 直接添加Firefox cookies参数
                        cmd.extend(["--cookies-from-browser", "firefox"])
                        cookie_added = True
                        self.append_log("已添加Firefox cookies参数", "INFO")
                    except Exception as e:
                        self.append_log(f"从Firefox获取cookies失败: {str(e)[:30]}", "WARNING")
                
                # 方式3: 如果Firefox失败，尝试Edge
                if not cookie_added:
                    try:
                        self.append_log("尝试从Edge浏览器获取cookies...", "INFO")
                        cmd.extend(["--cookies-from-browser", "edge"])
                        cookie_added = True
                        self.append_log("已添加Edge cookies参数", "INFO")
                    except Exception as e:
                        self.append_log(f"从Edge获取cookies失败: {str(e)[:30]}", "WARNING")
                
                if not cookie_added:
                    self.append_log("警告：无法获取抖音cookies，下载可能失败", "WARNING")
                    self.append_log("解决方案：在Firefox浏览器中登录抖音后重试", "WARNING")
                
                # 添加抖音专用参数（使用兼容的参数格式）
                cmd.extend([
                    "--user-agent", "Mozilla/5.0 (iPhone; CPU iPhone OS 16_0 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.0 Mobile/15E148 Safari/604.1",
                    "--max-downloads", "1",
                    "--no-check-certificate",
                    "--ignore-errors",
                    link
                ])
            else:
                cmd.append(link)
            
            # 执行命令（增加超时时间，添加重试机制）
            max_retries = 2
            retry_count = 0
            result = None
            
            while retry_count < max_retries:
                try:
                    self.append_log(f"执行yt-dlp命令（尝试 {retry_count+1}/{max_retries}）...", "INFO")
                    self.append_log(f"yt-dlp命令: {' '.join(cmd[:12])}...", "DEBUG")  # 记录命令
                    # B站下载限速严重，增加超时时间到600秒（10分钟）
                    result = subprocess.run(cmd, capture_output=True, text=True, timeout=600)
                    break
                except subprocess.TimeoutExpired as te:
                    retry_count += 1
                    self.append_log(f"yt-dlp执行超时（{te.timeout}秒），正在重试...（{retry_count}/{max_retries}）", "WARNING")
                    self.append_log(f"超时详情: B站可能限速，建议登录B站账号或使用更低清晰度", "INFO")
                    if retry_count >= max_retries:
                        self.append_log("yt-dlp执行多次超时，B站下载限速严重", "ERROR")
                        self.append_log("解决方案：1.在Firefox登录B站 2.使用其他视频源 3.降低视频清晰度", "ERROR")
                        download_end = time.time()
                        self.append_log(f"视频下载耗时: {download_end - download_start:.2f}秒（超时）", "INFO")
                        return None
                except Exception as e:
                    self.append_log(f"yt-dlp执行异常：{type(e).__name__}: {e}", "ERROR")
                    import traceback
                    self.append_log(f"异常详情: {traceback.format_exc()}", "DEBUG")
                    retry_count += 1
                    if retry_count >= max_retries:
                        self.append_log("yt-dlp执行失败", "ERROR")
                        download_end = time.time()
                        self.append_log(f"视频下载耗时: {download_end - download_start:.2f}秒（失败）", "INFO")
                        return None
            
            if result.returncode == 0:
                # 检查视频文件是否存在且大小大于0
                if os.path.exists(output_file) and os.path.getsize(output_file) > 0:
                    self.append_log(f"视频下载成功: {output_file}", "INFO")
                    # 添加到缓存
                    with self.video_cache_lock:
                        self.video_cache[link] = output_file
                    download_end = time.time()
                    self.append_log(f"视频下载耗时: {download_end - download_start:.2f}秒", "INFO")
                    return output_file
                else:
                    self.append_log("视频文件不存在或为空", "ERROR")
            else:
                # 优化：只显示部分错误信息
                error_msg = result.stderr[:500] + "..." if len(result.stderr) > 500 else result.stderr
                self.append_log(f"yt-dlp执行失败: {error_msg}", "ERROR")
                
                # 特殊处理抖音链接
                if "douyin.com" in link:
                    self.append_log("抖音视频下载失败，解决方案：", "ERROR")
                    self.append_log("1. 确保使用的是直接视频链接（非用户页面或收藏页面）", "ERROR")
                    self.append_log("2. 在Firefox浏览器中登录抖音网站（www.douyin.com）", "ERROR")
                    self.append_log("3. 登录后重新运行本程序，程序会自动使用Firefox的cookies", "ERROR")
                    self.append_log("4. 如果仍失败，请关闭Chrome浏览器后重试", "ERROR")
            
            # 执行失败，直接返回None表示失败，不使用示例文件
            self.append_log("视频下载失败", "ERROR")
            download_end = time.time()
            self.append_log(f"视频下载耗时: {download_end - download_start:.2f}秒（失败）", "INFO")
            return None
            
        except Exception as e:
            download_end = time.time()
            self.append_log(f"下载异常：{e}", "ERROR")
            self.append_log(f"视频下载耗时: {download_end - download_start:.2f}秒（异常）", "INFO")
            # 直接返回None表示失败，不使用示例文件
            return None

    # 步骤2：保存视频
    def save_video(self, video_url: str, link: str):
        try:
            # 检查video_url是否已经是本地文件路径
            if os.path.exists(video_url) and os.path.isfile(video_url):
                # 如果是本地文件，直接复制到videos目录
                url_hash = hashlib.md5(link.encode()).hexdigest()[:8]
                ts = int(time.time())
                file_path = os.path.join(VIDEO_DIR, f"video_{url_hash}_{ts}.mp4")
                
                # 复制文件
                import shutil
                shutil.copy2(video_url, file_path)
                
                self.append_log(f"视频已从临时位置复制到: {file_path}", "INFO")
                
                # 清理临时目录
                temp_dir = os.path.dirname(video_url)
                if 'temp' in temp_dir.lower():
                    try:
                        shutil.rmtree(temp_dir)
                        self.append_log(f"已清理临时目录: {temp_dir}", "DEBUG")
                    except:
                        pass
                
                return file_path
            else:
                # 如果是URL，使用异步方式下载视频
                url_hash = hashlib.md5(link.encode()).hexdigest()[:8]
                ts = int(time.time())
                file_path = os.path.join(VIDEO_DIR, f"video_{url_hash}_{ts}.mp4")
                self.append_log(f"保存至：{file_path}")
                
                # 运行异步下载
                try:
                    # 创建事件循环并运行异步任务
                    loop = asyncio.new_event_loop()
                    asyncio.set_event_loop(loop)
                    result = loop.run_until_complete(self.async_download_video(video_url, file_path))
                    loop.close()
                    
                    if result:
                        self.append_log(f"视频下载完成: {file_path}", "INFO")
                        return file_path
                    else:
                        self.append_log("异步下载视频失败", "ERROR")
                        return None
                except Exception as e:
                    self.append_log(f"异步下载异常：{e}", "ERROR")
                    # 回退到同步下载
                    return self.sync_download_video(video_url, file_path)
        except Exception as e:
            self.append_log(f"保存异常：{e}")
            return None
    
    # 异步下载视频
    async def async_download_video(self, video_url: str, file_path: str):
        """异步下载视频文件"""
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/144.0.0.0 Safari/537.36',
            'Accept-Encoding': 'identity;q=1, *;q=0',
            'Range': 'bytes=0-',
            'Referer': 'https://www.hellotik.app/',
            'Sec-Ch-Ua': '"Not(A:Brand";v="8", "Chromium";v="144", "Google Chrome";v="144")',
            'Sec-Ch-Ua-Mobile': '?0',
            'Sec-Ch-Ua-Platform': '"Windows"'
        }
        
        self.append_log(f"使用异步方式下载视频: {video_url}", "DEBUG")
        
        try:
            async with aiohttp.ClientSession() as session:
                async with session.get(video_url, headers=headers, timeout=aiohttp.ClientTimeout(total=60)) as response:
                    self.append_log(f"下载响应状态码: {response.status}", "DEBUG")
                    
                    if response.status != 200:
                        self.append_log(f"下载失败，状态码: {response.status}", "ERROR")
                        return False
                    
                    content_length = response.headers.get('Content-Length')
                    self.append_log(f"Content-Length: {content_length}", "DEBUG")
                    
                    # 异步写入文件
                    with open(file_path, 'wb') as f:
                        async for chunk in response.content.iter_chunked(8192):
                            if chunk:
                                f.write(chunk)
                    
            return True
        except Exception as e:
            self.append_log(f"异步下载异常：{e}", "ERROR")
            return False
    
    # 同步下载视频（作为回退方案）
    def sync_download_video(self, video_url: str, file_path: str):
        """同步下载视频文件（作为回退方案）"""
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/144.0.0.0 Safari/537.36',
            'Accept-Encoding': 'identity;q=1, *;q=0',
            'Range': 'bytes=0-',
            'Referer': 'https://www.hellotik.app/',
            'Sec-Ch-Ua': '"Not(A:Brand";v="8", "Chromium";v="144", "Google Chrome";v="144")',
            'Sec-Ch-Ua-Mobile': '?0',
            'Sec-Ch-Ua-Platform': '"Windows"'
        }
        
        self.append_log(f"使用同步方式下载视频: {video_url}", "DEBUG")
        
        try:
            r = requests.get(video_url, stream=True, headers=headers, timeout=60)
            self.append_log(f"下载响应状态码: {r.status_code}", "DEBUG")
            
            r.raise_for_status()
            
            with open(file_path, 'wb') as f:
                for chunk in r.iter_content(chunk_size=8192):
                    if chunk:
                        f.write(chunk)
            
            self.append_log(f"同步下载视频完成: {file_path}", "INFO")
            return file_path
        except Exception as e:
            self.append_log(f"同步下载异常：{e}", "ERROR")
            return None

    def _effective_whisper_pool_size(self) -> int:
        raw = os.environ.get("WHISPER_POOL_SIZE", "").strip()
        if raw.isdigit():
            n = int(raw)
        else:
            n = min(4, max(1, self.max_workers))
        return max(1, min(n, 8))

    def _ensure_whisper_pool(self) -> None:
        """懒加载：创建 N 份独立的 Whisper 模型，放入 Queue；转写时 get/put 实现借还，支持真并行。"""
        with self._whisper_pool_init_lock:
            if self._whisper_pool_queue is not None:
                return
            import whisper

            n = self._effective_whisper_pool_size()
            self.append_log(
                f"初始化 Whisper 实例池：{n} 路独立模型（环境变量 WHISPER_POOL_SIZE 可改；默认与线程池协调）",
                "INFO",
            )
            q: queue.Queue = queue.Queue(maxsize=n)
            name = self._whisper_pool_model_name
            for i in range(n):
                self.append_log(f"Whisper 池加载 {i + 1}/{n}（{name}）...", "INFO")
                try:
                    m = whisper.load_model(name)
                except Exception as e:
                    self.append_log(f"加载 {name} 失败，回退 small：{e}", "WARNING")
                    name = "small"
                    self._whisper_pool_model_name = "small"
                    m = whisper.load_model("small")
                q.put((f"w{i}", m))
            self._whisper_pool_queue = q
            with self.model_cache_lock:
                self.model_cache = None

    def _acquire_whisper_pool_slot(self, timeout: float = 7200.0):
        self._ensure_whisper_pool()
        return self._whisper_pool_queue.get(timeout=timeout)

    def _release_whisper_pool_slot(self, slot) -> None:
        if self._whisper_pool_queue is not None:
            self._whisper_pool_queue.put_nowait(slot)

    def _audio_duration_probe_sec(self, media_path: str):
        """ffprobe 检测首条音轨时长；无音轨返回 0.0；失败返回 None。"""
        import subprocess
        import shutil

        if not shutil.which("ffprobe"):
            return None
        try:
            r = subprocess.run(
                [
                    "ffprobe",
                    "-v",
                    "error",
                    "-select_streams",
                    "a:0",
                    "-show_entries",
                    "stream=duration",
                    "-of",
                    "default=noprint_wrappers=1:nokey=1",
                    media_path,
                ],
                capture_output=True,
                text=True,
                timeout=120,
            )
            out = (r.stdout or "").strip()
            if r.returncode != 0 or not out:
                return 0.0
            return float(out.split()[0])
        except Exception:
            return None

    # 步骤 3：语音转文字
    def speech_to_text(self, video_file: str, user_prompt: str = ""):
        try:
            import time
            start_time = time.time()
            
            # 检查视频文件是否存在
            if not os.path.exists(video_file):
                self.append_log(f"视频文件不存在: {video_file}", "ERROR")
                # 返回模拟数据
                return {
                    "segments": [
                        {"start_time": 0, "text": "视频文件不存在，这是模拟数据。"},
                        {"start_time": 10, "text": "请检查视频下载是否成功。"}
                    ],
                    "ai_summary": "视频文件不存在，无法转写。"
                }
            
            # 检查视频文件大小
            file_size = os.path.getsize(video_file)
            if file_size < 1024:  # 如果文件小于1KB，可能是空文件或示例文件
                self.append_log(f"检测到小文件（{file_size} bytes），可能是示例视频文件，使用模拟数据...", "WARNING")
                # 直接返回模拟数据
                return {
                    "segments": [
                        {"start_time": 0, "text": "这是一段模拟的视频转文字结果。"},
                        {"start_time": 10, "text": "视频内容包括产品介绍、使用方法和注意事项。"},
                        {"start_time": 20, "text": "这是一个示例文本，用于演示语音转文字功能。"}
                    ],
                    "ai_summary": "视频主要介绍了产品的基本信息、使用步骤和注意事项，帮助用户快速了解产品的核心功能和使用方法。"
                }
            

            
            # Whisper 实例池：多份独立模型 + Queue 借还，实现与线程池匹配的并行转写
            self.append_log("使用 Whisper 本地模型进行语音转文字...", "INFO")

            import shutil

            _fb = resolve_ffmpeg_bin_dir()
            if _fb:
                os.environ["PATH"] = _fb + os.pathsep + os.environ.get("PATH", "")
            if not shutil.which("ffmpeg"):
                self.append_log("警告：未找到 ffmpeg，Whisper 可能无法处理视频文件", "WARNING")

            dur = self._audio_duration_probe_sec(video_file)
            if dur is not None and dur < 0.05:
                self.append_log(
                    f"ffprobe：无有效音轨或时长过短（{dur}s），跳过转写以避免 reshape 错误", "ERROR"
                )
                return None

            max_retries = max(1, int(os.environ.get("WHISPER_TRANSCRIBE_RETRIES", "3")))
            base_delay = float(os.environ.get("WHISPER_RETRY_DELAY_SEC", "2"))

            slot = None
            transcribe_done = False
            try:
                slot = self._acquire_whisper_pool_slot()
                slot_id, model = slot
                self.append_log(
                    f"从 Whisper 池取得槽位 {slot_id}（池大小={self._effective_whisper_pool_size()}，可并行）",
                    "INFO",
                )
                self.update_progress(45, "加载语音转文字模型...")
                self.update_progress(55, "正在分析视频音频...")
                self.append_log("开始转写...", "INFO")

                def progress_updater():
                    progress = 60
                    while not transcribe_done:
                        if progress < 75:
                            progress += 1
                            self.update_progress(progress, f"正在转写音频... {progress - 55}%")
                        time.sleep(1)

                threading.Thread(target=progress_updater, daemon=True).start()

                result = None
                last_exc = None
                for attempt in range(max_retries):
                    try:
                        self.append_log(
                            f"使用优化参数进行转写…（尝试 {attempt + 1}/{max_retries}）", "INFO"
                        )
                        t0 = time.time()
                        result = model.transcribe(
                            video_file,
                            language="zh",
                            fp16=False,
                            verbose=False,
                            task="transcribe",
                            beam_size=1,
                            temperature=0.0,
                            best_of=1,
                            patience=0.0,
                            initial_prompt="请使用标准简体中文进行转写，保持语句通顺，不要遗漏任何内容。",
                            condition_on_previous_text=False,
                            compression_ratio_threshold=2.4,
                        )
                        self.append_log(f"转写耗时: {time.time() - t0:.2f}秒", "INFO")
                        last_exc = None
                        break
                    except RuntimeError as e:
                        last_exc = e
                        es = str(e)
                        recoverable = "cannot reshape tensor" in es or "0 elements" in es
                        if recoverable and attempt < max_retries - 1:
                            delay = base_delay * (2**attempt)
                            self.append_log(
                                f"Whisper 转写异常，{delay:.0f}s 后重试（隔时重试 {attempt + 1}/{max_retries}）：{e}",
                                "ERROR",
                            )
                            time.sleep(delay)
                        else:
                            raise
                transcribe_done = True

                if result is None:
                    self.append_log(f"Whisper 转写失败（已重试 {max_retries} 次）: {last_exc}", "ERROR")
                    return None

                self.update_progress(75, "转写完成，正在处理结果...")
                text = result["text"]
                segments = []
                for seg in result["segments"]:
                    segments.append({"start_time": seg["start"], "text": seg["text"].strip()})
                self.append_log("语音转文字完成！", "INFO")
                self.append_log(f"转写结果: {text[:100]}...", "INFO")

                self.update_progress(80, "使用AI进行文本总结...")
                self.append_log("使用火山引擎API进行文本总结...", "INFO")
                summary = self.summarize_with_volcengine(text, user_prompt)
                self.update_progress(85, "总结完成，准备生成文档...")

                if summary:
                    self.append_log("文本总结成功", "INFO")
                    return {"segments": segments, "ai_summary": summary}
                return {
                    "segments": segments,
                    "ai_summary": text[:100] + "...（省略部分内容）",
                }
            except queue.Empty:
                self.append_log("等待 Whisper 实例池槽位超时", "ERROR")
                return None
            except FileNotFoundError as e:
                self.append_log(
                    "语音转文字失败：找不到 ffmpeg/ffprobe 或临时文件，请检查 PATH", "ERROR"
                )
                self.append_log(f"错误详情：{e}", "ERROR")
                return None
            finally:
                transcribe_done = True
                if slot is not None:
                    try:
                        self._release_whisper_pool_slot(slot)
                    except Exception:
                        pass

        except Exception as e:
            self.append_log(f"语音转文字异常：{type(e).__name__}: {e}", "ERROR")
            return None
    
    # 使用火山引擎 API 进行文本总结（已停用，改用本地处理）
    def summarize_with_volcengine(self, text: str, user_prompt: str = ""):
        try:
            api_key = CONFIG.get("volcengine_api_key", VOLCENGINE_API_KEY)
            
            # 如果API密钥为空，使用本地简单总结
            if not api_key or api_key == "":
                self.append_log("API密钥未配置，使用本地简单总结...", "INFO")
                return self._local_summary(text)
            
            summary_prompt = CONFIG.get("summary_prompt", DEFAULT_CONFIG["summary_prompt"])
            system_prompt = CONFIG.get("system_prompt", DEFAULT_CONFIG["system_prompt"])
            rules = CONFIG.get("rules", DEFAULT_CONFIG["rules"])
            
            # 确保 volcenginesdkarkruntime 已安装
            try:
                from volcenginesdkarkruntime import Ark
            except ImportError:
                self.append_log("正在安装 volcengine-python-sdk[ark]...", "INFO")
                import subprocess
                subprocess.run(["pip", "install", "--upgrade", "volcengine-python-sdk[ark]"], check=True)
                from volcenginesdkarkruntime import Ark
            
            # 构建请求输入（chat.completions格式）
            messages = [
                {
                    "role": "system",
                    "content": system_prompt
                },
                {
                    "role": "user",
                    "content": f"分析规则：\n{rules}\n\n{summary_prompt.format(text=text)}"
                }
            ]
            
            # 如果有 user_prompt，添加到输入中
            if user_prompt:
                messages.append({
                    "role": "user",
                    "content": user_prompt
                })
            
            # 主接入点 + 备用接入点（均为 Doubao-Seed-2.0-mini）
            primary = CONFIG.get("ai_chat_model", AI_CHAT_MODEL)
            backup = CONFIG.get("ai_chat_model_backup", AI_CHAT_MODEL_BACKUP)
            client = Ark(
                base_url=AI_CHAT_API_URL,
                api_key=AI_CHAT_API_KEY,
                timeout=60.0,
            )
            max_retries = 3
            for attempt in range(max_retries):
                self.append_log(f"调用火山引擎API进行总结... (尝试 {attempt + 1}/{max_retries})", "INFO")
                last_err = None
                primary_failed_detail = None
                _tried = set()
                for mid, label in ((primary, "主"), (backup, "备")):
                    if not mid or mid in _tried:
                        continue
                    _tried.add(mid)
                    try:
                        self.append_log(f"  使用{label}接入点: {mid}", "INFO")
                        response = client.chat.completions.create(
                            model=mid,
                            messages=messages,
                            timeout=60.0,
                        )
                        if response.choices and len(response.choices) > 0:
                            summary = response.choices[0].message.content
                            if summary:
                                self.append_log(f"火山引擎API调用成功（{label}接入点）", "INFO")
                                if label == "备" and primary_failed_detail is not None:
                                    threading.Thread(
                                        target=self._schedule_ops_volcengine_degraded,
                                        args=(primary_failed_detail, primary, mid),
                                        daemon=True,
                                    ).start()
                                return summary
                        self.append_log("火山引擎API返回空结果或格式不正确", "ERROR")
                    except Exception as e:
                        last_err = e
                        et = type(e).__name__
                        self.append_log(f"  [{label}] 接入点失败 [{et}]: {e}", "ERROR")
                        if label == "主":
                            primary_failed_detail = f"{et}: {e}"
                if attempt < max_retries - 1:
                    import time
                    w = 2 ** attempt
                    self.append_log(f"主备均失败，{w}秒后重试...", "WARNING")
                    time.sleep(w)
            if last_err:
                self.append_log(f"火山引擎总结最终失败: {last_err}", "ERROR")
                threading.Thread(
                    target=self._ops_dispatch_log_incident,
                    args=(f"火山引擎总结最终失败: {last_err}", "ERROR"),
                    daemon=True,
                ).start()
            return None
        except Exception as e:
            self.append_log(f"火山引擎 API 调用异常：{e}", "ERROR")
            return None
    
    def _local_summary(self, text: str) -> str:
        """本地简单总结（当API不可用时使用）"""
        try:
            # 提取前3个句子作为总结
            sentences = text.split('。')[:3]
            summary = '。'.join(sentences) + '。'
            
            # 如果文本太长，截断
            if len(summary) > 500:
                summary = summary[:500] + '...'
            
            # 添加简单标题（取第一句的前20字）
            title = sentences[0][:20] if sentences else "内容总结"
            
            result = f"{title}\n\n总结：\n{summary}\n\n[注意：此为本地简单总结，未使用AI分析]"
            self.append_log("本地简单总结完成", "INFO")
            return result
        except Exception as e:
            self.append_log(f"本地总结失败：{e}", "ERROR")
            return f"内容总结\n\n{text[:500]}...\n\n[注意：总结生成失败]"

    # 步骤 3.5：处理已下载的视频（小红书视频专用）
    def _process_downloaded_video(self, video_file: str, link: str, user_prompt: str = "", feishu_folder_path: str = None):
        """处理已下载的视频（转写、AI 分析、生成 MD）"""
        try:
            # 阶段 2：语音转文字
            self.update_task_status(link, "transcribe", "in_progress")
            self.update_progress(40, "语音转文字 (上传/轮询)...")
            result_data = self.speech_to_text(video_file, user_prompt)
            if not result_data:
                self.append_log("语音转文字失败，流程结束。")
                self.update_task_status(link, "transcribe", "failed")
                self.update_progress(0, "失败")
                return
            self.update_task_status(link, "transcribe", "completed", result_data.get("segments", []))

            # 阶段 3：AI 分析
            self.update_task_status(link, "ai_analysis", "in_progress")
            self.update_progress(60, "使用 AI 进行文本分析...")
            summary = result_data.get("ai_summary", "")
            self.update_task_status(link, "ai_analysis", "completed", summary)
            
            # 阶段 4：生成 Markdown
            self.update_task_status(link, "generate_md", "in_progress")
            self.update_progress(80, "生成 Markdown 文档...")
            md_file = self.generate_md(result_data, link, "视频")
            if md_file:
                self.append_log(f"✓ Markdown 文档生成成功：{md_file}")
                self.update_task_status(link, "generate_md", "completed", md_file)
                # 与主流程一致：小红书下载视频后走本函数，此前未调用飞书上传
                self._run_feishu_upload_if_enabled(link, md_file, user_prompt, feishu_folder_path)
                self.update_progress(100, "处理完成！")
            else:
                self.append_log("Markdown 文档生成失败")
                self.update_task_status(link, "generate_md", "failed")
                self.update_progress(0, "失败")
        except Exception as e:
            self.append_log(f"视频处理异常：{e}", "ERROR")
            self.update_task_status(link, "generate_md", "failed")
            self.update_progress(0, "失败")

    # 步骤 4：生成 Markdown
    def generate_md(self, result_data: dict, link: str, platform: str):
        try:
            # 获取当前目录下的Markdown文件数量，作为总序号
            existing_md = [f for f in os.listdir(OUTPUT_DIR) if f.endswith('.md')]
            total_count = len(existing_md) + 1
            
            # 获取当前日期（月-日）
            current_date = time.strftime('%m-%d')
            
            # 从内容中提取标题（使用AI分析结果）
            summary = result_data.get('ai_summary', '')
            doc_name = self.extract_title_from_summary(summary, link)
            
            # 计算URL哈希（用于视频信息中显示）
            url_hash = hashlib.md5(link.encode()).hexdigest()[:8]
            
            # 判断是视频还是图文
            is_xiaohongshu_image = platform == "小红书" and result_data.get('image_analysis') is not None
            is_wechat_article = platform == "微信公众号" and result_data.get('image_analysis') is not None
            
            if is_xiaohongshu_image or is_wechat_article:
                # 图文格式（小红书或微信公众号）
                md_path = os.path.join(OUTPUT_DIR, f"{total_count:03d}-{current_date}-{doc_name}_内容分析.md")
                
                # 构建原始内容
                original_content = result_data.get('transcript', '')
                
                # 添加图片统计
                expected_count = result_data.get('expected_image_count', 0)
                actual_count = result_data.get('actual_image_count', 0)
                image_stats = ""
                if expected_count > 0:
                    image_stats = f"\n## 图片统计\n- 应有图片数: {expected_count}\n- 实际提取: {actual_count}\n- 状态: {'全部提取成功' if actual_count >= expected_count else '部分图片可能未成功提取'}\n\n"
                
                # 添加每张图片的OCR文本（仅当原始内容中未包含图片内容时）
                image_analysis = result_data.get('image_analysis', [])
                image_ocr_content = ""
                # 检查原始内容是否已经包含图片内容，避免重复
                if image_analysis and "## 图片内容" not in original_content:
                    image_ocr_content = "\n## 图片内容识别\n\n"
                    for i, img_data in enumerate(image_analysis, 1):
                        img_url = img_data.get('url', '')
                        img_text = img_data.get('text', '')
                        img_index = img_data.get('index', i)
                        if img_text:
                            image_ocr_content += f"### 图片 {img_index}\n"
                            image_ocr_content += f"![图片{img_index}]({img_url})\n\n"
                            image_ocr_content += f"**识别文本：**\n{img_text}\n\n"
                
                # 添加作者和发布时间（如果是微信公众号）
                author_info = ""
                if is_wechat_article:
                    author = result_data.get('author', '')
                    publish_time = result_data.get('publish_time', '')
                    if author or publish_time:
                        author_info = "## 文章信息\n"
                        if author:
                            author_info += f"- 作者: {author}\n"
                        if publish_time:
                            author_info += f"- 发布时间: {publish_time}\n"
                        author_info += "\n"
                
                md = f"""{original_content}
{author_info}{image_stats}{image_ocr_content}
## AI分析摘要
# {platform}内容分析
## 内容信息
- 分析时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
- 原始链接: {link}
- 平台: {platform}
- 类型: 图文

{summary}

---
*由视频转文字处理工具自动生成*
"""
            else:
                # 视频格式
                md_path = os.path.join(OUTPUT_DIR, f"{total_count:03d}-{current_date}-{doc_name}_视频分析.md")

                segs = result_data.get('segments', [])
                transcript_lines = []
                for seg in segs:
                    tstr = time.strftime('%H:%M:%S', time.gmtime(seg.get('start_time', 0)))
                    transcript_lines.append(f"- [{tstr}] {seg.get('text','')}")
                transcript = "\n".join(transcript_lines)
                summary = result_data.get('ai_summary', result_data.get('summary', ''))

                md = f"""# {platform}视频内容分析

## 视频信息
- 分析时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
- 原始链接: {link}
- 平台: {platform}
- URL哈希: {url_hash}

## 📝 语音转文字内容
{transcript}

## 🤖 AI智能分析摘要
{summary}

---
*由视频转文字处理工具自动生成*
"""
            with open(md_path, 'w', encoding='utf-8') as f:
                f.write(md)
            return md_path
        except Exception as e:
            self.append_log(f"写入MD失败：{e}")
            return None
    
    # 打开配置窗口
    def open_config_window(self):
        config_window = tk.Toplevel(self.root)
        config_window.title("配置")
        config_window.geometry("600x400")
        config_window.resizable(False, False)
        
        # 配置文件路径
        tk.Label(config_window, text="配置文件：", font=("微软雅黑", 10)).pack(anchor=tk.W, padx=20, pady=(20, 5))
        tk.Label(config_window, text=CONFIG_FILE, font=("Consolas", 9), fg="gray").pack(anchor=tk.W, padx=20, pady=(0, 15))
        
        # 总结提示文本
        tk.Label(config_window, text="总结要求文本：", font=("微软雅黑", 10)).pack(anchor=tk.W, padx=20, pady=(10, 5))
        summary_prompt_var = tk.StringVar(value=CONFIG.get("summary_prompt", DEFAULT_CONFIG["summary_prompt"]))
        summary_prompt_text = scrolledtext.ScrolledText(config_window, height=10, font=("Consolas", 10))
        summary_prompt_text.pack(fill=tk.X, padx=20, pady=(0, 15))
        summary_prompt_text.insert(tk.END, summary_prompt_var.get())
        
        # 保存按钮
        def save_config_changes():
            global CONFIG
            new_summary_prompt = summary_prompt_text.get(1.0, tk.END).strip()
            new_config = CONFIG.copy()
            new_config["summary_prompt"] = new_summary_prompt
            if save_config(new_config):
                CONFIG = new_config
                messagebox.showinfo("成功", "配置已保存")
                config_window.destroy()
            else:
                messagebox.showerror("失败", "保存配置失败")
        
        save_btn = ttk.Button(config_window, text="保存", command=save_config_changes)
        save_btn.pack(pady=20)
        
        # 居中显示
        config_window.transient(self.root)
        config_window.grab_set()
        self.root.wait_window(config_window)
    
    # 打开AI配置窗口 - Prompt配置（原来的配置）
    def open_ai_config_window(self):
        """打开AI Prompt配置窗口"""
        ai_config_window = tk.Toplevel(self.root)
        ai_config_window.title("AI配置 - 视频转文字处理工具")
        ai_config_window.geometry("1000x780")
        ai_config_window.minsize(720, 480)
        ai_config_window.resizable(True, True)
        ai_config_window.configure(bg="#f0f4f8")
        
        # 设置样式
        style = ttk.Style()
        style.theme_use("clam")
        
        # 自定义样式
        style.configure(
            "TButton", 
            padding=(12, 6),
            font=("微软雅黑", 10),
            foreground="#0066cc",
            background="#e6f0ff"
        )
        style.configure(
            "TLabel", 
            font=("微软雅黑", 10),
            foreground="#333"
        )
        style.configure(
            "TEntry", 
            padding=(8, 4),
            font=("微软雅黑", 10)
        )
        style.configure(
            "TLabelframe", 
            font=("微软雅黑", 10, "bold"),
            foreground="#0066cc"
        )
        
        # 顶部工具栏：保存/取消必须在建表单前就占位，否则部分环境下顶栏高度为 0 看不到按钮
        top_bar = tk.Frame(ai_config_window, bg="#dde8f5")
        top_bar.configure(highlightbackground="#0066cc", highlightthickness=1)
        _save_action_holder: list = [lambda: None]

        def _invoke_save_action():
            fn = _save_action_holder[0]
            if callable(fn):
                fn()

        top_inner = tk.Frame(top_bar, bg="#dde8f5")
        top_inner.pack(fill=tk.X, padx=10, pady=8)
        tk.Label(
            top_inner,
            text="修改后请点击「保存配置」",
            font=("微软雅黑", 10, "bold"),
            fg="#0066cc",
            bg="#dde8f5",
        ).pack(side=tk.LEFT)
        ttk.Button(top_inner, text="取消", command=ai_config_window.destroy).pack(side=tk.RIGHT, padx=(6, 0))
        ttk.Button(top_inner, text="保存配置", command=_invoke_save_action).pack(side=tk.RIGHT, padx=(6, 0))
        
        # 底部固定栏（说明 + 保存/取消）
        bottom_bar = tk.Frame(ai_config_window, bg="#e8eef5")
        bottom_bar.configure(highlightbackground="#0066cc", highlightthickness=1)
        
        # 主容器 - 可滚动区域
        main_frame = tk.Frame(ai_config_window, bg="#f0f4f8")
        ai_config_window.grid_columnconfigure(0, weight=1)
        ai_config_window.grid_rowconfigure(1, weight=1)
        ai_config_window.grid_rowconfigure(0, minsize=52)
        top_bar.grid(row=0, column=0, sticky="ew", padx=0, pady=0)
        main_frame.grid(row=1, column=0, sticky="nsew", padx=10, pady=(4, 0))
        bottom_bar.grid(row=2, column=0, sticky="ew")
        
        # 画布和滚动条
        canvas = tk.Canvas(main_frame, bg="#f0f4f8", highlightthickness=0)
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        scrollbar.pack(side="right", fill="y")
        canvas.pack(side="left", fill=tk.BOTH, expand=True)
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # 内容容器
        main_container = tk.Frame(canvas, bg="#f0f4f8")
        _cfg_win_id = canvas.create_window((0, 0), window=main_container, anchor="nw", width=960)
        
        def on_main_container_configure(event):
            canvas.configure(scrollregion=canvas.bbox("all"))
        
        main_container.bind("<Configure>", on_main_container_configure)
        
        def on_canvas_configure(event):
            try:
                sw = scrollbar.winfo_width() or 20
                w = max(int(event.width) - sw - 4, 240)
                canvas.itemconfigure(_cfg_win_id, width=w)
            except tk.TclError:
                pass
            canvas.configure(scrollregion=canvas.bbox("all"))
        
        canvas.bind("<Configure>", on_canvas_configure)
        
        # 添加鼠标滚轮支持
        def on_mouse_wheel(event):
            canvas.yview_scroll(-1 * (event.delta // 120), "units")
        
        canvas.bind_all("<MouseWheel>", on_mouse_wheel)
        
        # 顶部标题区域
        title_frame = tk.Frame(main_container, bg="#f0f4f8")
        title_frame.pack(fill=tk.X, pady=(0, 20))
        
        title_label = tk.Label(
            title_frame, 
            text="AI配置中心", 
            font=("微软雅黑", 16, "bold"),
            foreground="#0066cc",
            bg="#f0f4f8"
        )
        title_label.pack(anchor=tk.W)
        
        subtitle_label = tk.Label(
            title_frame, 
            text="智能分析参数配置", 
            font=("微软雅黑", 10, "italic"),
            foreground="#666",
            bg="#f0f4f8"
        )
        subtitle_label.pack(anchor=tk.W, pady=(5, 0))
        
        # 配置说明区域
        desc_frame = tk.Frame(main_container, bg="#ffffff", bd=0, relief=tk.RAISED)
        desc_frame.pack(fill=tk.X, pady=(0, 20))
        desc_frame.configure(bg="#ffffff", highlightbackground="#0066cc", highlightthickness=1, borderwidth=0)
        
        desc_title = tk.Label(
            desc_frame, 
            text="配置说明", 
            font=("微软雅黑", 12, "bold"),
            foreground="#0066cc",
            bg="#ffffff"
        )
        desc_title.pack(anchor=tk.W, padx=15, pady=(15, 10))
        
        desc_content = tk.Label(
            desc_frame, 
            text="本配置中心主要用于调整视频转文字后的AI文本分析参数，不影响语音转文字过程。\n\n" +
            "**配置作用范围：**\n" +
            "- 仅对AI文本分析步骤生效\n" +
            "- 不影响语音转文字的速度和准确率\n" +
            "- 不影响视频下载过程\n\n" +
            "**使用示例：**\n" +
            "1. 调整System Prompt来改变AI的分析角度\n" +
            "2. 修改Rules来控制分析的重点和范围\n" +
            "3. 自定义Output Template来改变最终文档的格式\n" +
            "4. 调整线程数来优化多任务处理速度",
            font=("微软雅黑", 10),
            foreground="#333",
            bg="#ffffff",
            justify=tk.LEFT,
            wraplength=950
        )
        desc_content.pack(anchor=tk.W, padx=15, pady=(0, 15))
        
        # 语音转文字说明
        speech_desc_frame = tk.Frame(main_container, bg="#ffffff", bd=0, relief=tk.RAISED)
        speech_desc_frame.pack(fill=tk.X, pady=(0, 20))
        speech_desc_frame.configure(bg="#ffffff", highlightbackground="#009966", highlightthickness=1, borderwidth=0)
        
        speech_title = tk.Label(
            speech_desc_frame, 
            text="语音转文字说明", 
            font=("微软雅黑", 12, "bold"),
            foreground="#009966",
            bg="#ffffff"
        )
        speech_title.pack(anchor=tk.W, padx=15, pady=(15, 10))
        
        speech_content = tk.Label(
            speech_desc_frame, 
            text="**语音转文字配置已固定为：**\n" +
            "- 语言：简体中文\n" +
            "- 模型：Whisper tiny（平衡速度和准确率）\n" +
            "- 优化方向：速度优先，同时保证基本准确率\n\n" +
            "**说明：**\n" +
            "- 语音转文字过程无需用户配置\n" +
            "- 系统会自动处理视频中的音频并转为文字\n" +
            "- 转文字结果将作为AI分析的输入",
            font=("微软雅黑", 10),
            foreground="#333",
            bg="#ffffff",
            justify=tk.LEFT,
            wraplength=950
        )
        speech_content.pack(anchor=tk.W, padx=15, pady=(0, 15))
        
        # 配置文件路径
        config_path_frame = tk.Frame(main_container, bg="#ffffff", bd=0, relief=tk.RAISED)
        config_path_frame.pack(fill=tk.X, pady=(0, 15))
        config_path_frame.configure(bg="#ffffff", highlightbackground="#0066cc", highlightthickness=1, borderwidth=0)
        
        ttk.Label(config_path_frame, text="配置文件：", font=("微软雅黑", 10), background="#ffffff").pack(anchor=tk.W, padx=15, pady=(10, 5))
        ttk.Label(
            config_path_frame, 
            text=CONFIG_FILE, 
            font=("Consolas", 9, "italic"), 
            foreground="#666",
            background="#ffffff"
        ).pack(anchor=tk.W, padx=15, pady=(0, 10))
        
        # System Prompt 配置
        system_prompt_frame = tk.Frame(main_container, bg="#ffffff", bd=0, relief=tk.RAISED)
        system_prompt_frame.pack(fill=tk.X, pady=(0, 15))
        system_prompt_frame.configure(bg="#ffffff", highlightbackground="#0066cc", highlightthickness=1, borderwidth=0)
        
        system_title_frame = tk.Frame(system_prompt_frame, bg="#ffffff")
        system_title_frame.pack(fill=tk.X, padx=15, pady=(15, 10))
        
        system_title = tk.Label(
            system_title_frame, 
            text="System Prompt", 
            font=("微软雅黑", 12, "bold"),
            foreground="#0066cc",
            bg="#ffffff"
        )
        system_title.pack(side=tk.LEFT)
        
        system_desc = tk.Label(
            system_title_frame, 
            text="定义AI助手的角色和能力", 
            font=("微软雅黑", 9),
            foreground="#666",
            bg="#ffffff"
        )
        system_desc.pack(side=tk.LEFT, padx=10)
        
        system_prompt_text = scrolledtext.ScrolledText(
            system_prompt_frame, 
            height=5, 
            font=("Consolas", 10),
            bd=0, 
            bg="#f9f9f9",
            relief=tk.FLAT
        )
        system_prompt_text.pack(fill=tk.X, padx=15, pady=(0, 15))
        system_prompt_text.insert(tk.END, CONFIG.get("system_prompt", DEFAULT_CONFIG["system_prompt"]))
        
        # Rules 配置
        rules_frame = tk.Frame(main_container, bg="#ffffff", bd=0, relief=tk.RAISED)
        rules_frame.pack(fill=tk.X, pady=(0, 15))
        rules_frame.configure(bg="#ffffff", highlightbackground="#0066cc", highlightthickness=1, borderwidth=0)
        
        rules_title_frame = tk.Frame(rules_frame, bg="#ffffff")
        rules_title_frame.pack(fill=tk.X, padx=15, pady=(15, 10))
        
        rules_title = tk.Label(
            rules_title_frame, 
            text="Rules 配置", 
            font=("微软雅黑", 12, "bold"),
            foreground="#0066cc",
            bg="#ffffff"
        )
        rules_title.pack(side=tk.LEFT)
        
        rules_desc = tk.Label(
            rules_title_frame, 
            text="设定AI分析的规则和标准", 
            font=("微软雅黑", 9),
            foreground="#666",
            bg="#ffffff"
        )
        rules_desc.pack(side=tk.LEFT, padx=10)
        
        rules_text = scrolledtext.ScrolledText(
            rules_frame, 
            height=5, 
            font=("Consolas", 10),
            bd=0, 
            bg="#f9f9f9",
            relief=tk.FLAT
        )
        rules_text.pack(fill=tk.X, padx=15, pady=(0, 15))
        rules_text.insert(tk.END, CONFIG.get("rules", DEFAULT_CONFIG["rules"]))
        
        # 产出模板配置
        output_template_frame = tk.Frame(main_container, bg="#ffffff", bd=0, relief=tk.RAISED)
        output_template_frame.pack(fill=tk.X, pady=(0, 15))
        output_template_frame.configure(bg="#ffffff", highlightbackground="#0066cc", highlightthickness=1, borderwidth=0)
        
        template_title_frame = tk.Frame(output_template_frame, bg="#ffffff")
        template_title_frame.pack(fill=tk.X, padx=15, pady=(15, 10))
        
        template_title = tk.Label(
            template_title_frame, 
            text="产出模板配置", 
            font=("微软雅黑", 12, "bold"),
            foreground="#0066cc",
            bg="#ffffff"
        )
        template_title.pack(side=tk.LEFT)
        
        template_desc = tk.Label(
            template_title_frame, 
            text="定义最终生成文档的格式和结构", 
            font=("微软雅黑", 9),
            foreground="#666",
            bg="#ffffff"
        )
        template_desc.pack(side=tk.LEFT, padx=10)
        
        output_template_text = scrolledtext.ScrolledText(
            output_template_frame, 
            height=8, 
            font=("Consolas", 10),
            bd=0, 
            bg="#f9f9f9",
            relief=tk.FLAT
        )
        output_template_text.pack(fill=tk.X, padx=15, pady=(0, 15))
        output_template_text.insert(tk.END, CONFIG.get("output_template", DEFAULT_CONFIG["output_template"]))
        
        # 飞书：应用凭证与默认目录（开关与单次路径覆盖在主界面）
        feishu_cfg_frame = tk.Frame(main_container, bg="#ffffff", bd=0, relief=tk.RAISED)
        feishu_cfg_frame.pack(fill=tk.X, pady=(0, 15))
        feishu_cfg_frame.configure(bg="#ffffff", highlightbackground="#0066cc", highlightthickness=1, borderwidth=0)
        
        feishu_head = tk.Frame(feishu_cfg_frame, bg="#ffffff")
        feishu_head.pack(fill=tk.X, padx=15, pady=(15, 8))
        tk.Label(
            feishu_head,
            text="飞书同步",
            font=("微软雅黑", 12, "bold"),
            fg="#0066cc",
            bg="#ffffff",
        ).pack(side=tk.LEFT)
        tk.Label(
            feishu_head,
            text="填写后用于上传 MD；是否实际上传由主界面「同步到飞书」勾选及任务级设置决定",
            font=("微软雅黑", 9),
            fg="#666",
            bg="#ffffff",
        ).pack(side=tk.LEFT, padx=10)
        
        feishu_app_id_var = tk.StringVar(value=CONFIG.get("feishu_app_id", "") or "")
        feishu_app_secret_var = tk.StringVar(value=CONFIG.get("feishu_app_secret", "") or "")
        feishu_default_folder_var = tk.StringVar(value=CONFIG.get("feishu_default_folder_path", "") or "")
        feishu_folder_token_var = tk.StringVar(value=CONFIG.get("feishu_folder_token", "") or "")
        
        tk.Label(feishu_cfg_frame, text="飞书应用 App ID：", font=("微软雅黑", 10), bg="#ffffff", fg="#333").pack(anchor=tk.W, padx=15)
        ttk.Entry(feishu_cfg_frame, textvariable=feishu_app_id_var, width=80).pack(fill=tk.X, padx=15, pady=(2, 8))
        tk.Label(feishu_cfg_frame, text="飞书应用 App Secret：", font=("微软雅黑", 10), bg="#ffffff", fg="#333").pack(anchor=tk.W, padx=15)
        ttk.Entry(feishu_cfg_frame, textvariable=feishu_app_secret_var, show="*", width=80).pack(fill=tk.X, padx=15, pady=(2, 8))
        tk.Label(feishu_cfg_frame, text="云空间文件夹 Token 或 URL（fldcn… 或 …/drive/folder/fldcn…，上传必填）：", font=("微软雅黑", 10), bg="#ffffff", fg="#333").pack(anchor=tk.W, padx=15)
        ttk.Entry(feishu_cfg_frame, textvariable=feishu_folder_token_var, width=80).pack(fill=tk.X, padx=15, pady=(2, 8))
        tk.Label(feishu_cfg_frame, text="默认展示用路径（如 就业知识库/就业技术文档集/AI相关，便于识别；API 落点以上一项为准）：", font=("微软雅黑", 10), bg="#ffffff", fg="#333").pack(anchor=tk.W, padx=15)
        ttk.Entry(feishu_cfg_frame, textvariable=feishu_default_folder_var, width=80).pack(fill=tk.X, padx=15, pady=(2, 8))

        feishu_wiki_sync_var = tk.BooleanVar(value=bool(CONFIG.get("feishu_wiki_sync_enabled", False)))
        feishu_wiki_space_name_var = tk.StringVar(value=CONFIG.get("feishu_wiki_space_name", "") or "")
        feishu_wiki_space_id_var = tk.StringVar(value=CONFIG.get("feishu_wiki_space_id", "") or "")
        feishu_wiki_anchor_var = tk.StringVar(value=CONFIG.get("feishu_wiki_anchor_node_token", "") or "")
        feishu_wiki_path_var = tk.StringVar(value=CONFIG.get("feishu_wiki_path_ensure", "") or "")
        tk.Checkbutton(
            feishu_cfg_frame,
            text="导入后将云文档迁入知识库（下方路径；无 API 文件夹时用空白云文档作目录占位）",
            variable=feishu_wiki_sync_var,
            font=("微软雅黑", 10),
            bg="#ffffff",
            activebackground="#ffffff",
            fg="#333",
            highlightthickness=0,
        ).pack(anchor=tk.W, padx=15, pady=(4, 4))
        tk.Label(feishu_cfg_frame, text="知识空间名称（与列表中名称匹配，含即可；可留空若填了下方空间 ID）：", font=("微软雅黑", 10), bg="#ffffff", fg="#333").pack(anchor=tk.W, padx=15)
        ttk.Entry(feishu_cfg_frame, textvariable=feishu_wiki_space_name_var, width=80).pack(fill=tk.X, padx=15, pady=(2, 4))
        tk.Label(feishu_cfg_frame, text="知识空间 space_id（可选，填写则不再按名称查找）：", font=("微软雅黑", 10), bg="#ffffff", fg="#333").pack(anchor=tk.W, padx=15)
        ttk.Entry(feishu_cfg_frame, textvariable=feishu_wiki_space_id_var, width=80).pack(fill=tk.X, padx=15, pady=(2, 4))
        tk.Label(feishu_cfg_frame, text="锚点 wiki 节点 token（可选，…/wiki/ 后一段；填则路径建在该节点之下）：", font=("微软雅黑", 10), bg="#ffffff", fg="#333").pack(anchor=tk.W, padx=15)
        ttk.Entry(feishu_cfg_frame, textvariable=feishu_wiki_anchor_var, width=80).pack(fill=tk.X, padx=15, pady=(2, 4))
        tk.Label(feishu_cfg_frame, text="知识库内路径（用 / 分隔，不存在则自动建空文档页作目录）：", font=("微软雅黑", 10), bg="#ffffff", fg="#333").pack(anchor=tk.W, padx=15)
        ttk.Entry(feishu_cfg_frame, textvariable=feishu_wiki_path_var, width=80).pack(fill=tk.X, padx=15, pady=(2, 12))
        
        # User Prompt 配置
        user_prompt_frame = tk.Frame(main_container, bg="#ffffff", bd=0, relief=tk.RAISED)
        user_prompt_frame.pack(fill=tk.X, pady=(0, 15))
        user_prompt_frame.configure(bg="#ffffff", highlightbackground="#0066cc", highlightthickness=1, borderwidth=0)
        
        user_title_frame = tk.Frame(user_prompt_frame, bg="#ffffff")
        user_title_frame.pack(fill=tk.X, padx=15, pady=(15, 10))
        
        user_title = tk.Label(
            user_title_frame, 
            text="User Prompt", 
            font=("微软雅黑", 12, "bold"),
            foreground="#0066cc",
            bg="#ffffff"
        )
        user_title.pack(side=tk.LEFT)
        
        user_desc = tk.Label(
            user_title_frame, 
            text="每次处理视频时的额外提示信息", 
            font=("微软雅黑", 9),
            foreground="#666",
            bg="#ffffff"
        )
        user_desc.pack(side=tk.LEFT, padx=10)
        
        user_prompt_text = scrolledtext.ScrolledText(
            user_prompt_frame, 
            height=3, 
            font=("Consolas", 10),
            bd=0, 
            bg="#f9f9f9",
            relief=tk.FLAT
        )
        user_prompt_text.pack(fill=tk.X, padx=15, pady=(0, 15))
        user_prompt_text.insert(tk.END, CONFIG.get("user_prompt", DEFAULT_CONFIG["user_prompt"]))
        
        def save_ai_config_changes():
            global CONFIG
            new_system_prompt = system_prompt_text.get(1.0, tk.END).strip()
            new_rules = rules_text.get(1.0, tk.END).strip()
            new_output_template = output_template_text.get(1.0, tk.END).strip()
            new_user_prompt = user_prompt_text.get(1.0, tk.END).strip()
            
            new_config = CONFIG.copy()
            new_config["system_prompt"] = new_system_prompt
            new_config["rules"] = new_rules
            new_config["output_template"] = new_output_template
            new_config["user_prompt"] = new_user_prompt
            new_config["feishu_app_id"] = feishu_app_id_var.get().strip()
            new_config["feishu_app_secret"] = feishu_app_secret_var.get().strip()
            new_config["feishu_folder_token"] = feishu_folder_token_var.get().strip()
            new_config["feishu_default_folder_path"] = feishu_default_folder_var.get().strip()
            new_config["feishu_wiki_sync_enabled"] = bool(feishu_wiki_sync_var.get())
            new_config["feishu_wiki_space_name"] = feishu_wiki_space_name_var.get().strip()
            new_config["feishu_wiki_space_id"] = feishu_wiki_space_id_var.get().strip()
            new_config["feishu_wiki_anchor_node_token"] = feishu_wiki_anchor_var.get().strip()
            new_config["feishu_wiki_path_ensure"] = feishu_wiki_path_var.get().strip()
            
            if save_config(new_config):
                CONFIG = new_config
                messagebox.showinfo("成功", "AI配置已保存（config.json；已连接 MariaDB 时同步 video_agent_config）")
                ai_config_window.destroy()
            else:
                messagebox.showerror("失败", "保存AI配置失败")

        _save_action_holder[0] = save_ai_config_changes
        
        tk.Label(
            bottom_bar,
            text=(
                "保存：写入 config.json，并在 db 模块可用时同步到 MariaDB 表 video_agent_config（与项目 init_database / ai_api_config 一致）；"
                "启动时库中有记录则覆盖 JSON 同名字段。飞书走开放平台 API。"
            ),
            font=("微软雅黑", 9),
            fg="#444",
            bg="#e8eef5",
            justify=tk.LEFT,
            wraplength=620,
        ).pack(side=tk.LEFT, padx=12, pady=10, anchor=tk.W)
        btn_wrap = tk.Frame(bottom_bar, bg="#e8eef5")
        btn_wrap.pack(side=tk.RIGHT, padx=12, pady=8)
        ttk.Button(btn_wrap, text="取消", command=ai_config_window.destroy).pack(side=tk.RIGHT, padx=(8, 0))
        ttk.Button(btn_wrap, text="保存配置", command=_invoke_save_action).pack(side=tk.RIGHT, padx=(8, 0))
        
        ai_config_window.update_idletasks()
        
        # 居中显示
        ai_config_window.transient(self.root)
        ai_config_window.grab_set()
        self.root.wait_window(ai_config_window)
    
    # 打开AI API配置窗口（新的API配置界面）
    def _apply_ai_api_runtime_config(self, main_config, backup_configs):
        """将 API 窗口中的主/备配置写回 config.json，与 summarize/chat 使用同一数据源。"""
        global CONFIG
        CONFIG = {**CONFIG}
        if main_config.get("api_key"):
            CONFIG["volcengine_api_key"] = main_config["api_key"]
        if main_config.get("endpoint_id"):
            CONFIG["ai_chat_model"] = main_config["endpoint_id"]
        if main_config.get("base_url"):
            CONFIG["volcengine_base_url"] = main_config["base_url"]
        if main_config.get("model"):
            CONFIG["ai_chat_model_display_name"] = main_config["model"]
        if backup_configs:
            ep = (backup_configs[0] or {}).get("endpoint_id") or ""
            if ep:
                CONFIG["ai_chat_model_backup"] = ep
        save_config(CONFIG)
        self.append_log("已同步 AI API 到 config.json（主/备接入点与密钥）", "INFO")

    def open_ai_api_config_window(self):
        """打开AI API配置窗口（API Key、Model等）"""
        if AI_API_CONFIG_AVAILABLE:
            open_ai_api_config_window(
                self.root,
                get_runtime_config=lambda: CONFIG.copy(),
                on_save_runtime=self._apply_ai_api_runtime_config,
            )
        else:
            messagebox.showwarning(
                "模块未加载",
                "AI API配置模块(ai_api_config_gui)未找到，请确保文件存在。\n\n" +
                "请检查文件: ai_api_config_gui.py"
            )
    
    def open_thread_config_window(self):
        """打开线程配置窗口"""
        thread_config_window = tk.Toplevel(self.root)
        thread_config_window.title("线程配置 - 视频转文字处理工具")
        thread_config_window.geometry("800x500")
        thread_config_window.resizable(True, True)
        thread_config_window.configure(bg="#f0f4f8")
        
        # 设置样式
        style = ttk.Style()
        style.theme_use("clam")
        
        # 自定义样式
        style.configure(
            "TButton", 
            padding=(12, 6),
            font=("微软雅黑", 10),
            foreground="#0066cc",
            background="#e6f0ff"
        )
        style.configure(
            "TLabel", 
            font=("微软雅黑", 10),
            foreground="#333"
        )
        style.configure(
            "TEntry", 
            padding=(8, 4),
            font=("微软雅黑", 10)
        )
        style.configure(
            "TLabelframe", 
            font=("微软雅黑", 10, "bold"),
            foreground="#0066cc"
        )
        
        # 主容器
        main_container = tk.Frame(thread_config_window, bg="#f0f4f8")
        main_container.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # 顶部标题区域
        title_frame = tk.Frame(main_container, bg="#f0f4f8")
        title_frame.pack(fill=tk.X, pady=(0, 20))
        
        title_label = tk.Label(
            title_frame, 
            text="线程配置中心", 
            font=("微软雅黑", 16, "bold"),
            foreground="#0066cc",
            bg="#f0f4f8"
        )
        title_label.pack(anchor=tk.W)
        
        subtitle_label = tk.Label(
            title_frame, 
            text="调整线程池大小和任务队列限制", 
            font=("微软雅黑", 10, "italic"),
            foreground="#666",
            bg="#f0f4f8"
        )
        subtitle_label.pack(anchor=tk.W, pady=(5, 0))
        
        # 系统信息区域
        system_info_frame = tk.Frame(main_container, bg="#ffffff", bd=0, relief=tk.RAISED)
        system_info_frame.pack(fill=tk.X, pady=(0, 20))
        system_info_frame.configure(bg="#ffffff", highlightbackground="#0066cc", highlightthickness=1, borderwidth=0)
        
        ttk.Label(system_info_frame, text="系统信息：", font=("微软雅黑", 10, "bold"), background="#ffffff").pack(anchor=tk.W, padx=15, pady=(10, 5))
        
        info_text = (
            f"系统CPU核心数：{self.cpu_count}\n"
            f"当前执行中任务数：{len(self.active_futures)}\n"
            f"待处理队列长度：{self._task_queue_len()}\n"
            f"队列最大大小：{self.queue_max_size}"
        )
        info_label = tk.Label(
            system_info_frame, 
            text=info_text, 
            font=("Consolas", 9), 
            foreground="#666",
            background="#ffffff",
            justify=tk.LEFT
        )
        info_label.pack(anchor=tk.W, padx=15, pady=(0, 10))
        
        # 线程池配置区域
        thread_pool_frame = tk.Frame(main_container, bg="#ffffff", bd=0, relief=tk.RAISED)
        thread_pool_frame.pack(fill=tk.X, pady=(0, 20))
        thread_pool_frame.configure(bg="#ffffff", highlightbackground="#0066cc", highlightthickness=1, borderwidth=0)
        
        thread_pool_title_frame = tk.Frame(thread_pool_frame, bg="#ffffff")
        thread_pool_title_frame.pack(fill=tk.X, padx=15, pady=(15, 10))
        
        thread_pool_title = tk.Label(
            thread_pool_title_frame, 
            text="线程池配置", 
            font=("微软雅黑", 12, "bold"),
            foreground="#0066cc",
            bg="#ffffff"
        )
        thread_pool_title.pack(side=tk.LEFT)
        
        thread_pool_desc = tk.Label(
            thread_pool_title_frame, 
            text=f"范围：1-{self.cpu_count}", 
            font=("微软雅黑", 9),
            foreground="#666",
            bg="#ffffff"
        )
        thread_pool_desc.pack(side=tk.LEFT, padx=10)
        
        # 线程数输入
        thread_count_frame = tk.Frame(thread_pool_frame, bg="#ffffff")
        thread_count_frame.pack(fill=tk.X, padx=15, pady=(0, 15))
        
        ttk.Label(thread_count_frame, text="最大线程数：", font=("微软雅黑", 10), background="#ffffff").pack(side=tk.LEFT, padx=(0, 10))
        
        thread_count_var = tk.StringVar(value=str(self.max_workers))
        thread_count_entry = ttk.Entry(
            thread_count_frame, 
            textvariable=thread_count_var, 
            width=10, 
            font=("微软雅黑", 10)
        )
        thread_count_entry.pack(side=tk.LEFT, padx=(0, 10))
        
        # 任务队列配置区域
        queue_frame = tk.Frame(main_container, bg="#ffffff", bd=0, relief=tk.RAISED)
        queue_frame.pack(fill=tk.X, pady=(0, 20))
        queue_frame.configure(bg="#ffffff", highlightbackground="#0066cc", highlightthickness=1, borderwidth=0)
        
        queue_title_frame = tk.Frame(queue_frame, bg="#ffffff")
        queue_title_frame.pack(fill=tk.X, padx=15, pady=(15, 10))
        
        queue_title = tk.Label(
            queue_title_frame, 
            text="任务队列配置", 
            font=("微软雅黑", 12, "bold"),
            foreground="#0066cc",
            bg="#ffffff"
        )
        queue_title.pack(side=tk.LEFT)
        
        # 队列大小输入
        queue_size_frame = tk.Frame(queue_frame, bg="#ffffff")
        queue_size_frame.pack(fill=tk.X, padx=15, pady=(0, 15))
        
        ttk.Label(queue_size_frame, text="队列最大大小：", font=("微软雅黑", 10), background="#ffffff").pack(side=tk.LEFT, padx=(0, 10))
        
        queue_size_var = tk.StringVar(value=str(self.queue_max_size))
        queue_size_entry = ttk.Entry(
            queue_size_frame, 
            textvariable=queue_size_var, 
            width=10, 
            font=("微软雅黑", 10)
        )
        queue_size_entry.pack(side=tk.LEFT, padx=(0, 10))
        
        # 按钮框架
        button_frame = tk.Frame(main_container, bg="#f0f4f8")
        button_frame.pack(fill=tk.X, pady=20)
        
        # 应用按钮
        def apply_thread_config_changes():
            try:
                # 验证并更新线程数
                new_thread_count = int(thread_count_var.get())
                if new_thread_count < 1 or new_thread_count > self.cpu_count:
                    messagebox.showerror("错误", f"线程数必须在1-{self.cpu_count}之间")
                    return
                
                # 验证并更新队列大小
                new_queue_size = int(queue_size_var.get())
                if new_queue_size < 1:
                    messagebox.showerror("错误", "队列大小必须大于0")
                    return
                
                # 应用线程池调整
                if new_thread_count != self.max_workers:
                    self.update_thread_pool_size(new_thread_count)
                
                # 应用队列大小调整
                if new_queue_size != self.queue_max_size:
                    self.queue_max_size = new_queue_size
                    self.append_log(f"任务队列最大大小已调整为：{new_queue_size}")
                
                messagebox.showinfo("成功", "线程配置已应用")
                thread_config_window.destroy()
            except ValueError:
                messagebox.showerror("错误", "请输入有效的数字")
        
        apply_btn = ttk.Button(button_frame, text="应用配置", command=apply_thread_config_changes)
        apply_btn.pack(side=tk.RIGHT, padx=10)
        
        cancel_btn = ttk.Button(button_frame, text="取消", command=thread_config_window.destroy)
        cancel_btn.pack(side=tk.RIGHT, padx=10)
        
        # 居中显示
        thread_config_window.transient(self.root)
        thread_config_window.grab_set()
        self.root.wait_window(thread_config_window)
    
    def update_thread_pool_size(self, new_size):
        """更新线程池大小
        
        Args:
            new_size: 新的线程池大小
        """
        try:
            self.append_log(f"开始调整线程池大小：从 {self.max_workers} 到 {new_size}")
            
            if new_size > self.max_workers:
                # 增加线程数
                self.max_workers = new_size
                # 重新创建线程池
                self.executor.shutdown(wait=False)
                self.executor = concurrent.futures.ThreadPoolExecutor(max_workers=self.max_workers)
                self.append_log(f"线程池已扩容至：{self.max_workers} 线程")
                
                # 扩容后，如果队列中有任务且未在处理中，继续处理
                if self._task_queue_len() > 0 and not self.processing_queue:
                    self.append_log("线程池扩容完成，继续处理队列任务")
                    self.start_queue_processing()
            elif new_size < self.max_workers:
                # 减少线程数（active_futures 为 dict: link -> future）
                pairs = list(self.active_futures.items())
                keep = min(new_size, len(pairs))
                cancelled_count = 0
                for link, future in pairs[keep:]:
                    if not future.done():
                        future.cancel()
                        cancelled_count += 1
                        self.append_log(f"已尝试取消任务：{str(link)[:70]}...")
                    self.active_futures.pop(link, None)
                
                # 更新线程池大小
                self.max_workers = new_size
                # 重新创建线程池
                self.executor.shutdown(wait=False)
                self.executor = concurrent.futures.ThreadPoolExecutor(max_workers=self.max_workers)
                self.append_log(f"线程池已缩容至：{self.max_workers} 线程")
                self.append_log(f"已取消 {cancelled_count} 个任务")
                
                # 缩容后，无论是否正在处理中，都检查队列并继续处理
                # 因为即使正在处理中，我们也需要确保新的线程池能够处理剩余任务
                if self._task_queue_len() > 0:
                    # 如果当前没有在处理中，直接开始处理
                    if not self.processing_queue:
                        self.append_log("线程池缩容完成，继续处理队列任务")
                        self.start_queue_processing()
                    else:
                        # 如果当前正在处理中，记录日志，等待当前批次完成后自动继续
                        self.append_log("线程池缩容完成，当前有任务正在处理中，将在完成后继续处理队列任务")
            
            # 保存线程数量到配置文件
            global CONFIG
            new_config = CONFIG.copy()
            new_config["max_workers"] = self.max_workers
            save_config(new_config)
            CONFIG = new_config
            self.append_log(f"线程数量已保存到配置文件：{self.max_workers}")
            
            # 更新队列状态显示
            self.update_queue_status()
        except Exception as e:
            self.append_log(f"调整线程池大小失败：{e}")
            messagebox.showerror("错误", f"调整线程池大小失败：{e}")


def main():
    root = tk.Tk()
    app = App(root)
    root.mainloop()

if __name__ == '__main__':
    main()
